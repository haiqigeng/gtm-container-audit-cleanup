from __future__ import annotations

import copy
import json
import re
import subprocess
import sys
import tempfile
import tomllib
import unittest
import urllib.error
from pathlib import Path
from unittest.mock import MagicMock, patch

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS))

from build_skill_package import build as build_skill_bundle  # noqa: E402
from check_release import (  # noqa: E402
    check_project_version,
    check_release_tag,
    check_repository_layout,
    git_ls_files,
)
from gtm_architecture_review import (  # noqa: E402
    scaffold_review as scaffold_architecture,
)
from gtm_architecture_review import (  # noqa: E402
    validate_review as validate_architecture,
)
from gtm_audit_gate_check import validate_workbook  # noqa: E402
from gtm_audit_package_build import build_package  # noqa: E402
from gtm_baseline_audit import audit_export  # noqa: E402
from gtm_change_log_build import build_change_log  # noqa: E402
from gtm_configuration_review import (  # noqa: E402
    scaffold_review as scaffold_configuration,
)
from gtm_configuration_review import (  # noqa: E402
    validate_review as validate_configuration,
)
from gtm_consent_model import tag_consent_route  # noqa: E402
from gtm_context_model import build_context_model  # noqa: E402
from gtm_custom_code_extract import extract_export  # noqa: E402
from gtm_diff_operations import operations as diff_operations  # noqa: E402
from gtm_future_state_check import apply_operations, check_future_state  # noqa: E402
from gtm_human_rows import build_rows  # noqa: E402
from gtm_lib import container_version  # noqa: E402
from gtm_operation_compile import compile_operations, source_object_catalog  # noqa: E402
from gtm_operational_review import (  # noqa: E402
    MANDATORY_OPERATIONAL_MODULES,
    mandatory_module_errors,
)
from gtm_operational_review import (  # noqa: E402
    scaffold_review as scaffold_operational,
)
from gtm_operational_review import (  # noqa: E402
    validate_review as validate_operational,
)
from gtm_privacy import (  # noqa: E402
    privacy_findings,
    redact_text,
    sanitize_url,
    spreadsheet_safe_text,
)
from gtm_privacy_scan import scan_xlsx  # noqa: E402
from gtm_relationships import (  # noqa: E402
    object_records,
    relationship_candidates,
)
from gtm_relationships import (  # noqa: E402
    scan_export as scan_relationships,
)
from gtm_review_common import (  # noqa: E402
    object_consumer_map,
    object_keys,
    object_name_map,
    object_source_path_map,
    validate_operation_set,
    validate_structured_actions,
)
from gtm_review_shards import merge_review, split_review  # noqa: E402
from gtm_shared_facts import build_shared_facts  # noqa: E402
from gtm_source_model import build_model  # noqa: E402
from gtm_three_run_gate import run_gate  # noqa: E402
from gtm_validate_artifact import missing_references  # noqa: E402
from gtm_validate_artifact import validate as validate_artifact  # noqa: E402
from gtm_vendor_registry import (  # noqa: E402
    load_registry,
    official_url_error,
    validate_registry,
    vendor_record,
)
from gtm_workbook_build import (  # noqa: E402
    CANONICAL_SHEETS,
    MAX_CELL_TEXT,
    add_table,
    build_workbook,
)


def condition(operator: str, left: str, right: str) -> dict:
    return {
        "type": operator,
        "parameter": [
            {"type": "TEMPLATE", "key": "arg0", "value": left},
            {"type": "TEMPLATE", "key": "arg1", "value": right},
        ],
    }


def sample_export() -> dict:
    return {
        "exportFormatVersion": 2,
        "containerVersion": {
            "accountId": "100",
            "containerId": "200",
            "containerVersionId": "1",
            "container": {"publicId": "GTM-TEST", "usageContext": ["WEB"]},
            "tag": [
                {
                    "tagId": "1",
                    "name": "GA4 - Purchase - All",
                    "type": "gaawe",
                    "parameter": [
                        {"type": "TEMPLATE", "key": "eventName", "value": "purchase"},
                        {"type": "TEMPLATE", "key": "measurementId", "value": "G-TEST123"},
                        {
                            "type": "MAP",
                            "key": "eventParameters",
                            "map": [
                                {
                                    "type": "TEMPLATE",
                                    "key": "value",
                                    "value": "{{DLV - Value}}",
                                },
                                {
                                    "type": "TEMPLATE",
                                    "key": "transaction_id",
                                    "value": "{{DLV - Transaction ID}}",
                                }
                            ],
                        },
                    ],
                    "firingTriggerId": ["10"],
                    "blockingTriggerId": ["13"],
                    "setupTag": [{"tagName": "Utility - Consent Defaults"}],
                },
                {
                    "tagId": "2",
                    "name": "Meta - Purchase - All",
                    "type": "html",
                    "parameter": [
                        {
                            "type": "TEMPLATE",
                            "key": "html",
                            "value": (
                                "<script>\n"
                                "var items = {{DLV - Items}} || [];\n"
                                "fbq('track', 'Purchase', {contents: items});\n"
                                "</script>"
                            ),
                        }
                    ],
                    "firingTriggerId": ["12"],
                },
                {
                    "tagId": "3",
                    "name": "Utility - Consent Defaults",
                    "type": "html",
                    "parameter": [
                        {
                            "type": "TEMPLATE",
                            "key": "html",
                            "value": "<script>\nwindow.consentDefault = 'denied';\n</script>",
                        }
                    ],
                    "parentFolderId": "101",
                },
                {
                    "tagId": "4",
                    "name": "Paused - Helper Consumer",
                    "type": "html",
                    "paused": True,
                    "parameter": [
                        {
                            "type": "TEMPLATE",
                            "key": "html",
                            "value": "<script>void({{CJS - Paused Only}});</script>",
                        }
                    ],
                    "firingTriggerId": ["10"],
                },
            ],
            "trigger": [
                {
                    "triggerId": "10",
                    "name": "Purchase",
                    "type": "CUSTOM_EVENT",
                    "customEventFilter": [condition("EQUALS", "{{_event}}", "purchase")],
                },
                {
                    "triggerId": "11",
                    "name": "Purchase copy",
                    "type": "CUSTOM_EVENT",
                    "customEventFilter": [condition("EQUALS", "{{_event}}", "purchase")],
                },
                {
                    "triggerId": "12",
                    "name": "TG - Purchase only",
                    "type": "TRIGGER_GROUP",
                    "parameter": [
                        {
                            "type": "LIST",
                            "key": "triggerIds",
                            "list": [{"type": "TEMPLATE", "value": "10"}],
                        }
                    ],
                },
                {
                    "triggerId": "13",
                    "name": "Block - Page view",
                    "type": "CUSTOM_EVENT",
                    "customEventFilter": [condition("EQUALS", "{{_event}}", "page_view")],
                },
                {
                    "triggerId": "14",
                    "name": "Click - Invalid regex",
                    "type": "LINK_CLICK",
                    "filter": [condition("MATCH_REGEX", "{{Click URL}}", "(")],
                },
                {
                    "triggerId": "15",
                    "name": "Funnel question 1",
                    "type": "CUSTOM_EVENT",
                    "customEventFilter": [condition("EQUALS", "{{_event}}", "funnel_question_1")],
                },
                {
                    "triggerId": "16",
                    "name": "Funnel step impression Q1",
                    "type": "CUSTOM_EVENT",
                    "customEventFilter": [
                        condition("EQUALS", "{{_event}}", "funnel_step_impression")
                    ],
                },
            ],
            "variable": [
                {
                    "variableId": "20",
                    "name": "DLV - Items",
                    "type": "v",
                    "parameter": [
                        {"type": "INTEGER", "key": "dataLayerVersion", "value": "2"},
                        {"type": "TEMPLATE", "key": "name", "value": "ecommerce.items"},
                    ],
                },
                {
                    "variableId": "21",
                    "name": "DLV - Items copy",
                    "type": "v",
                    "parameter": [
                        {"type": "INTEGER", "key": "dataLayerVersion", "value": "2"},
                        {"type": "TEMPLATE", "key": "name", "value": "ecommerce.items"},
                    ],
                },
                {
                    "variableId": "22",
                    "name": "CJS - Page URL Mirror",
                    "type": "jsm",
                    "parameter": [
                        {
                            "type": "TEMPLATE",
                            "key": "javascript",
                            "value": "function() {\n  return {{Page URL}};\n}",
                        }
                    ],
                },
                {
                    "variableId": "23",
                    "name": "CJS - Paused Only",
                    "type": "jsm",
                    "parameter": [
                        {
                            "type": "TEMPLATE",
                            "key": "javascript",
                            "value": "function() {\n  return 'paused-value';\n}",
                        }
                    ],
                },
                {
                    "variableId": "24",
                    "name": "DLV - Value",
                    "type": "v",
                    "parameter": [
                        {"type": "INTEGER", "key": "dataLayerVersion", "value": "2"},
                        {"type": "TEMPLATE", "key": "name", "value": "ecommerce.value"},
                    ],
                },
                {
                    "variableId": "25",
                    "name": "DLV - Transaction ID",
                    "type": "v",
                    "parameter": [
                        {"type": "INTEGER", "key": "dataLayerVersion", "value": "2"},
                        {
                            "type": "TEMPLATE",
                            "key": "name",
                            "value": "ecommerce.transaction_id",
                        },
                    ],
                },
            ],
            "folder": [
                {"folderId": "100", "name": "Unused folder"},
                {"folderId": "101", "name": "Utilities"},
            ],
            "builtInVariable": [
                {"name": "Page URL", "type": "PAGE_URL"},
                {"name": "Click URL", "type": "CLICK_URL"},
            ],
        },
    }


def fixed_slot_formula_export() -> dict:
    data = sample_export()
    variables = data["containerVersion"]["variable"]
    for variable_id, index in (("30", 1), ("31", 2), ("32", 3)):
        variables.append(
            {
                "variableId": variable_id,
                "name": f"DLV - Product Price {index}",
                "type": "v",
                "parameter": [
                    {"type": "INTEGER", "key": "dataLayerVersion", "value": "2"},
                    {
                        "type": "TEMPLATE",
                        "key": "name",
                        "value": f"ecommerce.product_price_{index}",
                    },
                ],
            }
        )
    variables.append(
        {
            "variableId": "33",
            "name": "CJS - Total Price",
            "type": "jsm",
            "parameter": [
                {
                    "type": "TEMPLATE",
                    "key": "javascript",
                    "value": (
                        "function() {\n"
                        "  return Number({{DLV - Product Price 1}} || 0) + "
                        "Number({{DLV - Product Price 2}} || 0) + "
                        "Number({{DLV - Product Price 3}} || 0);\n"
                        "}"
                    ),
                }
            ],
        }
    )
    event_parameters = data["containerVersion"]["tag"][0]["parameter"][2]["map"]
    event_parameters[0]["value"] = "{{CJS - Total Price}}"
    return data


def object_specific_text(row: dict, subject: str, field: str = "") -> str:
    field_terms = [
        str(value)
        for value in (row.get("field_evidence_requirements") or {}).get(field, [])[:3]
    ]
    if not field_terms:
        field_terms = [row["object_name"] or row["object_key"], row["object_type"]]
    field_terms.extend(
        str(value)
        for value in row.get("specificity_tokens", [])[:2]
        if str(value) not in field_terms
    )
    evidence = " and ".join(field_terms)
    return (
        f"{row['object_name'] or row['object_key']} is a {row['object_type']} that {subject}; "
        f"the exported configuration specifically names {evidence}."
    )


def concrete_purpose_subject(row: dict) -> str:
    return {
        "tag": "sends, loads, routes, or records the configured measurement action",
        "trigger": "matches and activates the configured event or condition scope",
        "variable": "reads, calculates, maps, or returns its configured value",
        "zone": "restricts, scopes, allows, and governs the configured child container",
        "customTemplate": "defines and executes the exported template behavior",
        "client": "claims, parses, and routes the configured request",
        "gtagConfig": "configures, routes, sets, and governs the Google tag behavior",
        "transformation": "transforms, allows, or redacts the configured fields",
    }.get(row.get("layer"), "implements its concrete exported action")


def branch_role(path: str) -> str:
    lowered = path.lower()
    if "consent" in lowered or "storage" in lowered:
        return "Consent"
    if "filter" in lowered or "condition" in lowered or "operator" in lowered:
        return "Condition"
    if any(value in lowered for value in ("firingtriggerid", "blockingtriggerid", "triggerids")):
        return "Routing"
    if any(
        value in lowered
        for value in (
            "setuptag",
            "teardowntag",
            "tagfiringoption",
            "schedulestartms",
            "scheduleendms",
        )
    ):
        return "Execution control"
    if "childcontainer" in lowered or "typerestriction" in lowered:
        return "Condition"
    return "Input"


def contract_topic_anchors(
    topic: dict, fact_by_path: dict, row: dict, contract_anchor: str
) -> list[str]:
    events = [str(value).lower() for value in topic.get("configured_event_values", [])]
    anchors = [
        path
        for path, fact in fact_by_path.items()
        if path in row["required_logic_anchors"]
        and any(
            event in str(fact.get("value_preview") or "").lower() for event in events
        )
    ]
    return anchors or [contract_anchor]


def behavior_signal_text(fact: dict) -> str:
    signals = []
    for signal in fact.get("required_behavior_signals", []):
        required_terms = [
            str(group[0])
            for group in signal.get("required_term_groups", [])
            if isinstance(group, list) and group
        ]
        signals.append(
            f"{str(signal.get('signal') or 'source behavior').replace('_', ' ')} "
            f"uses {' '.join(required_terms)}"
        )
    return "; ".join(signals)


def complete_configuration(export_path: Path) -> dict:
    review = scaffold_configuration(export_path)
    for row in review["rows"]:
        row.update(
            {
                "review_status": "complete",
                "purpose": object_specific_text(
                    row, concrete_purpose_subject(row), "purpose"
                ),
                "execution_logic": object_specific_text(
                    row, "runs under the named trigger, condition, or call route", "execution_logic"
                ),
                "inputs_and_terminal_sources": object_specific_text(
                    row,
                    "reads the listed GTM references and terminal configuration",
                    "inputs_and_terminal_sources",
                ),
                "configured_output_or_side_effect": object_specific_text(
                    row,
                    "produces the named event, return value, or browser side effect",
                    "configured_output_or_side_effect",
                ),
                "consumer_contract": object_specific_text(
                    row,
                    "supplies the listed consumer objects with that configured value",
                    "consumer_contract",
                ),
                "consent_and_sequence": object_specific_text(
                    row,
                    "uses the listed blocking, consent, and sequencing controls",
                    "consent_and_sequence",
                ),
                "correctness_verdict": "Correct",
                "correctness_basis": object_specific_text(
                    row,
                    "has matching inputs, route, output, and consumers in this source",
                    "correctness_basis",
                ),
                "defects": [],
                "contract_checks": [],
                "code_behavior_blocks": [],
                "technical_facts_assessment": "",
                "technical_finding_reviews": [],
                "configuration_branch_reviews": [
                    {
                        "json_path": branch["json_path"],
                        "value_hash": branch["value_hash"],
                        "logic_role": branch_role(branch["json_path"]),
                        "interpretation": (
                            f"At {branch['json_path']}, value {branch.get('value_preview')} is the "
                            f"{branch_role(branch['json_path']).lower()} branch for "
                            f"{row['object_name'] or row['object_key']}."
                        ),
                        "configured_effect": (
                            f"The {branch_role(branch['json_path']).lower()} setting "
                            f"{branch.get('value_preview')} at {branch['json_path']} is read, "
                            "matched, routed, or applied when that configured branch executes for "
                            f"{row['object_name'] or row['object_key']}."
                        ),
                        "correctness": "Correct",
                    }
                    for branch in row["required_branch_reviews"]
                ],
                "evidence_anchors": list(row["required_logic_anchors"]),
                "consumer_evidence_keys": [
                    item["consumer_key"] for item in row["export_consumers"]
                ],
                "reference_traces": [
                    {
                        "reference": item["reference"],
                        "object_chain": item["required_object_keys"],
                        "evidence_anchors": item["required_evidence_anchors"],
                        "terminal_states": item["terminal_states"],
                        "terminal_source": (
                            f"Reference {item['reference']} terminates in the source states "
                            f"{', '.join(item['terminal_states'])} after the listed variable chain."
                        ),
                        "node_reviews": [
                            {
                                "object_key": node["object_key"],
                                "object_name": node["object_name"],
                                "object_type": node["object_type"],
                                "config_hash": node["config_hash"],
                                "source_json_path": node["source_json_path"],
                                "referenced_variables": node["referenced_variables"],
                                "configured_parameters": node["configured_parameters"],
                                "semantic_role": node["semantic_role"],
                                "evidence_anchors": node["required_evidence_anchors"],
                                "configured_function": (
                                    f"{node['object_name']} ({node['object_type']}) reads "
                                    f"{' and '.join(node['specificity_tokens'][:3])} from its parameters."
                                ),
                                "configured_output": (
                                    f"{node['object_name']} ({node['object_type']}) returns the value "
                                    f"selected by {' and '.join(node['specificity_tokens'][:3])}."
                                ),
                                "output_type_and_shape": (
                                    f"{node['object_name']} keeps the {node['object_type']} output "
                                    f"shape associated with {' and '.join(node['specificity_tokens'][:2])}."
                                ),
                                "availability_and_fallback": (
                                    f"{node['object_name']} makes {' and '.join(node['specificity_tokens'][:2])} "
                                    "available only where its listed source exists, with no extra fallback."
                                ),
                                "consumer_compatibility": (
                                    f"{node['object_name']} supplies its {node['object_type']} and "
                                    f"{' and '.join(node['specificity_tokens'][:2])} value to "
                                    f"{row['object_name'] or row['object_key']}."
                                ),
                            }
                            for node in item["required_nodes"]
                        ],
                        "edge_reviews": [
                            {
                                **edge,
                                "dependency_meaning": (
                                    f"{edge['from_object_key']} reads {edge['reference']} from "
                                    f"{edge['to_object_key']} before returning its configured result."
                                ),
                            }
                            for edge in item["required_edges"]
                        ],
                        "terminal_reviews": [
                            {
                                **terminal,
                                "terminal_meaning": (
                                    f"{terminal['reference']} resolves to {terminal['configured_source']} "
                                    f"as the final {terminal['state']} source."
                                ),
                                "consumer_compatibility": (
                                    f"The {terminal['reference']} {terminal['state']} source supplies "
                                    f"the configured value to {row['object_name'] or row['object_key']}."
                                ),
                            }
                            for terminal in item["terminal_requirements"]
                        ],
                    }
                    for item in row["reference_trace_requirements"]
                ],
                "related_operational_finding_ids": [],
                "logic_cross_checks": [
                    {
                        "check_key": check["check_key"],
                        "verdict": "Aligned",
                        "conclusion": (
                            f"For {row['object_name'] or row['object_key']}, "
                            f"{check['question']} The exported facts "
                            f"{' and '.join(check['required_terms'][:2])} remain aligned in this "
                            "controlled fixture configuration."
                        ),
                        "evidence_anchors": list(check["allowed_evidence_anchors"][:2]),
                    }
                    for check in row["required_logic_cross_checks"]
                ],
                "disposition": "keep",
                "owner_question": "",
                "operation": {},
                "confidence": "High",
                "evidence_citations": {
                    field: list((row.get("field_evidence_paths") or {}).get(field, []))[
                        : 2 if field == "correctness_basis" else 1
                    ]
                    for field in (
                        "purpose",
                        "execution_logic",
                        "inputs_and_terminal_sources",
                        "configured_output_or_side_effect",
                        "consumer_contract",
                        "consent_and_sequence",
                        "correctness_basis",
                    )
                },
            }
        )
        if row["required_contract_topics"]:
            fact_by_path = {
                fact["json_path"]: fact for fact in row.get("source_facts", [])
            }
            contract_anchor = row["required_logic_anchors"][0]

            row["contract_checks"] = [
                {
                    "contract_topic": topic["topic_key"],
                    "contract_field": (f"{topic['vendor']} {topic['topic']} exported contract"),
                    "configured_value": (
                        f"At {contract_topic_anchors(topic, fact_by_path, row, contract_anchor)[0]}, "
                        "the exported value "
                        f"{fact_by_path[contract_topic_anchors(topic, fact_by_path, row, contract_anchor)[0]].get('value_preview')} configures "
                        f"{topic['topic']} for {row['object_name']}; vendor-specific events are "
                        f"{', '.join(topic.get('configured_event_values', [])) or 'not exported'}"
                    ),
                    "expected_rule": (
                        f"The official {topic['vendor']} documentation defines the required "
                        f"{topic['topic']} behavior and value types"
                    ),
                    "source": (
                        topic["official_doc_candidates"][0]
                        if topic["official_doc_candidates"]
                        else ""
                    ),
                    "identified_vendor": topic["vendor"],
                    "official_source_basis": (
                        f"The cited page is the registered official {topic['vendor']} reference, "
                        "or no authoritative vendor identity and source is visible for this "
                        "unclassified integration."
                    ),
                    "research_status": (
                        "No authoritative vendor identity or official documentation source can "
                        "be established from the exported integration hostname and code alone."
                        if topic.get("research_required")
                        and not topic["official_doc_candidates"]
                        else "Official source is registered for this detected vendor."
                    ),
                    "verdict": (
                        "Non-compliant"
                        if topic.get("deterministic_contract_state")
                        == "known_noncompliant"
                        else "Unproven"
                        if topic.get("deterministic_contract_state")
                        == "unproven_from_container"
                        or topic.get("research_required")
                        and not topic["official_doc_candidates"]
                        else "Compliant"
                    ),
                    "evidence_anchors": contract_topic_anchors(
                        topic, fact_by_path, row, contract_anchor
                    ),
                }
                for topic in row["required_contract_topics"]
            ]
        if row["required_code_line_hashes"]:
            previews = " ".join(
                str(item.get("line_preview") or "") for item in row["code_line_facts"]
            )
            markers = [
                token
                for token in re.findall(r"[A-Za-z_$][A-Za-z0-9_.$:/-]{3,}", previews)
                if token.lower() not in {"function", "return", "const", "false", "true", "script"}
            ]
            marker = markers[0] if markers else row["object_name"]
            segment_marker_text = " and ".join(dict.fromkeys(markers)) or marker
            required_behavior_text = "; ".join(
                text
                for fact in row["code_line_facts"]
                if (text := behavior_signal_text(fact))
            )
            row["code_behavior_blocks"] = [
                {
                    "line_hashes": list(row["required_code_line_hashes"]),
                    "start_line": min(item["line_number"] for item in row["code_line_facts"]),
                    "end_line": max(item["line_number"] for item in row["code_line_facts"]),
                    "purpose": f"The {marker} block implements the exact exported helper behavior.",
                    "inputs": f"The {marker} block reads only the variables and literals visible here.",
                    "outputs": f"The {marker} block returns or sends the output shown by these lines.",
                    "side_effects": f"The {marker} block has the browser effects identified in static facts.",
                    "health_assessment": (
                        f"The {marker} implementation is coherent for this controlled fixture; "
                        f"every source segment is identified by {segment_marker_text}. "
                        f"Source-visible behavior: {required_behavior_text or 'no additional static signal'}."
                    ),
                }
            ]
            row["technical_facts_assessment"] = (
                f"The {marker} code facts, parser result, side effects, and line behavior are "
                "accounted for in this container-only assessment."
            )
            fallback_segment_reviews = []
            for fact in row["code_line_facts"]:
                fact_markers = [
                    token
                    for token in re.findall(
                        r"[A-Za-z_$][A-Za-z0-9_.$:/-]{3,}",
                        str(fact.get("line_preview") or ""),
                    )
                    if token.lower()
                    not in {
                        "function",
                        "return",
                        "const",
                        "false",
                        "true",
                        "undefined",
                        "script",
                    }
                ][:4]
                fact_marker_text = (
                    " and ".join(fact_markers) or "the exported syntax boundary"
                )
                fact_behavior_text = behavior_signal_text(fact)
                fallback_segment_reviews.append(
                    {
                        "line_hash": fact["line_hash"],
                        "behavior": (
                            "Mandatory line-by-line review of this segment identifies "
                            f"{fact_marker_text} and maps its inputs, output, side effects, "
                            "and execution behavior. Source-visible behavior: "
                            f"{fact_behavior_text or 'no additional static signal'}."
                        ),
                    }
                )
            row["technical_finding_reviews"] = [
                {
                    "finding_key": item["finding_key"],
                    "source_statement": item["statement"],
                    "verdict": (
                        "Documented exception"
                        if item["category"] in {"parser", "health", "security"}
                        else "False positive"
                    ),
                    "rationale": (
                        (
                            f"The {marker} mandatory line-by-line code blocks cover every "
                            "exported line; the parser boundary is recorded without claiming "
                            "AST coverage."
                        )
                        if item["category"] == "parser"
                        else (
                            f"The {marker} controlled fixture documents the exact source signal "
                            f"{item['statement']} with identifiers {segment_marker_text}; its "
                            "exported behavior remains explicit and accepted for this test."
                        )
                    ),
                    "proposed_action": "",
                    "exception_basis": (
                        f"This controlled fixture accepts the source-proven risk "
                        f"{item['statement']} because its behavior is required and retained "
                        "under an explicit test constraint."
                        if item["category"] in {"health", "security"}
                        else ""
                    ),
                    "owner_question": "",
                    "fallback_line_hashes": (
                        list(row["required_code_line_hashes"])
                        if item["category"] == "parser"
                        else []
                    ),
                    "parser_boundary": (
                        f"The exact parser status "
                        f"{row['technical_code_facts'].get('javascript_parser')} leaves syntax "
                        "coverage incomplete for GTM substitutions or template wrapper code."
                        if item["category"] == "parser"
                        else ""
                    ),
                    "manual_review_method": (
                        f"A mandatory line-by-line review follows source identifiers "
                        f"{' and '.join(markers[:4]) or marker}, maps every segment hash to its "
                        "inputs, output, side effects, and execution branch, and claims no AST proof."
                        if item["category"] == "parser"
                        else ""
                    ),
                    "fallback_segment_reviews": (
                        copy.deepcopy(fallback_segment_reviews)
                        if item["category"] == "parser"
                        else []
                    ),
                }
                for item in row["required_technical_findings"]
            ]
        issue_obligations = [
            item
            for item in row["required_configuration_obligations"]
            if item["required_outcome"] == "Issue"
        ]
        unclear_obligations = [
            item
            for item in row["required_configuration_obligations"]
            if item["required_outcome"] == "Unclear"
        ]
        required_topic_by_key = {
            item["topic_key"]: item for item in row["required_contract_topics"]
        }
        issue_contract_topics = {
            topic
            for item in issue_obligations
            for topic in item.get("affected_contract_topics", [])
        }
        unclear_contract_topics = {
            topic
            for item in unclear_obligations
            for topic in item.get("affected_contract_topics", [])
        }
        for check in row["contract_checks"]:
            topic = required_topic_by_key[check["contract_topic"]]["topic"]
            if topic in issue_contract_topics:
                check["verdict"] = "Non-compliant"
            elif topic in unclear_contract_topics and check["verdict"] == "Compliant":
                check["verdict"] = "Unproven"
        issue_paths = {
            anchor
            for item in issue_obligations
            for anchor in item["evidence_anchors"]
        }
        unclear_paths = {
            anchor
            for item in unclear_obligations
            for anchor in item["evidence_anchors"]
        }
        noncompliant_checks = [
            check for check in row["contract_checks"] if check["verdict"] == "Non-compliant"
        ]
        unproven_checks = [
            check for check in row["contract_checks"] if check["verdict"] == "Unproven"
        ]
        issue_paths.update(
            anchor
            for check in noncompliant_checks
            for anchor in check["evidence_anchors"]
        )
        unclear_paths.update(
            anchor
            for check in unproven_checks
            for anchor in check["evidence_anchors"]
            if anchor not in issue_paths
        )
        for branch in row["configuration_branch_reviews"]:
            if branch["json_path"] in issue_paths:
                branch["correctness"] = "Issue"
            elif branch["json_path"] in unclear_paths:
                branch["correctness"] = "Unclear"
        issue_check_keys = {
            key
            for item in issue_obligations
            for key in item["affected_logic_checks"]
        }
        unclear_check_keys = {
            key
            for item in unclear_obligations
            for key in item["affected_logic_checks"]
        }
        if noncompliant_checks:
            issue_check_keys.add("vendor_contract_alignment")
        elif unproven_checks:
            unclear_check_keys.add("vendor_contract_alignment")
        for check in row["logic_cross_checks"]:
            requirement = next(
                item
                for item in row["required_logic_cross_checks"]
                if item["check_key"] == check["check_key"]
            )
            if check["check_key"] in issue_check_keys:
                check["verdict"] = "Issue"
                issue_evidence = issue_paths | {
                    anchor
                    for contract in noncompliant_checks
                    for anchor in contract["evidence_anchors"]
                }
                allowed = set(requirement["allowed_evidence_anchors"])
                check["evidence_anchors"] = sorted(issue_evidence & allowed)[:2]
                check["conclusion"] = (
                    f"For {row['object_name'] or row['object_key']}, the exported facts "
                    f"{' and '.join(requirement['required_terms'][:2])} "
                    "contain a deterministic configuration defect and are not aligned."
                )
            elif check["check_key"] in unclear_check_keys:
                check["verdict"] = "Unclear"
                check["conclusion"] = (
                    f"For {row['object_name'] or row['object_key']}, the exported facts "
                    f"{' and '.join(requirement['required_terms'][:2])} "
                    "leave a specific container-only contract unproven."
                )
            relevant_obligations = [
                item
                for item in row["required_configuration_obligations"]
                if check["check_key"] in item["affected_logic_checks"]
            ]
            if relevant_obligations:
                check["conclusion"] += " " + " ".join(
                    f"Obligation {item['obligation_key']}: {item['statement']}"
                    for item in relevant_obligations
                )
        row["defects"] = [
            {
                "defect_id": f"AUTO-{index:03d}",
                "statement": item["statement"],
                "configured_effect": (
                    f"The exported state for {item['obligation_key']} makes the configured "
                    "execution, payload, or dependency behavior invalid."
                ),
                "expected_behavior": (
                    "The source must expose a valid, resolvable, officially supported "
                    "configuration before this object can be certified."
                ),
                "evidence_anchors": list(item["evidence_anchors"]),
                "code_line_hashes": [],
                "technical_finding_keys": [],
            }
            for index, item in enumerate(issue_obligations, start=1)
        ]
        defect_index = len(row["defects"])
        for check in noncompliant_checks:
            defect_index += 1
            row["defects"].append(
                {
                    "defect_id": f"AUTO-{defect_index:03d}",
                    "statement": (
                        f"Official contract topic {check['contract_topic']} is non-compliant."
                    ),
                    "configured_effect": (
                        f"The exported value in {check['configured_value']} does not satisfy "
                        "the cited vendor contract."
                    ),
                    "expected_behavior": (
                        f"Use the officially supported names, required fields, and value types "
                        f"for {check['contract_topic']}."
                    ),
                    "evidence_anchors": list(check["evidence_anchors"]),
                    "code_line_hashes": [],
                    "technical_finding_keys": [],
                }
            )
        unresolved = bool(unclear_obligations or unproven_checks)
        if row["defects"]:
            row["correctness_verdict"] = "Issue"
            row["correctness_basis"] = object_specific_text(
                row,
                "contains the listed deterministic defects and cannot be certified as correct",
                "correctness_basis",
            )
            row["disposition"] = "owner_decision_needed"
            row["owner_question"] = (
                "Should the listed source-proven defects be corrected through an approved "
                "operation, or is there a documented owner constraint requiring a redesign?"
            )
            row["confidence"] = "High"
        elif unresolved:
            row["correctness_verdict"] = "Owner decision needed"
            row["correctness_basis"] = object_specific_text(
                row,
                "leaves the listed runtime, external, or owner-controlled contract unproven",
                "correctness_basis",
            )
            row["disposition"] = "owner_decision_needed"
            row["owner_question"] = (
                "What approved runtime, vendor, or ownership evidence resolves the explicitly "
                "unproven configuration contract for this object?"
            )
            row["confidence"] = "Medium"
        for technical_review in row["technical_finding_reviews"]:
            if "no reviewable executable behavior" in technical_review[
                "source_statement"
            ].lower():
                technical_review["verdict"] = "Owner decision needed"
                technical_review["rationale"] = (
                    "The exported template metadata does not expose executable behavior, so an "
                    "owner must provide the original template source or approve removal."
                )
                technical_review["owner_question"] = (
                    "Which owner can provide the custom-template executable source and permissions, "
                    "or approve removal of the opaque implementation?"
                )
    review["run_status"] = "complete"
    return review


def complete_operational(export_path: Path) -> dict:
    review = scaffold_operational(export_path)
    for row in review["findings"]:
        row.update(
            {
                "review_status": "complete",
                "disposition": "owner_decision_needed",
                "rationale": (
                    f"The source evidence {' and '.join(row['rationale_evidence_terms'][:2])} "
                    "requires an explicit retained-versus-cleaned decision before this controlled "
                    "fixture can claim a completed cleanup."
                ),
                "owner_question": (
                    f"Should the source objects for {row['finding_type']} be retained for a "
                    "documented business reason, or approved for the proposed cleanup?"
                ),
            }
        )
    review["run_status"] = "complete"
    return review


def member_assessments(
    item: dict,
    keys_field: str,
    anchors_field: str,
    paused_field: str,
    terms_field: str,
) -> list:
    distinguishing_map = (
        item.get("member_distinguishing_terms")
        or item.get("candidate_distinguishing_terms")
        or {}
    )
    return [
        {
            "object_key": key,
            "configured_role": (
                f"{key} uses {' and '.join(item[terms_field][key][:2])} to perform its "
                "specific role in this measurement chain."
            ),
            "necessity": (
                f"{key} remains necessary while {item[terms_field][key][0]} has a distinct "
                "consumer, route, or terminal source."
            ),
            "distinguishing_configuration": (
                f"{key} is distinguished by "
                f"{' and '.join((distinguishing_map.get(key) or item[terms_field][key])[:2])} "
                "in its route, payload, or dependency configuration; its own source facts include "
                f"{' and '.join(item[terms_field][key][:2])}."
            ),
            "status": "paused" if item[paused_field].get(key, False) else "active",
            "evidence_anchors": item[anchors_field][key][:1],
        }
        for key in item[keys_field]
    ]


def architecture_text(row: dict, field: str, statement: str) -> str:
    terms = list((row.get("field_evidence_requirements") or {}).get(field, [])[:3])
    for token in row.get("chain_specificity_tokens", [])[:3]:
        if token not in terms:
            terms.append(token)
    return f"{statement}; the source specifically includes {' and '.join(terms)}."


def architecture_caution_text(cautions: list[dict]) -> str:
    keys = {str(item.get("caution_key") or "") for item in cautions}
    statements = []
    if "deduplication_alignment_unproven" in keys:
        statements.append(
            "Runtime deduplication through event ID or transaction ID remains unproven from "
            "the visible container"
        )
    if "consent_alignment_unproven_or_conflicting" in keys:
        statements.append(
            "end-to-end browser and server consent alignment remains unproven and unresolved"
        )
    return (" " + "; ".join(statements) + ".") if statements else ""


def unsafe_owner_question(comparison_types: set[str], candidate_keys: list[str]) -> str:
    identities = " and ".join(candidate_keys[:2])
    prefix = f"For {identities}, "
    if "browser_server_consent_deduplication_review" in comparison_types:
        return prefix + (
            "Which owner approves the browser and server routes, and what evidence resolves "
            "their consent forwarding and event-ID deduplication contract?"
        )
    if "shared_zone_child_container" in comparison_types:
        return prefix + (
            "Which Zone owns the shared child container, and what non-overlapping boundary "
            "scope justifies retaining both Zone routes?"
        )
    if "cyclic_trigger_group_dependency" in comparison_types:
        return prefix + (
            "Which trigger-group dependency should be removed to break the exported cycle, "
            "and which trigger route is canonical?"
        )
    return prefix + (
        "Which tag route is canonical, and what trigger or consent distinction justifies "
        "retaining the same payload on the other route?"
    )


def complete_architecture(export_path: Path) -> dict:
    review = scaffold_architecture(export_path)
    non_retention_types = {
        "same_tag_payload_different_route",
        "shared_zone_child_container",
        "cyclic_trigger_group_dependency",
        "browser_server_consent_deduplication_review",
    }
    for row in review["families"]:
        basis = row["family_label"] or row["family_key"]
        member_keys = set(row["member_object_keys"])
        unsafe_family_comparisons = [
            candidate
            for candidate in review["comparisons"]
            if set(candidate["candidate_object_keys"]) <= member_keys
            and set(candidate["comparison_types"]) & non_retention_types
        ]
        unresolved_member_relationship = bool(unsafe_family_comparisons)
        family_comparison_types = {
            comparison_type
            for candidate in unsafe_family_comparisons
            for comparison_type in candidate["comparison_types"]
            if comparison_type in non_retention_types
        }
        family_cautions = [
            caution
            for candidate in unsafe_family_comparisons
            for caution in candidate.get("required_caution_states", [])
        ]
        row.update(
            {
                "review_status": "complete",
                "business_action": architecture_text(
                    row, "business_action", f"The {basis} family records one named business action"
                ),
                "family_purpose": architecture_text(
                    row, "family_purpose", f"The {basis} family serves one destination outcome"
                ),
                "member_assessments": member_assessments(
                    row,
                    "member_object_keys",
                    "available_member_evidence_anchors",
                    "member_paused_status",
                    "member_evidence_terms",
                ),
                "chain_assessments": member_assessments(
                    row,
                    "chain_object_keys",
                    "available_chain_evidence_anchors",
                    "chain_paused_status",
                    "chain_evidence_terms",
                ),
                "execution_path_summary": architecture_text(
                    row, "execution_path_summary", f"The {basis} route connects its tag and dependencies"
                ),
                "payload_coherence": architecture_text(
                    row, "payload_coherence", f"The {basis} payload matches its event and destination"
                ),
                "consent_and_sequence_coherence": architecture_text(
                    row,
                    "consent_and_sequence_coherence",
                    f"The {basis} route uses its listed consent and sequence controls",
                ),
                "necessity_and_ownership": architecture_text(
                    row,
                    "necessity_and_ownership",
                    f"The {basis} members retain distinct chain ownership",
                ),
                "relationship_verdict": (
                    "Owner decision needed"
                    if unresolved_member_relationship
                    else "Complementary"
                ),
                "analyst_rationale": architecture_text(
                    row,
                    "analyst_rationale",
                    f"The {basis} members have no proven duplicate firing in this export",
                )
                + architecture_caution_text(family_cautions),
                "target_architecture": architecture_text(
                    row,
                    "target_architecture",
                    f"Keep the {basis} chain minimal while preserving these distinct roles",
                ),
                "disposition": (
                    "owner_decision_needed" if unresolved_member_relationship else "keep"
                ),
                "owner_question": (
                    unsafe_owner_question(
                        family_comparison_types, row["member_object_keys"]
                    )
                    if unresolved_member_relationship
                    else ""
                ),
                "operations": [],
                "confidence": "High",
            }
        )
    for row in review["comparisons"]:
        exact = "exact_configuration" in row.get("comparison_types", [])
        owner_required = exact or bool(
            set(row.get("comparison_types", [])) & non_retention_types
        )
        comparison_types = set(row.get("comparison_types", []))
        caution_text = architecture_caution_text(
            row.get("required_caution_states", [])
        )
        row.update(
            {
                "review_status": "complete",
                "member_assessments": member_assessments(
                    row,
                    "candidate_object_keys",
                    "available_member_evidence_anchors",
                    "candidate_paused_status",
                    "candidate_evidence_terms",
                ),
                "relationship_verdict": (
                    "Owner decision needed" if owner_required else "Intentional variant"
                ),
                "analyst_rationale": architecture_text(
                    row,
                    "analyst_rationale",
                    f"{row['comparison_id']} candidates retain distinct roles after route and source comparison",
                )
                + caution_text,
                "architecture_effect": architecture_text(
                    row,
                    "architecture_effect",
                    f"{row['comparison_id']} keeps separate paths because no common target is proven",
                )
                + caution_text,
                "disposition": "owner_decision_needed" if owner_required else "keep",
                "owner_question": (
                    unsafe_owner_question(
                        comparison_types, row["candidate_object_keys"]
                    )
                    if owner_required
                    else ""
                ),
                "operations": [],
                "confidence": "High",
            }
        )
    source_records = object_records(container_version(json.loads(export_path.read_text(encoding="utf-8"))))
    source_record_by_key = {
        record["object_key"]: record
        for layer_records in source_records.values()
        for record in layer_records
    }
    reviewed_keys = sorted(
        record["object_key"]
        for layer_records in source_records.values()
        for record in layer_records
    )
    attestation = review["open_discovery_attestation"]
    source_terms = [
        term
        for family in review["families"]
        for terms in family["chain_evidence_terms"].values()
        for term in terms
    ]
    source_terms = list(dict.fromkeys([*source_terms, *reviewed_keys]))
    while len(source_terms) < 3:
        source_terms.append(f"source-object-{len(source_terms) + 1}")

    def method_source_terms(method_row: dict) -> list[str]:
        keys = method_row["candidate_object_keys"] or reviewed_keys
        values = [
            value
            for key in keys
            for value in (
                source_record_by_key[key]["object_name"],
                source_record_by_key[key]["object_key"],
            )
        ]
        return list(dict.fromkeys(values))[:6]
    attestation.update(
        {
            "review_status": "complete",
            "reviewed_object_keys": reviewed_keys,
            "discovered_comparison_ids": [],
            "zero_discovery_rationale": (
                f"Every object, including {source_terms[0]}, {source_terms[1]}, and "
                f"{source_terms[2]}, was compared by "
                + ", ".join(
                    row["method"].replace("_", " ")
                    for row in review["discovery_method_coverage"]
                )
                + " without finding another source-grounded relationship."
            ),
            "method_reviews": [
                {
                    **row,
                    "review_status": "complete",
                    "reviewed_comparison_ids": list(row["comparison_ids"]),
                    "reviewed_object_keys": reviewed_keys,
                    "additional_discovery_ids": [],
                    "conclusion": (
                        f"The {row['method'].replace('_', ' ')} scan reviewed "
                        f"{len(reviewed_keys)} source objects including "
                        f"{' and '.join(method_source_terms(row))}, and candidates "
                        f"{', '.join(row['comparison_ids']) or 'with no generated comparison'}; "
                        "no additional source-grounded relationship was found in this fixture."
                    ),
                }
                for row in review["discovery_method_coverage"]
            ],
        }
    )
    review["run_status"] = "complete"
    return review


def value_discovery_row(export_path: Path) -> dict:
    records = object_records(
        container_version(json.loads(export_path.read_text(encoding="utf-8")))
    )
    by_key = {
        record["object_key"]: record
        for layer_records in records.values()
        for record in layer_records
    }
    keys = ["variable:22", "variable:24"]
    return {
        "comparison_id": "DISC-VALUE-001",
        "comparison_origin": "analyst_discovered",
        "discovery_methods": ["terminal_source_formula_and_output_overlap"],
        "comparison_types": ["shared_terminal_source"],
        "candidate_object_keys": keys,
        "candidate_basis": [
            "Recursive chain review found two differently named value helpers that require "
            "consumer-level comparison outside deterministic similarity groups."
        ],
        "review_status": "complete",
        "member_assessments": [
            {
                "object_key": key,
                "configured_role": (
                    f"{by_key[key]['object_name']} ({by_key[key]['object_type']}) returns "
                    f"{' and '.join(by_key[key]['specificity_tokens'][:2])} to its consumers."
                ),
                "necessity": (
                    f"{by_key[key]['object_name']} remains necessary while "
                    f"{by_key[key]['specificity_tokens'][0]} serves a distinct consumer."
                ),
                "distinguishing_configuration": (
                    f"{by_key[key]['object_name']} differs as a {by_key[key]['object_type']} "
                    f"through {' and '.join(by_key[key]['specificity_tokens'][:2])}."
                ),
                "status": "paused" if by_key[key]["paused"] else "active",
                "evidence_anchors": by_key[key]["evidence_anchors"][:1],
            }
            for key in keys
        ],
        "relationship_verdict": "Intentional variant",
        "analyst_rationale": (
            "CJS - Page URL Mirror mirrors Page URL while DLV - Value reads "
            "ecommerce.value, so shared value wording does not establish duplication."
        ),
        "architecture_effect": (
            "Keep CJS - Page URL Mirror and DLV - Value as separate variable roles while "
            "Page URL and ecommerce.value consumers remain distinct."
        ),
        "disposition": "keep",
        "owner_question": "",
        "operations": [],
        "confidence": "High",
    }


def duplicate_variable_operation() -> dict:
    return {
        "operation_key": "consolidate-ecommerce-items-dlv",
        "title": "Consolidate duplicate ecommerce items variables",
        "area": "GTM hygiene",
        "problem_type": "Exact duplicate",
        "problem": "Two data-layer variables read the same ecommerce.items source with identical settings.",
        "why_it_matters": "Maintaining both variables creates needless ambiguity and duplicate ownership work.",
        "expected_clean_state": "One canonical ecommerce.items variable serves every existing consumer.",
        "exact_proposed_action": "Keep variable 20 and delete unused duplicate variable 21.",
        "canonical_object_key": "variable:20",
        "changes": [],
        "remaps": [],
        "deletions": [
            {
                "object_key": "variable:21",
                "reason": "Variable 21 duplicates variable 20 and has no active consumer.",
            }
        ],
        "renames": [],
        "preconditions": "Confirm variable 21 still has no export-visible consumer before execution.",
        "qa_steps": "Re-export the workspace and confirm every ecommerce.items reference resolves to variable 20.",
        "rollback": "Restore variable 21 from the original container export if a missed dependency appears.",
        "priority": "Medium",
        "confidence": "High",
        "execution_readiness": "approval_required",
        "minimum_aggressiveness": "Standard",
        "challenge_review": {},
    }


def align_duplicate_operation(operational: dict, architecture: dict) -> None:
    operation = duplicate_variable_operation()
    for finding in operational["findings"]:
        if finding.get("module_name") not in {
            "duplicate_variable_logic",
            "duplicate_variable_paths",
        }:
            continue
        if set(finding.get("object_ids", [])) != {"20", "21"}:
            continue
        finding.update(copy.deepcopy(operation))
        finding["disposition"] = "cleanup_operation"
        finding["rationale"] = (
            "Variables 20 and 21 have identical data-layer settings, and variable 21 has no "
            "consumer, so one canonical variable is sufficient."
        )
    for comparison in architecture["comparisons"]:
        if set(comparison.get("candidate_object_keys", [])) != {"variable:20", "variable:21"}:
            continue
        comparison.update(
            {
                "relationship_verdict": "Exact duplicate",
                "analyst_rationale": (
                    "DLV - Items and DLV - Items Copy read ecommerce.items with identical type "
                    "and settings, while only DLV - Items has an active consumer."
                ),
                "architecture_effect": (
                    "Keeping DLV - Items and DLV - Items Copy adds maintenance work because both "
                    "use ecommerce.items without a distinct route, payload, or consumer contract."
                ),
                "disposition": "cleanup_operation",
                "operations": [copy.deepcopy(operation)],
            }
        )


class PipelineTests(unittest.TestCase):
    maxDiff = None

    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        self.addCleanup(self.tempdir.cleanup)
        self.root = Path(self.tempdir.name)
        self.export_path = self.root / "container.json"
        self.export_path.write_text(json.dumps(sample_export()), encoding="utf-8")

    def completed_reviews(self) -> tuple[dict, dict, dict]:
        return (
            complete_operational(self.export_path),
            complete_configuration(self.export_path),
            complete_architecture(self.export_path),
        )

    def write_review(self, name: str, payload: dict) -> Path:
        path = self.root / name
        path.write_text(json.dumps(payload), encoding="utf-8")
        return path

    def test_system_and_builtin_references_are_not_missing(self) -> None:
        report = missing_references(container_version(sample_export()))
        self.assertEqual([], report["undefinedVariableReferences"])
        configuration = scaffold_configuration(self.export_path)
        mirror = next(row for row in configuration["rows"] if row["object_id"] == "22")
        trace = next(
            item
            for item in mirror["reference_trace_requirements"]
            if item["reference"] == "Page URL"
        )
        self.assertEqual(["built_in"], trace["terminal_states"])

    def test_container_only_contract_has_no_d4_or_runtime_gate(self) -> None:
        paths = [ROOT / "SKILL.md", ROOT / "README.md"]
        paths.extend((ROOT / "scripts").glob("*.py"))
        paths.extend((ROOT / "references").rglob("*.md"))
        for path in paths:
            content = path.read_text(encoding="utf-8").lower()
            self.assertNotIn("d4_required", content, str(path))
            self.assertNotIn("runtime_qa_required", content, str(path))
        removed_runtime_reference = (
            ROOT / "references" / "02-commands" / ("runtime-qa-" + "templates.md")
        )
        self.assertFalse(removed_runtime_reference.exists())

    def test_operational_scan_catches_basic_cleanup_failures(self) -> None:
        findings = [
            row
            for row in audit_export(self.export_path)["findings"]
            if row["finding_type"] != "zero_findings"
        ]
        types = {row["finding_type"] for row in findings}
        self.assertIn("duplicate_configuration", types)
        self.assertIn("duplicate_variable_path", types)
        self.assertIn("single_member_trigger_group", types)
        self.assertIn("invalid_trigger_regex", types)
        self.assertIn("ineffective_blocking_trigger", types)
        self.assertIn("variable_mirrors_builtin", types)
        unused_builtins = {
            object_id
            for row in findings
            if row["finding_type"] == "unused_built_in_variable"
            for object_id in row["object_ids"]
        }
        self.assertEqual({"Page URL", "Click URL"}, unused_builtins)
        paused = [row for row in findings if row["module_name"] == "used_only_by_paused_tags"]
        self.assertTrue(any("23" in row["object_ids"] for row in paused))
        triggerless = [
            row for row in findings if row["module_name"] == "tags_without_firing_triggers"
        ]
        self.assertFalse(any("3" in row["object_ids"] for row in triggerless))
        allowed_resolutions = {
            "cleanup_operation",
            "documented_exception",
            "owner_decision_needed",
        }
        self.assertTrue(
            all(
                {
                    value.strip()
                    for value in row["required_resolution"].split("|")
                    if value.strip()
                }
                <= allowed_resolutions
                for row in findings
            )
        )

    def test_fixed_slot_business_formula_cannot_pass_as_generic_false_positive(self) -> None:
        path = self.root / "fixed-slot-formula.json"
        path.write_text(json.dumps(fixed_slot_formula_export()), encoding="utf-8")
        findings = [
            row for row in audit_export(path)["findings"] if row["finding_type"] != "zero_findings"
        ]
        formula = next(
            row for row in findings if row["finding_type"] == "fixed_slot_business_formula"
        )
        self.assertEqual(["33"], formula["object_ids"])
        technical = extract_export(path)
        technical_row = next(row for row in technical["rows"] if row["object_id"] == "33")
        self.assertTrue(technical_row["fixed_slot_aggregation"])
        self.assertEqual([1, 2, 3], technical_row["fixed_slot_groups"][0]["indexes"])

        review = complete_configuration(path)
        total = next(row for row in review["rows"] if row["object_key"] == "variable:33")
        fixed_review = next(
            row
            for row in total["technical_finding_reviews"]
            if "fixed numbered value slots" in row["source_statement"].lower()
        )
        fixed_review["verdict"] = "False positive"
        fixed_review["rationale"] = "The code was inspected and appears acceptable."
        errors, _ = validate_configuration(
            path,
            self.write_review("fixed-slot-generic-dismissal.json", review),
        )
        self.assertTrue(any("fixed-slot business formula" in error for error in errors))

    def test_consent_purposes_with_same_logic_are_compared(self) -> None:
        data = sample_export()
        shared_config = [
            {"type": "TEMPLATE", "key": "name", "value": "OnetrustActiveGroups"},
        ]
        data["containerVersion"]["variable"].extend(
            [
                {
                    "variableId": "40",
                    "name": "CJS - analytics_storage",
                    "type": "jsm",
                    "parameter": [
                        {
                            "type": "TEMPLATE",
                            "key": "javascript",
                            "value": "function(){ return {{OnetrustActiveGroups}}.indexOf(',2,') > -1; }",
                        },
                        *shared_config,
                    ],
                },
                {
                    "variableId": "41",
                    "name": "CJS - ad_storage",
                    "type": "jsm",
                    "parameter": [
                        {
                            "type": "TEMPLATE",
                            "key": "javascript",
                            "value": "function(){ return {{OnetrustActiveGroups}}.indexOf(',2,') > -1; }",
                        },
                        *shared_config,
                    ],
                },
            ]
        )
        path = self.root / "consent-shared-logic.json"
        path.write_text(json.dumps(data), encoding="utf-8")
        findings = audit_export(path)["findings"]
        consent = [
            row
            for row in findings
            if row["finding_type"] == "different_consent_purposes_share_logic"
        ]
        self.assertEqual(1, len(consent))
        self.assertEqual({"40", "41"}, set(consent[0]["object_ids"]))

    def test_server_forwarded_consent_is_distinct_from_missing_client_control(self) -> None:
        variable = {
            "variableId": "50",
            "name": "DLV - Server consent state",
            "type": "v",
            "parameter": [
                {
                    "type": "TEMPLATE",
                    "key": "name",
                    "value": "consent_state",
                }
            ],
        }
        tag = {
            "tagId": "51",
            "name": "Google tag - Server transport",
            "type": "googtag",
            "parameter": [
                {
                    "type": "TEMPLATE",
                    "key": "transport_url",
                    "value": "https://collect.example.test",
                },
                {
                    "type": "TEMPLATE",
                    "key": "server_consent",
                    "value": "{{DLV - Server consent state}}",
                },
            ],
        }
        route = tag_consent_route(
            tag,
            "$.containerVersion.tag[0]",
            variables=[variable],
        )
        self.assertEqual("server_forwarding_candidate", route["effective_control_status"])
        self.assertEqual(["collect.example.test"], route["server_routing_hosts"])
        self.assertEqual(
            ["DLV - Server consent state"],
            route["server_consent_forwarding_variables"],
        )
        self.assertTrue(route["forwarded_cmp_signal_visible"])
        self.assertEqual("not_visible_in_web_export", route["server_enforcement_visibility"])

        data = sample_export()
        data["containerVersion"]["variable"].append(variable)
        media_transport = {
            **tag,
            "tagId": "53",
            "name": "Meta - Lead server transporter",
            "type": "html",
            "firingTriggerId": ["10"],
        }
        data["containerVersion"]["tag"].append(media_transport)
        export = self.root / "server-forwarded-consent.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        findings = audit_export(export)["findings"]
        self.assertFalse(
            any(
                row["finding_type"] == "media_consent_route_requires_review"
                and "53" in row["object_ids"]
                for row in findings
            )
        )

    def test_server_route_without_forwarded_consent_remains_unproven(self) -> None:
        tag = {
            "tagId": "52",
            "name": "Google tag - Incomplete server transport",
            "type": "googtag",
            "parameter": [
                {
                    "type": "TEMPLATE",
                    "key": "transport_url",
                    "value": "https://collect.example.test",
                }
            ],
        }
        route = tag_consent_route(tag)
        self.assertEqual("server_contract_unproven", route["effective_control_status"])
        self.assertEqual([], route["server_consent_forwarding_evidence"])

    def test_manual_consent_enum_only_enables_additional_checks_when_needed(self) -> None:
        expectations = {
            "notSet": ("NOT_SET", False, "native_consent_capability"),
            "notNeeded": ("NOT_NEEDED", False, "native_consent_capability"),
            "needed": ("NEEDED", True, "explicit_export_control"),
            "unexpected": ("UNEXPECTED", False, "unrecognized_consent_status"),
        }
        for raw_status, expected in expectations.items():
            with self.subTest(raw_status=raw_status):
                route = tag_consent_route(
                    {
                        "tagId": "consent",
                        "name": "Google tag consent enum",
                        "type": "googtag",
                        "consentSettings": {"consentStatus": raw_status},
                    }
                )
                self.assertEqual(expected[0], route["consent_status"])
                self.assertEqual(expected[1], route["additional_consent_checks_visible"])
                self.assertEqual(expected[2], route["effective_control_status"])

    def test_consent_control_requires_payload_and_preserves_every_vendor(self) -> None:
        route = tag_consent_route(
            {
                "tagId": "consent-boundary",
                "name": "Consent bootstrap",
                "type": "html",
                "parameter": [
                    {
                        "type": "TEMPLATE",
                        "key": "eventName",
                        "value": "consent_ready",
                    },
                    {
                        "type": "TEMPLATE",
                        "key": "html",
                        "value": (
                            "<script>fbq('track','Purchase');"
                            "ttq.track('CompletePayment');</script>"
                        ),
                    },
                ],
                "blockingTriggerId": ["99"],
            }
        )
        self.assertEqual({"Meta", "TikTok"}, set(route["detected_vendors"]))
        self.assertEqual("blocker_control_candidate", route["effective_control_status"])
        self.assertFalse(route["forwarded_cmp_signal_visible"])
        self.assertEqual([], route["forwarded_consent_purposes"])
        self.assertEqual([], route["server_consent_forwarding_evidence"])

    def test_context_inference_uses_reachable_behavior_and_separates_gateway(self) -> None:
        data = sample_export()
        data["containerVersion"]["trigger"].append(
            {
                "triggerId": "99",
                "name": "Lead form submit orphan",
                "type": "CUSTOM_EVENT",
                "customEventFilter": [
                    condition("EQUALS", "{{_event}}", "generate_lead")
                ],
            }
        )
        data["containerVersion"]["tag"][0]["parameter"].append(
            {
                "type": "TEMPLATE",
                "key": "server_container_url",
                "value": "https://collect.example.test",
            }
        )
        export = self.root / "reachable-context.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        context = build_context_model(export)["context"]
        self.assertEqual("ecommerce", context["business_model"])
        self.assertNotIn("lead_or_quote", context["business_signals"])
        self.assertEqual(["collect.example.test"], context["server_routing_hosts"])
        self.assertEqual(
            "not_visible_in_container_export",
            context["google_tag_gateway"]["status"],
        )

    def test_cross_object_dependencies_and_empty_objects_remain_reviewable(self) -> None:
        data = sample_export()
        data["containerVersion"]["tag"][0]["setupTag"].append(
            {"tagName": "GA4 - Purchase - All"}
        )
        data["containerVersion"]["tag"][0]["teardownTag"] = [{}]
        data["containerVersion"]["zone"] = [
            {
                "zoneId": "70",
                "name": "Partner Zone",
                "childContainer": [{"publicId": "GTM-CHILD"}],
                "boundary": {"customEvaluationTriggerId": ["10", "999"]},
                "typeRestriction": {},
            }
        ]
        export = self.root / "cross-object-dependencies.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        review = scaffold_configuration(export)
        purchase = next(row for row in review["rows"] if row["object_key"] == "tag:1")
        self.assertTrue(
            any(
                fact["json_path"].endswith(".teardownTag[0]")
                and fact["value_preview"] == "{}"
                for fact in purchase["source_facts"]
            )
        )
        setup = next(
            trace
            for trace in purchase["execution_dependency_traces"]
            if trace["relation"] == "setupTag"
            and trace["reference"] == "GA4 - Purchase - All"
        )
        self.assertEqual("tag:1", setup["targets"][0]["object_key"])
        self.assertEqual("cycle", setup["resolution_state"])
        self.assertTrue(
            any(
                fact["json_path"].startswith("$.containerVersion.trigger[0]")
                for fact in purchase["execution_dependency_facts"]
            )
        )
        self.assertTrue(
            any(
                fact["json_path"].startswith("$.containerVersion.tag[2].parameter")
                and "consentDefault" in str(fact.get("value_preview") or "")
                for fact in purchase["execution_dependency_facts"]
            )
        )
        zone = next(row for row in review["rows"] if row["object_key"] == "zone:70")
        missing = next(
            trace
            for trace in zone["execution_dependency_traces"]
            if trace["reference"] == "999"
        )
        self.assertEqual("missing", missing["resolution_state"])
        architecture = scaffold_architecture(export)
        zone_family = next(
            family
            for family in architecture["families"]
            if "zone:70" in family["member_object_keys"]
        )
        self.assertTrue(
            any(
                edge.get("target_reference") == "999"
                and edge.get("resolution_state") == "missing"
                for edge in zone_family["chain_edges"]
            )
        )
        transaction = next(
            row for row in review["rows"] if row["object_key"] == "variable:25"
        )
        self.assertTrue(transaction["consumer_dependency_facts"])

    def test_custom_code_facts_do_not_infer_loader_dom_or_return_shapes(self) -> None:
        data = sample_export()
        data["containerVersion"]["variable"].append(
            {
                "variableId": "90",
                "name": "CJS - Delegated URL",
                "type": "jsm",
                "parameter": [
                    {
                        "type": "TEMPLATE",
                        "key": "javascript",
                        "value": "function(){ return {{Click URL}}; }",
                    }
                ],
            }
        )
        data["containerVersion"]["tag"].append(
            {
                "tagId": "91",
                "name": "Dynamic vendor loader",
                "type": "html",
                "parameter": [
                    {
                        "type": "TEMPLATE",
                        "key": "html",
                        "value": (
                            "<script>var s=document.createElement('script');"
                            "s.src='https://cdn.example.test/sdk.js';"
                            "document.head.appendChild(s);window.vendorReady=true;</script>"
                        ),
                    }
                ],
                "firingTriggerId": ["10"],
            }
        )
        data["containerVersion"]["customTemplate"] = [
            {
                "templateId": "tpl-90",
                "name": "Opaque template",
                "templateData": "function(){ return 1; }",
            }
        ]
        export = self.root / "custom-code-representation.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        with patch.dict(sys.modules, {"esprima": None}):
            technical = extract_export(export)
        delegated = next(row for row in technical["rows"] if row["object_id"] == "90")
        self.assertEqual(
            "gtm_variable_reference_type_unresolved", delegated["returned_value_type"]
        )
        self.assertTrue(delegated["parser_input_normalized"])
        loader = next(row for row in technical["rows"] if row["object_id"] == "91")
        self.assertTrue(loader["dom_mutations"])
        self.assertFalse(loader["dom_selector_reads"])
        self.assertFalse(
            any(
                "more than one script" in finding.lower()
                for finding in loader["technical_code_optimization_findings"]
            )
        )
        self.assertIn("window/global write", loader["technical_current_behavior"])
        template = next(
            row for row in technical["rows"] if row["object_id"] == "tpl-90"
        )
        self.assertEqual("owner_decision_needed", template["technical_action_candidate"])

    def test_operational_scan_covers_paused_groups_and_template_duplicates(self) -> None:
        data = sample_export()
        data["containerVersion"]["tag"][3].pop("firingTriggerId")
        data["containerVersion"]["trigger"].extend(
            [
                {
                    "triggerId": "30",
                    "name": "TG - Empty",
                    "type": "TRIGGER_GROUP",
                    "parameter": [],
                },
                {
                    "triggerId": "31",
                    "name": "TG - Duplicate members",
                    "type": "TRIGGER_GROUP",
                    "parameter": [
                        {
                            "type": "LIST",
                            "key": "triggerIds",
                            "list": [
                                {"type": "TEMPLATE", "value": "10"},
                                {"type": "TEMPLATE", "value": "10"},
                            ],
                        }
                    ],
                },
                {
                    "triggerId": "32",
                    "name": "TG - Cycle A",
                    "type": "TRIGGER_GROUP",
                    "parameter": [
                        {
                            "type": "LIST",
                            "key": "triggerIds",
                            "list": [{"type": "TEMPLATE", "value": "33"}],
                        }
                    ],
                },
                {
                    "triggerId": "33",
                    "name": "TG - Cycle B",
                    "type": "TRIGGER_GROUP",
                    "parameter": [
                        {
                            "type": "LIST",
                            "key": "triggerIds",
                            "list": [{"type": "TEMPLATE", "value": "32"}],
                        }
                    ],
                },
            ]
        )
        data["containerVersion"]["customTemplate"] = [
            {"templateId": "tpl-1", "name": "Template one", "templateData": "same"},
            {"templateId": "tpl-2", "name": "Template two", "templateData": "same"},
        ]
        path = self.root / "operational-adversary.json"
        path.write_text(json.dumps(data), encoding="utf-8")
        findings = [
            row for row in audit_export(path)["findings"] if row["finding_type"] != "zero_findings"
        ]
        types = {row["finding_type"] for row in findings}
        self.assertIn("paused_objects_for_lifecycle_review", types)
        self.assertIn("empty_trigger_group", types)
        self.assertIn("duplicate_trigger_group_members", types)
        self.assertIn("cyclic_trigger_groups", types)
        self.assertTrue(
            any(
                row["module_name"] == "duplicate_custom_template_configurations" for row in findings
            )
        )
        triggerless = [
            row for row in findings if row["module_name"] == "tags_without_firing_triggers"
        ]
        self.assertFalse(any("4" in row["object_ids"] for row in triggerless))

    def test_operational_scan_resolves_reachability_sequences_and_contradictions(self) -> None:
        data = sample_export()
        data["containerVersion"]["variable"].extend(
            [
                {
                    "variableId": "80",
                    "name": "Orphan A",
                    "type": "c",
                    "parameter": [
                        {"type": "TEMPLATE", "key": "value", "value": "{{Orphan B}}"}
                    ],
                },
                {
                    "variableId": "81",
                    "name": "Orphan B",
                    "type": "c",
                    "parameter": [
                        {"type": "TEMPLATE", "key": "value", "value": "{{Orphan A}}"}
                    ],
                },
            ]
        )
        data["containerVersion"]["tag"].extend(
            [
                {
                    "tagId": "80",
                    "name": "Sequence A",
                    "type": "html",
                    "setupTag": [{"tagName": "Sequence B"}],
                    "scheduleStartMs": "200",
                    "scheduleEndMs": "100",
                },
                {
                    "tagId": "81",
                    "name": "Sequence B",
                    "type": "html",
                    "teardownTag": [{"tagName": "Sequence A"}],
                },
            ]
        )
        data["containerVersion"]["trigger"].extend(
            [
                {
                    "triggerId": "80",
                    "name": "CE - Impossible",
                    "type": "CUSTOM_EVENT",
                    "filter": [
                        condition("EQUALS", "{{_event}}", "purchase"),
                        condition("NOT_EQUALS", "{{_event}}", "purchase"),
                    ],
                },
                {
                    "triggerId": "90",
                    "name": "TG - Parent",
                    "type": "TRIGGER_GROUP",
                    "parameter": [
                        {
                            "type": "LIST",
                            "key": "triggerIds",
                            "list": [{"type": "TEMPLATE", "value": "91"}],
                        }
                    ],
                },
                {
                    "triggerId": "91",
                    "name": "TG - Child",
                    "type": "TRIGGER_GROUP",
                    "parameter": [
                        {
                            "type": "LIST",
                            "key": "triggerIds",
                            "list": [{"type": "TEMPLATE", "value": "10"}],
                        }
                    ],
                },
            ]
        )
        data["containerVersion"]["zone"] = [
            {
                "zoneId": "95",
                "name": "Impossible Zone",
                "childContainer": [{"publicId": "GTM-CHILD"}],
                "boundary": {
                    "condition": [
                        condition("EQUALS", "market", "fr"),
                        condition("NOT_EQUALS", "market", "fr"),
                    ]
                },
            }
        ]
        path = self.root / "reachability-and-sequence.json"
        path.write_text(json.dumps(data), encoding="utf-8")
        scan = audit_export(path)
        findings = [
            row for row in scan["findings"] if row["finding_type"] != "zero_findings"
        ]
        finding_types = {row["finding_type"] for row in findings}
        self.assertIn("cyclic_tag_sequence", finding_types)
        self.assertIn("invalid_tag_schedule_order", finding_types)
        self.assertIn("contradictory_trigger_conditions", finding_types)
        self.assertIn("contradictory_zone_boundary_conditions", finding_types)
        self.assertIn("nested_trigger_groups", finding_types)
        unused_variable_ids = {
            object_id
            for row in findings
            if row["module_name"] == "unused_variables"
            for object_id in row["object_ids"]
        }
        self.assertTrue({"80", "81"}.issubset(unused_variable_ids))
        unreachable_tags = {
            object_id
            for row in findings
            if row["module_name"] == "tags_without_firing_triggers"
            for object_id in row["object_ids"]
        }
        self.assertTrue({"80", "81"}.issubset(unreachable_tags))

    def test_operational_scan_fails_visible_on_malformed_nested_controls(self) -> None:
        data = sample_export()
        data["containerVersion"]["tag"][0]["setupTag"] = {
            "tagName": "Utility - Consent Defaults"
        }
        data["containerVersion"]["tag"][0]["teardownTag"] = ["bad-entry", {}]
        data["containerVersion"]["tag"][0]["consentSettings"] = []
        data["containerVersion"]["tag"][1]["consentSettings"] = {
            "consentStatus": "futureMaybe"
        }
        group = next(
            row for row in data["containerVersion"]["trigger"] if row["triggerId"] == "12"
        )
        group["parameter"][0]["list"] = ["bad-member", {"value": ""}]
        data["containerVersion"]["zone"] = [
            {
                "zoneId": "95",
                "name": "Malformed Zone",
                "childContainer": [{"publicId": "GTM-CHILD"}],
                "boundary": {"condition": "not-an-array"},
                "typeRestriction": {
                    "enable": True,
                    "whitelistedTypeId": "html",
                },
            }
        ]
        path = self.root / "malformed-nested-controls.json"
        path.write_text(json.dumps(data), encoding="utf-8")
        scan = audit_export(path)
        self.assertEqual("complete", scan["run_status"])
        finding_types = {
            row["finding_type"]
            for row in scan["findings"]
            if row["finding_type"] != "zero_findings"
        }
        self.assertTrue(
            {
                "invalid_tag_sequence_shape",
                "invalid_tag_sequence_entry",
                "tag_sequence_target_missing_name",
                "invalid_trigger_group_member_structure",
                "invalid_zone_boundary_field_shape",
                "invalid_zone_type_allowlist_shape",
                "invalid_consent_settings_shape",
                "unrecognized_manual_consent_status",
            }.issubset(finding_types)
        )
        package = self.root / "malformed-nested-package"
        manifest = build_package(path, package, pretty=True)
        self.assertEqual("pass", manifest["status"])
        self.assertTrue((package / "operational_review.json").exists())
        self.assertTrue((package / "configuration_review.json").exists())
        self.assertTrue((package / "architecture_review.json").exists())

    def test_ineffective_blocker_requires_every_firing_route_to_be_exact(self) -> None:
        data = sample_export()
        data["containerVersion"]["tag"][0]["firingTriggerId"].append("14")
        path = self.root / "mixed-firing-routes.json"
        path.write_text(json.dumps(data), encoding="utf-8")
        findings = audit_export(path)["findings"]
        self.assertFalse(
            any(
                row["finding_type"] == "ineffective_blocking_trigger"
                and "1" in row["object_ids"]
                for row in findings
            )
        )

    def test_relationship_run_detects_scope_and_status_without_names_only(self) -> None:
        rows = relationship_candidates(container_version(sample_export()))
        q1 = [
            row
            for row in rows
            if "shared_business_scope" in row["comparison_types"]
            and {"trigger:15", "trigger:16"}.issubset(row["candidate_object_keys"])
        ]
        self.assertEqual(1, len(q1))
        duplicate = next(
            row
            for row in rows
            if {"variable:20", "variable:21"} == set(row["candidate_object_keys"])
        )
        self.assertIn("exact_configuration", duplicate["comparison_types"])
        self.assertEqual(False, duplicate["candidate_paused_status"]["variable:20"])

    def test_duplicate_logic_ignores_export_metadata_and_folder_placement(self) -> None:
        data = sample_export()
        variables = {
            row["variableId"]: row for row in data["containerVersion"]["variable"]
        }
        variables["20"].update(
            {
                "workspaceId": "7",
                "tagManagerUrl": "https://tagmanager.google.com/variable/20",
                "notes": "Original implementation",
                "parentFolderId": "100",
            }
        )
        variables["21"].update(
            {
                "workspaceId": "8",
                "tagManagerUrl": "https://tagmanager.google.com/variable/21",
                "notes": "Later copy",
                "parentFolderId": "101",
            }
        )
        path = self.root / "metadata-different-duplicates.json"
        path.write_text(json.dumps(data), encoding="utf-8")
        finding = next(
            row
            for row in audit_export(path)["findings"]
            if row["module_name"] == "duplicate_variable_logic"
            and {"20", "21"}.issubset(row["object_ids"])
        )
        self.assertEqual("duplicate_configuration", finding["finding_type"])
        comparison = next(
            row
            for row in relationship_candidates(data["containerVersion"])
            if "exact_configuration" in row["comparison_types"]
            and {"variable:20", "variable:21"}.issubset(
                row["candidate_object_keys"]
            )
        )
        self.assertEqual(1.0, comparison["similarity_score"])
        configuration = scaffold_configuration(path)
        variable_review = next(
            row
            for row in configuration["rows"]
            if row["object_key"] == "variable:20"
        )
        metadata_suffixes = (
            ".workspaceId",
            ".tagManagerUrl",
            ".notes",
            ".parentFolderId",
        )
        self.assertFalse(
            any(
                source_path.endswith(metadata_suffixes)
                for source_path in variable_review["required_logic_anchors"]
            )
        )

    def test_relationship_run_reviews_multiple_firing_routes(self) -> None:
        data = sample_export()
        data["containerVersion"]["tag"][0]["firingTriggerId"] = ["15", "16"]
        rows = relationship_candidates(container_version(data))
        route = [
            row
            for row in rows
            if "multi_firing_route_consolidation_review" in row["comparison_types"]
            and set(row["candidate_object_keys"]) == {"trigger:15", "trigger:16"}
        ]
        self.assertEqual(1, len(route))

    def test_architecture_accepts_source_grounded_open_discovery(self) -> None:
        review = complete_architecture(self.export_path)
        review["comparisons"].append(value_discovery_row(self.export_path))
        review["open_discovery_attestation"]["discovered_comparison_ids"] = [
            "DISC-VALUE-001"
        ]
        review["open_discovery_attestation"]["zero_discovery_rationale"] = ""
        next(
            item
            for item in review["open_discovery_attestation"]["method_reviews"]
            if item["method"] == "terminal_source_formula_and_output_overlap"
        )["additional_discovery_ids"] = ["DISC-VALUE-001"]
        errors, _ = validate_architecture(
            self.export_path,
            self.write_review("open-discovery.json", review),
        )
        self.assertEqual([], errors)

        mismatched = copy.deepcopy(review)
        mismatched_discovery = next(
            row
            for row in mismatched["comparisons"]
            if row["comparison_id"] == "DISC-VALUE-001"
        )
        mismatched_discovery["comparison_types"] = [
            "semantic_name_family_candidate"
        ]
        mismatched_errors, _ = validate_architecture(
            self.export_path,
            self.write_review("mismatched-discovery-method.json", mismatched),
        )
        self.assertTrue(
            any(
                "declared comparison types require discovery methods" in error
                for error in mismatched_errors
            )
        )

    def test_configuration_scaffold_requires_every_object_branch_code_and_trace(self) -> None:
        review = scaffold_configuration(self.export_path)
        source_count = sum(
            len(sample_export()["containerVersion"].get(layer, []))
            for layer in (
                "tag",
                "trigger",
                "variable",
                "customTemplate",
                "client",
                "transformation",
            )
        )
        self.assertEqual(source_count, len(review["rows"]))
        meta = next(
            row for row in review["rows"] if row["object_id"] == "2" and row["layer"] == "tag"
        )
        self.assertTrue(meta["required_branch_reviews"])
        self.assertTrue(meta["required_code_line_hashes"])
        self.assertTrue(meta["reference_trace_requirements"])
        self.assertTrue(meta["technical_code_facts"])

    def test_configuration_gate_rejects_missing_branch_and_generic_code(self) -> None:
        review = complete_configuration(self.export_path)
        custom = next(
            row for row in review["rows"] if row["object_id"] == "2" and row["layer"] == "tag"
        )
        custom["configuration_branch_reviews"].pop()
        custom["code_behavior_blocks"][0]["purpose"] = "Code inspected and reviewed"
        path = self.write_review("bad-config.json", review)
        errors, _ = validate_configuration(self.export_path, path)
        self.assertTrue(any("branch reviews" in error for error in errors))
        self.assertTrue(any("incomplete purpose" in error for error in errors))

    def test_configuration_gate_rejects_generic_semantic_prose_with_valid_citations(self) -> None:
        review = complete_configuration(self.export_path)
        purchase = next(row for row in review["rows"] if row["object_key"] == "tag:1")
        purchase["purpose"] = (
            "GA4 - Purchase - All serves one concrete measurement purpose through its tag setup."
        )
        errors, _ = validate_configuration(
            self.export_path,
            self.write_review("generic-semantic-prose.json", review),
        )
        self.assertTrue(any("purpose lacks object-specific analysis" in error for error in errors))

    def test_review_context_is_content_locked_not_hash_only(self) -> None:
        review = complete_configuration(self.export_path)
        review["audit_context"]["business_model"] = "publisher"
        errors, _ = validate_configuration(
            self.export_path,
            self.write_review("tampered-context.json", review),
        )
        self.assertTrue(any("audit_context differs" in error for error in errors))

    def test_unknown_external_vendor_creates_official_research_contract(self) -> None:
        data = sample_export()
        data["containerVersion"]["tag"].append(
            {
                "tagId": "90",
                "name": "Partner widget",
                "type": "html",
                "parameter": [
                    {
                        "type": "TEMPLATE",
                        "key": "html",
                        "value": (
                            '<script src="https://unknown-cdn.example/widget.js"></script>'
                        ),
                    }
                ],
                "firingTriggerId": ["10"],
            }
        )
        path = self.root / "unknown-vendor.json"
        path.write_text(json.dumps(data), encoding="utf-8")
        review = scaffold_configuration(path)
        partner = next(row for row in review["rows"] if row["object_key"] == "tag:90")
        self.assertTrue(
            any(context["category"] == "unknown_vendor" for context in partner["vendor_contexts"])
        )
        self.assertTrue(any(topic["research_required"] for topic in partner["required_contract_topics"]))
        context = build_context_model(path)
        self.assertIn("unknown-cdn.example", context["context"]["external_hosts"])
        self.assertNotIn("unknown-cdn.example", context["context"]["server_routing_hosts"])
        self.assertEqual("Unclassified", vendor_record('{"key":"contents"}')["name"])
        self.assertEqual("Unclassified", vendor_record('{"key":"activity"}')["name"])

    def test_configuration_requires_contract_topics_and_recursive_node_meaning(self) -> None:
        scaffold = scaffold_configuration(self.export_path)
        purchase = next(row for row in scaffold["rows"] if row["object_key"] == "tag:1")
        topics = {item["topic"] for item in purchase["required_contract_topics"]}
        self.assertIn("ecommerce_event_contract", topics)
        self.assertIn("transaction_value_currency_and_quantity", topics)
        value_variable = next(row for row in scaffold["rows"] if row["object_key"] == "variable:24")
        self.assertTrue(
            any(
                context["vendor"] == "GA4 / Google tag"
                for context in value_variable["vendor_contexts"]
            )
        )
        self.assertTrue(value_variable["required_contract_topics"])

        review = complete_configuration(self.export_path)
        completed_purchase = next(row for row in review["rows"] if row["object_key"] == "tag:1")
        completed_purchase["contract_checks"].pop()
        errors, _ = validate_configuration(
            self.export_path,
            self.write_review("missing-contract-topic.json", review),
        )
        self.assertTrue(any("every generated topic" in error for error in errors))

        review = complete_configuration(self.export_path)
        completed_purchase = next(row for row in review["rows"] if row["object_key"] == "tag:1")
        completed_purchase["contract_checks"][0]["expected_rule"] = (
            "The official vendor documentation was reviewed for this configuration."
        )
        errors, _ = validate_configuration(
            self.export_path,
            self.write_review("generic-contract-rule.json", review),
        )
        self.assertTrue(any("topic-specific contract" in error for error in errors))

        review = complete_configuration(self.export_path)
        meta = next(row for row in review["rows"] if row["object_key"] == "tag:2")
        trace = next(
            item for item in meta["reference_traces"] if item["reference"] == "DLV - Items"
        )
        trace["node_reviews"].clear()
        errors, _ = validate_configuration(
            self.export_path,
            self.write_review("missing-trace-node.json", review),
        )
        self.assertTrue(any("every variable node" in error for error in errors))

    def test_complete_configuration_and_architecture_reviews_validate(self) -> None:
        configuration = complete_configuration(self.export_path)
        architecture = complete_architecture(self.export_path)
        config_errors, _ = validate_configuration(
            self.export_path, self.write_review("configuration.json", configuration)
        )
        architecture_errors, _ = validate_architecture(
            self.export_path, self.write_review("architecture.json", architecture)
        )
        self.assertEqual([], config_errors)
        self.assertEqual([], architecture_errors)
        purchase_family = next(
            row for row in architecture["families"] if "tag:1" in row["member_object_keys"]
        )
        self.assertIn("trigger:10", purchase_family["chain_object_keys"])
        self.assertIn("trigger:13", purchase_family["chain_object_keys"])
        self.assertIn("variable:24", purchase_family["chain_object_keys"])
        self.assertIn("variable:25", purchase_family["chain_object_keys"])
        self.assertIn("tag:3", purchase_family["chain_object_keys"])
        self.assertEqual(
            set(purchase_family["chain_object_keys"]),
            {item["object_key"] for item in purchase_family["chain_assessments"]},
        )

    def test_architecture_verdicts_and_zero_discovery_are_fail_closed(self) -> None:
        review = complete_architecture(self.export_path)
        comparison = next(
            row
            for row in review["comparisons"]
            if "exact_configuration" not in row.get("comparison_types", [])
        )
        comparison.update(
            {
                "relationship_verdict": "Owner decision needed",
                "disposition": "keep",
                "owner_question": "",
            }
        )
        errors, _ = validate_architecture(
            self.export_path,
            self.write_review("incoherent-architecture-verdict.json", review),
        )
        self.assertTrue(any("requires owner_decision_needed" in error for error in errors))
        self.assertTrue(any("precise question" in error for error in errors))

        review = complete_architecture(self.export_path)
        comparison = next(
            row
            for row in review["comparisons"]
            if row["relationship_verdict"] in {"Intentional variant", "Complementary"}
            and len(row["candidate_object_keys"]) > 1
        )
        for assessment in comparison["member_assessments"]:
            assessment["distinguishing_configuration"] = (
                "This object is kept because the general source review found a different role."
            )
        errors, _ = validate_architecture(
            self.export_path,
            self.write_review("generic-retention-distinction.json", review),
        )
        self.assertTrue(
            any("configuration term unique to that member" in error for error in errors)
        )

        review = complete_architecture(self.export_path)
        review["open_discovery_attestation"]["zero_discovery_rationale"] = (
            "Every object was reviewed carefully and no additional relationship was found."
        )
        errors, _ = validate_architecture(
            self.export_path,
            self.write_review("generic-zero-discovery.json", review),
        )
        self.assertTrue(any("naming every discovery method" in error for error in errors))

    def test_purchase_transaction_id_absence_cannot_be_marked_compliant(self) -> None:
        data = sample_export()
        event_parameters = data["containerVersion"]["tag"][0]["parameter"][2]["map"]
        data["containerVersion"]["tag"][0]["parameter"][2]["map"] = [
            item for item in event_parameters if item.get("key") != "transaction_id"
        ]
        export = self.root / "missing-transaction-id.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        scaffold = scaffold_configuration(export)
        purchase = next(row for row in scaffold["rows"] if row["object_key"] == "tag:1")
        topic = next(
            topic
            for topic in purchase["required_contract_topics"]
            if topic["topic"] == "purchase_transaction_id_uniqueness"
        )
        self.assertEqual("missing", topic["configuration_presence_state"])

        completed = complete_configuration(export)
        purchase_review = next(
            row for row in completed["rows"] if row["object_key"] == "tag:1"
        )
        purchase_check = next(
            check
            for check in purchase_review["contract_checks"]
            if check["contract_topic"] == topic["topic_key"]
        )
        purchase_check["verdict"] = "Compliant"
        errors, _ = validate_configuration(
            export,
            self.write_review("missing-transaction-id-review.json", completed),
        )
        self.assertTrue(
            any(
                "required exported configuration terms" in error
                and "transaction_id" in error
                for error in errors
            )
        )

    def test_architecture_requires_every_chain_object_and_server_root(self) -> None:
        review = complete_architecture(self.export_path)
        family = next(row for row in review["families"] if "tag:1" in row["member_object_keys"])
        family["chain_assessments"].pop()
        errors, _ = validate_architecture(
            self.export_path,
            self.write_review("missing-chain-assessment.json", review),
        )
        self.assertTrue(any("chain" in error and "every member" in error for error in errors))

        data = sample_export()
        data["containerVersion"]["client"] = [
            {"clientId": "50", "name": "GA4 client", "type": "gaaw_client"}
        ]
        data["containerVersion"]["transformation"] = [
            {
                "transformationId": "60",
                "name": "Redact user data",
                "type": "exclude_parameters",
            }
        ]
        server_path = self.root / "server.json"
        server_path.write_text(json.dumps(data), encoding="utf-8")
        server_review = scaffold_architecture(server_path)
        root_keys = {
            key
            for family_row in server_review["families"]
            for key in family_row["member_object_keys"]
        }
        self.assertIn("client:50", root_keys)
        self.assertIn("transformation:60", root_keys)

    def test_operational_validator_requires_all_findings(self) -> None:
        review = complete_operational(self.export_path)
        review["findings"].pop()
        errors, _ = validate_operational(
            self.export_path, self.write_review("bad-operational.json", review)
        )
        self.assertTrue(any("missing operational findings" in error for error in errors))

    def test_compiler_blocks_unaligned_basic_consolidation(self) -> None:
        operational, configuration, architecture = self.completed_reviews()
        operation = duplicate_variable_operation()
        finding = next(
            row
            for row in operational["findings"]
            if row["module_name"] == "duplicate_variable_paths"
        )
        finding.update(copy.deepcopy(operation))
        finding["disposition"] = "cleanup_operation"
        payload, errors = compile_operations(
            operational,
            configuration,
            architecture,
            "Direct GTM/MCP/API",
            "Deep",
            source_object_catalog(self.export_path),
        )
        self.assertEqual([], payload["operations"])
        self.assertTrue(any("lacks an aligned business-architecture" in error for error in errors))

    def test_three_runs_reconcile_and_future_state_passes(self) -> None:
        operational, configuration, architecture = self.completed_reviews()
        align_duplicate_operation(operational, architecture)
        for name, review, validator in (
            ("operational.json", operational, validate_operational),
            ("configuration.json", configuration, validate_configuration),
            ("architecture.json", architecture, validate_architecture),
        ):
            errors, _ = validator(self.export_path, self.write_review(name, review))
            self.assertEqual([], errors)
        payload, errors = compile_operations(
            operational,
            configuration,
            architecture,
            "Direct GTM/MCP/API",
            "Deep",
            source_object_catalog(self.export_path),
        )
        self.assertEqual([], errors)
        self.assertEqual(1, len(payload["operations"]))
        self.assertEqual(
            ["business_architecture", "operational_sanitation"],
            payload["operations"][0]["source_runs"],
        )
        self.assertTrue(payload["shared_facts_sha256"])
        self.assertTrue(payload["decision_ledger"])
        self.assertEqual(-1, payload["projected_object_counts"]["variable"]["delta"])
        self.assertEqual(
            ["delete"],
            payload["operations"][0]["execution_phases"],
        )
        report, future_errors = check_future_state(self.export_path, payload)
        self.assertEqual([], future_errors)
        self.assertEqual("pass", report["status"])
        self.assertEqual(-1, report["object_counts"]["variable"]["delta"])

    def test_compiler_rejects_same_key_with_different_mutations(self) -> None:
        operational, configuration, architecture = self.completed_reviews()
        align_duplicate_operation(operational, architecture)
        comparison = next(
            row
            for row in architecture["comparisons"]
            if set(row["candidate_object_keys"]) == {"variable:20", "variable:21"}
        )
        comparison["operations"][0]["canonical_object_key"] = "variable:21"
        comparison["operations"][0]["deletions"] = [
            {
                "object_key": "variable:20",
                "reason": "Use variable 21 as the conflicting canonical object in this test.",
            }
        ]
        _, errors = compile_operations(operational, configuration, architecture, "Manual", "Deep")
        self.assertTrue(
            any("reused for different structured mutations" in error for error in errors)
        )

    def test_compiler_merges_wording_variants_for_the_same_mutation(self) -> None:
        operational, configuration, architecture = self.completed_reviews()
        align_duplicate_operation(operational, architecture)
        comparison = next(
            row
            for row in architecture["comparisons"]
            if set(row["candidate_object_keys"]) == {"variable:20", "variable:21"}
        )
        comparison["operations"][0]["problem"] = (
            "DLV - Items Copy repeats the canonical ecommerce.items variable without consumers."
        )
        payload, errors = compile_operations(
            operational,
            configuration,
            architecture,
            "Manual",
            "Deep",
        )
        self.assertEqual([], errors)
        self.assertEqual(1, len(payload["operations"]))
        self.assertEqual(3, len(payload["operations"][0]["lens_rationales"]))
        self.assertEqual(
            {"operational_sanitation", "business_architecture"},
            {
                row["source_run"]
                for row in payload["operations"][0]["lens_rationales"]
            },
        )

    def test_future_state_blocks_deletion_that_breaks_a_consumer(self) -> None:
        operational, configuration, architecture = self.completed_reviews()
        align_duplicate_operation(operational, architecture)
        payload, errors = compile_operations(
            operational, configuration, architecture, "Manual", "Deep"
        )
        self.assertEqual([], errors)
        payload["operations"][0]["deletions"] = [
            {"object_key": "variable:20", "reason": "Intentional broken test deletion."}
        ]
        payload["operations"][0]["affected_object_keys"] = ["variable:20"]
        report, future_errors = check_future_state(self.export_path, payload)
        self.assertEqual("fail", report["status"])
        self.assertTrue(any("missing references" in error for error in future_errors))

    def test_future_state_rejects_stale_before_value(self) -> None:
        operational, configuration, architecture = self.completed_reviews()
        align_duplicate_operation(operational, architecture)
        payload, errors = compile_operations(
            operational, configuration, architecture, "Manual", "Deep"
        )
        self.assertEqual([], errors)
        payload["operations"][0]["changes"] = [
            {
                "object_key": "tag:1",
                "json_path": "$.containerVersion.tag[0].name",
                "before": "A stale tag name",
                "after": "GA4 - Purchase - Global",
            }
        ]
        report, future_errors = check_future_state(self.export_path, payload)
        self.assertEqual("fail", report["status"])
        self.assertTrue(any("before value does not match" in error for error in future_errors))

    def test_structured_operations_support_creation_and_missing_field_addition(self) -> None:
        operation = {
            "problem_type": "Missing tracking",
            "minimum_aggressiveness": "Standard",
            "creations": [
                {
                    "layer": "variable",
                    "object": {
                        "variableId": "99",
                        "name": "Constant - Test Value",
                        "type": "c",
                        "parameter": [
                            {"type": "TEMPLATE", "key": "value", "value": "test"}
                        ],
                    },
                    "reason": "Create the missing constant required by the approved test tag.",
                }
            ],
            "additions": [
                {
                    "object_key": "tag:1",
                    "json_path": "$.containerVersion.tag[0].parameter",
                    "mode": "append",
                    "value": {
                        "type": "TEMPLATE",
                        "key": "testParameter",
                        "value": "{{Constant - Test Value}}",
                    },
                    "reason": "Add the approved parameter to the existing purchase tag.",
                }
            ],
            "changes": [],
            "remaps": [],
            "renames": [],
            "deletions": [],
        }
        validation_errors = validate_structured_actions(
            operation,
            object_keys(self.export_path),
            "creation test",
            object_consumer_map(self.export_path),
            object_source_path_map(self.export_path),
        )
        self.assertEqual([], validation_errors)
        mismatched = copy.deepcopy(operation)
        mismatched["additions"][0]["json_path"] = (
            "$.containerVersion.tag[1].parameter"
        )
        mismatch_errors = validate_structured_actions(
            mismatched,
            object_keys(self.export_path),
            "source-path binding test",
            object_consumer_map(self.export_path),
            object_source_path_map(self.export_path),
        )
        self.assertTrue(
            any("paired with another object's json_path" in error for error in mismatch_errors)
        )
        future, apply_errors = apply_operations(
            sample_export(),
            {"operations": [operation]},
        )
        self.assertEqual([], apply_errors)
        future_cv = container_version(future)
        self.assertTrue(any(row.get("variableId") == "99" for row in future_cv["variable"]))
        self.assertEqual(
            "testParameter",
            future_cv["tag"][0]["parameter"][-1]["key"],
        )

    def test_aggressiveness_defers_operations_below_their_minimum_level(self) -> None:
        operational, configuration, architecture = self.completed_reviews()
        align_duplicate_operation(operational, architecture)
        conservative, errors = compile_operations(
            operational,
            configuration,
            architecture,
            "Manual",
            "Conservative",
        )
        self.assertEqual([], errors)
        self.assertEqual([], conservative["operations"])
        self.assertEqual(1, len(conservative["deferred_operations"]))
        self.assertEqual(
            "deferred_by_aggressiveness",
            conservative["deferred_operations"][0]["resolution_status"],
        )
        human, human_errors = build_rows(conservative)
        self.assertEqual([], human_errors)
        self.assertEqual("Deferred", human[0]["Level"])

    def test_human_plan_exposes_owner_decisions_without_internal_proof_columns(self) -> None:
        payload = {
            "operations": [],
            "deferred_operations": [],
            "decision_ledger": [
                {
                    "decision_id": "CFG-001",
                    "disposition": "owner_decision_needed",
                    "area": "Governance / ownership",
                    "problem_type": "Unclear business purpose",
                    "affected_objects": "tag:1 - Legacy lead",
                    "summary": "Legacy lead sends a second conversion for the same form submit.",
                    "owner_question": (
                        "Should Legacy lead remain a separate paid-media conversion?"
                    ),
                }
            ],
        }
        rows, errors = build_rows(payload)
        self.assertEqual([], errors)
        self.assertEqual(1, len(rows))
        self.assertEqual("Owner decision", rows[0]["Level"])
        self.assertEqual(6, len(rows[0]))

    def test_remap_requires_the_exact_source_consumer_set(self) -> None:
        review = complete_configuration(self.export_path)
        row = next(item for item in review["rows"] if item["object_key"] == "variable:20")
        operation = duplicate_variable_operation()
        operation.update(
            {
                "operation_key": "remap-items-variable",
                "canonical_object_key": "variable:21",
                "changes": [],
                "deletions": [],
                "remaps": [
                    {
                        "from_object_key": "variable:20",
                        "to_object_key": "variable:21",
                        "consumer_object_keys": ["tag:1"],
                    }
                ],
                "exact_proposed_action": (
                    "Remap the items variable to the selected canonical variable."
                ),
            }
        )
        row["disposition"] = "cleanup_operation"
        row["operation"] = operation
        errors, _ = validate_configuration(
            self.export_path,
            self.write_review("bad-remap-consumers.json", review),
        )
        self.assertTrue(
            any("exactly match every source-graph consumer" in error for error in errors)
        )

    def test_package_build_contains_three_independent_review_artifacts(self) -> None:
        package_dir = self.root / "package"
        manifest = build_package(self.export_path, package_dir, pretty=True)
        self.assertEqual("pass", manifest["status"])
        for filename in (
            "context.json",
            "shared_facts.json",
            "operational_review.json",
            "configuration_review.json",
            "architecture_review.json",
            "operational_scan.json",
        ):
            self.assertTrue((package_dir / filename).is_file())
        self.assertNotIn("semantic_review", manifest["files"])
        shared = json.loads((package_dir / "shared_facts.json").read_text(encoding="utf-8"))
        context = json.loads((package_dir / "context.json").read_text(encoding="utf-8"))
        self.assertEqual(shared["shared_facts_sha256"], manifest["shared_facts_sha256"])
        self.assertEqual(context["context_sha256"], manifest["context_sha256"])
        for filename in (
            "operational_review.json",
            "configuration_review.json",
            "architecture_review.json",
        ):
            review = json.loads((package_dir / filename).read_text(encoding="utf-8"))
            self.assertEqual(shared["shared_facts_sha256"], review["shared_facts_sha256"])

    def test_package_gate_rejects_shared_fact_content_with_a_copied_hash(self) -> None:
        package_dir = self.root / "tampered-package"
        build_package(self.export_path, package_dir, pretty=True)
        shared_path = package_dir / "shared_facts.json"
        shared = json.loads(shared_path.read_text(encoding="utf-8"))
        shared["objects"][0]["object_name"] = "Fabricated object name"
        shared_path.write_text(json.dumps(shared), encoding="utf-8")
        report = run_gate(self.export_path, package_dir, audit_only=True)
        self.assertEqual("fail", report["status"])
        self.assertTrue(any("recorded hash" in error for error in report["errors"]))

    def test_broken_references_remain_auditable_integrity_findings(self) -> None:
        data = sample_export()
        data["containerVersion"]["tag"][0]["firingTriggerId"] = ["999"]
        path = self.root / "broken-reference.json"
        path.write_text(json.dumps(data), encoding="utf-8")
        model = build_model(path)
        self.assertEqual("pass_with_integrity_findings", model["coverage_gate"])
        package_dir = self.root / "broken-package"
        manifest = build_package(path, package_dir, pretty=True)
        self.assertEqual("pass", manifest["status"])
        shared = build_shared_facts(path, context=build_context_model(path))
        self.assertEqual("pass_with_integrity_findings", shared["coverage_gate"])
        findings = audit_export(path)["findings"]
        self.assertTrue(
            any(row["finding_type"] == "missing_trigger_reference" for row in findings)
        )

    def test_verdict_engines_share_only_neutral_review_helpers(self) -> None:
        configuration_source = (SCRIPTS / "gtm_configuration_review.py").read_text(encoding="utf-8")
        architecture_source = (SCRIPTS / "gtm_architecture_review.py").read_text(encoding="utf-8")
        self.assertNotIn("from gtm_operational_review import", configuration_source)
        self.assertNotIn("from gtm_operational_review import", architecture_source)
        self.assertIn("from gtm_review_common import", configuration_source)
        self.assertIn("from gtm_review_common import", architecture_source)

    def test_large_review_shards_merge_only_with_complete_exact_coverage(self) -> None:
        completed = complete_configuration(self.export_path)
        base_path = self.write_review("configuration-complete.json", completed)
        shard_dir = self.root / "configuration-shards"
        manifest = split_review(base_path, shard_dir, max_items=3)
        self.assertGreater(len(manifest["shards"]), 1)
        output = self.root / "configuration-merged.json"
        merged = merge_review(base_path, shard_dir, output)
        self.assertEqual(completed["rows"], merged["rows"])
        self.assertEqual("complete", merged["run_status"])

        shard_manifest_path = shard_dir / "shard_manifest.json"
        broken_manifest = json.loads(shard_manifest_path.read_text(encoding="utf-8"))
        broken_manifest["shards"][0]["filename"] = "missing-shard.json"
        shard_manifest_path.write_text(json.dumps(broken_manifest), encoding="utf-8")
        with self.assertRaisesRegex(ValueError, "missing review shard"):
            merge_review(base_path, shard_dir, self.root / "must-not-exist.json")

    def test_architecture_shards_merge_discovery_rows_and_attestation(self) -> None:
        completed = complete_architecture(self.export_path)
        completed["comparisons"] = [
            row
            for row in completed["comparisons"]
            if row.get("comparison_origin") != "analyst_discovered"
        ]
        completed["open_discovery_attestation"]["discovered_comparison_ids"] = []
        base_path = self.write_review("architecture-complete.json", completed)
        shard_dir = self.root / "architecture-shards"
        manifest = split_review(base_path, shard_dir, max_items=2)
        discovery_path = shard_dir / manifest["discovery_shard"]
        discovery = json.loads(discovery_path.read_text(encoding="utf-8"))
        discovery["discovered_comparisons"] = [value_discovery_row(self.export_path)]
        discovery["open_discovery_attestation"].update(
            {
                "review_status": "complete",
                "discovered_comparison_ids": ["DISC-VALUE-001"],
                "zero_discovery_rationale": "",
            }
        )
        next(
            item
            for item in discovery["open_discovery_attestation"]["method_reviews"]
            if item["method"] == "terminal_source_formula_and_output_overlap"
        )["additional_discovery_ids"] = ["DISC-VALUE-001"]
        discovery_path.write_text(json.dumps(discovery), encoding="utf-8")
        merged_path = self.root / "architecture-merged.json"
        merged = merge_review(base_path, shard_dir, merged_path)
        self.assertIn(
            "DISC-VALUE-001",
            {row["comparison_id"] for row in merged["comparisons"]},
        )
        errors, _ = validate_architecture(self.export_path, merged_path)
        self.assertEqual([], errors)

    def test_three_run_gate_rejects_scaffolds_and_passes_completed_reviews(self) -> None:
        package_dir = self.root / "package"
        build_package(self.export_path, package_dir, pretty=True)
        pending = run_gate(self.export_path, package_dir)
        self.assertEqual("fail", pending["status"])
        operational, configuration, architecture = self.completed_reviews()
        align_duplicate_operation(operational, architecture)
        (package_dir / "operational_review.json").write_text(
            json.dumps(operational), encoding="utf-8"
        )
        (package_dir / "configuration_review.json").write_text(
            json.dumps(configuration), encoding="utf-8"
        )
        (package_dir / "architecture_review.json").write_text(
            json.dumps(architecture), encoding="utf-8"
        )
        payload, compile_errors = compile_operations(
            operational,
            configuration,
            architecture,
            "Direct GTM/MCP/API",
            "Deep",
            source_object_catalog(self.export_path),
        )
        self.assertEqual([], compile_errors)
        operations_path = self.write_review("operations.json", payload)
        completed = run_gate(self.export_path, package_dir, operations_path)
        self.assertEqual("pass", completed["status"], completed["errors"])

        tampered = copy.deepcopy(payload)
        tampered["operations"][0]["exact_proposed_action"] = (
            "A hand-edited action that was not compiled from the reviews."
        )
        tampered_path = self.write_review("tampered-operations.json", tampered)
        rejected = run_gate(self.export_path, package_dir, tampered_path)
        self.assertEqual("fail", rejected["status"])
        self.assertTrue(any("deterministic recompilation" in error for error in rejected["errors"]))

    def test_human_rows_and_workbook_are_compact_and_separate_from_change_log(self) -> None:
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")
        operational, configuration, architecture = self.completed_reviews()
        align_duplicate_operation(operational, architecture)
        payload, errors = compile_operations(
            operational,
            configuration,
            architecture,
            "Direct GTM/MCP/API",
            "Deep",
            source_object_catalog(self.export_path),
        )
        self.assertEqual([], errors)
        human, human_errors = build_rows(payload)
        self.assertEqual([], human_errors)
        self.assertEqual(6, len(human[0]))
        workbook_path = self.root / "cleanup-plan.xlsx"
        manifest = {"source_file": self.export_path.name, "source_sha256": payload["source_sha256"]}
        source = build_model(self.export_path)
        build_workbook(
            manifest,
            source,
            operational,
            configuration,
            architecture,
            payload,
            {"rows": human},
            workbook_path,
        )
        workbook = load_workbook(workbook_path)
        self.assertEqual(CANONICAL_SHEETS, workbook.sheetnames)
        self.assertEqual(
            ["01 Summary", "02 Cleanup Plan"],
            [s.title for s in workbook if s.sheet_state == "visible"],
        )
        for sheet in workbook:
            self.assertLessEqual(sheet.max_column, 6)
            self.assertLessEqual(
                max(sheet.column_dimensions[col].width or 0 for col in sheet.column_dimensions), 92
            )
            self.assertLessEqual(
                max(sheet.row_dimensions[row].height or 0 for row in sheet.row_dimensions), 120
            )
        self.assertNotIn("Change Log", workbook.sheetnames)
        gate_errors, gate_warnings = validate_workbook(
            workbook_path,
            self.write_review("workbook-operations.json", payload),
        )
        self.assertEqual([], gate_errors)
        self.assertEqual([], gate_warnings)

    def test_hidden_proof_is_split_losslessly_and_visible_text_is_not_truncated(self) -> None:
        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")
        proof = "source-bound-proof-" + ("x" * (MAX_CELL_TEXT + 41))
        workbook = Workbook()
        hidden = workbook.active
        hidden.title = "Hidden"
        add_table(hidden, [{"Evidence": proof}], ["Evidence"], split_long_cells=True)
        rendered = "".join(str(hidden.cell(row, 1).value or "") for row in range(2, hidden.max_row + 1))
        self.assertEqual(proof, rendered)
        self.assertNotIn("[truncated]", rendered.lower())

        visible = workbook.create_sheet("Visible")
        with self.assertRaisesRegex(ValueError, "summarize the user-facing row"):
            add_table(visible, [{"Problem": proof}], ["Problem"])

    def test_source_model_contains_nested_facts_and_dependency_edges(self) -> None:
        model = build_model(self.export_path)
        self.assertEqual("pass", model["coverage_gate"])
        self.assertGreater(model["counts"]["field_edges"], 0)
        self.assertGreater(model["counts"]["trigger_edges"], 0)

    def test_source_integrity_blocks_ambiguous_unmodelled_and_empty_sources(self) -> None:
        duplicate = sample_export()
        duplicate["containerVersion"]["tag"].append(
            {**copy.deepcopy(duplicate["containerVersion"]["tag"][0]), "name": "Duplicate ID"}
        )
        unmodelled = sample_export()
        unmodelled["containerVersion"]["mysteryEntity"] = [{"mysteryId": "1"}]
        unmodelled_empty = sample_export()
        unmodelled_empty["containerVersion"]["futureEntity"] = []
        cases = {
            "duplicate": (duplicate, "duplicate_entity_id"),
            "unmodelled": (unmodelled, "unmodelled_entity_layer"),
            "unmodelled-empty": (unmodelled_empty, "unmodelled_entity_layer"),
            "empty": ({}, "invalid_container_version_shape"),
            "container-resource": (
                {
                    "accountId": "100",
                    "containerId": "200",
                    "name": "This is a Container resource, not a ContainerVersion",
                    "publicId": "GTM-WRONG",
                    "usageContext": ["WEB"],
                },
                "invalid_container_version_shape",
            ),
            "empty-nested-container": (
                {"container": {}},
                "invalid_container_version_shape",
            ),
        }
        for name, (payload, finding_type) in cases.items():
            with self.subTest(name=name):
                export = self.root / f"{name}.json"
                export.write_text(json.dumps(payload), encoding="utf-8")
                model = build_model(export)
                self.assertEqual("blocked_source_integrity", model["coverage_gate"])
                self.assertIn(
                    finding_type,
                    {row["finding_type"] for row in model["source_integrity_findings"]},
                )
                package = self.root / f"{name}-package"
                manifest = build_package(export, package, pretty=True)
                self.assertEqual("blocked", manifest["status"])
                self.assertFalse((package / "configuration_review.json").exists())
                self.assertFalse((package / "architecture_review.json").exists())
                operational = audit_export(export)
                self.assertEqual("blocked_source_integrity", operational["run_status"])
                for scaffold in (
                    scaffold_operational,
                    scaffold_configuration,
                    scaffold_architecture,
                ):
                    with self.assertRaisesRegex(ValueError, "source integrity gate blocked"):
                        scaffold(export)
                with self.assertRaisesRegex(ValueError, "source integrity gate blocked"):
                    source_object_catalog(export)
                with self.assertRaisesRegex(ValueError, "source integrity gate blocked"):
                    build_context_model(export)
                with self.assertRaisesRegex(ValueError, "source integrity gate blocked"):
                    extract_export(export)
                with self.assertRaisesRegex(ValueError, "source integrity gate blocked"):
                    scan_relationships(export)
                future_report, future_errors = check_future_state(
                    export, {"operations": []}
                )
                self.assertEqual(
                    "blocked_source_integrity", future_report["status"]
                )
                self.assertTrue(future_errors)

    def test_duplicate_reference_names_are_never_resolved_by_fallback(self) -> None:
        data = sample_export()
        data["containerVersion"]["variable"].extend(
            [
                {
                    "variableId": "900",
                    "name": "Click URL",
                    "type": "c",
                    "parameter": [
                        {"type": "TEMPLATE", "key": "value", "value": "analytics_storage"}
                    ],
                },
                {
                    "variableId": "901",
                    "name": "Click URL",
                    "type": "c",
                    "parameter": [
                        {"type": "TEMPLATE", "key": "value", "value": "ad_storage"}
                    ],
                },
            ]
        )
        data["containerVersion"]["tag"][0]["parameter"].append(
            {
                "type": "TEMPLATE",
                "key": "consent_payload",
                "value": "{{Click URL}}",
            }
        )
        export = self.root / "duplicate-reference-names.json"
        export.write_text(json.dumps(data), encoding="utf-8")

        self.assertEqual("pass", build_model(export)["coverage_gate"])
        operational = audit_export(export)
        duplicate = next(
            row
            for row in operational["findings"]
            if row["module_name"] == "duplicate_variable_names"
            and row["finding_type"] != "zero_findings"
        )
        self.assertEqual({"900", "901"}, set(duplicate["object_ids"]))

        configuration = scaffold_configuration(export)
        tag_row = next(
            row for row in configuration["rows"] if row["object_key"] == "tag:1"
        )
        trace = next(
            row
            for row in tag_row["reference_trace_requirements"]
            if row["reference"] == "Click URL"
        )
        self.assertEqual(["ambiguous"], trace["terminal_states"])
        self.assertEqual(
            {"variable:900", "variable:901"}, set(trace["required_object_keys"])
        )
        self.assertIn(
            "builtInVariable:Click URL",
            trace["terminal_requirements"][0]["configured_source"],
        )

        architecture = scaffold_architecture(export)
        family = next(
            row for row in architecture["families"] if "tag:1" in row["member_object_keys"]
        )
        self.assertTrue(
            {"variable:900", "variable:901"}.issubset(family["chain_object_keys"])
        )

        consent = tag_consent_route(
            data["containerVersion"]["tag"][0],
            variables=data["containerVersion"]["variable"],
        )
        self.assertTrue(
            {"analytics_storage", "ad_storage"}.issubset(
                consent["detected_consent_payload_purposes"]
            )
        )
        self.assertEqual([], consent["forwarded_consent_purposes"])
        builtin = next(
            row
            for row in operational["lifecycle_matrix"]
            if row["object_key"] == "builtInVariable:Click URL"
        )
        self.assertEqual("used", builtin["usage_state"])

        completed = complete_configuration(export)
        errors, _ = validate_configuration(
            export,
            self.write_review("ambiguous-configuration.json", completed),
        )
        self.assertTrue(
            any("ambiguous variable identity" in error for error in errors), errors
        )

    def test_missing_javascript_parser_is_an_explicit_review_limit(self) -> None:
        with patch.dict(sys.modules, {"esprima": None}):
            review = complete_configuration(self.export_path)
            parser_requirements = [
                finding
                for row in review["rows"]
                for finding in row["required_technical_findings"]
                if finding["category"] == "parser"
            ]
            self.assertTrue(parser_requirements)
            for row in review["rows"]:
                for finding in row["technical_finding_reviews"]:
                    if finding["finding_key"] == "parser:coverage":
                        finding["verdict"] = "False positive"
            review_path = self.write_review("parser-limit.json", review)
            errors, _ = validate_configuration(self.export_path, review_path)
            self.assertTrue(any("parser coverage limit" in error for error in errors))

            for row in review["rows"]:
                for finding in row["technical_finding_reviews"]:
                    if finding["finding_key"] != "parser:coverage":
                        continue
                    finding["verdict"] = "Documented exception"
                    finding["rationale"] = (
                        "The mandatory line-by-line code blocks cover every exported line; "
                        "the parser availability boundary is recorded without claiming AST coverage."
                    )
            review_path = self.write_review("parser-limit-resolved.json", review)
            resolved_errors, _ = validate_configuration(self.export_path, review_path)
            self.assertEqual([], resolved_errors)

    def test_direct_container_version_paths_compile_and_apply_without_envelope(self) -> None:
        direct = copy.deepcopy(sample_export()["containerVersion"])
        export = self.root / "direct-container-version.json"
        export.write_text(json.dumps(direct), encoding="utf-8")
        model = build_model(export)
        self.assertEqual("pass", model["coverage_gate"])
        self.assertTrue(
            all(
                not row["json_path"].startswith("$.containerVersion")
                for row in model["field_edges"]
            )
        )
        configuration = scaffold_configuration(export)
        tag_review = next(row for row in configuration["rows"] if row["object_key"] == "tag:1")
        self.assertEqual("$.tag[0]", tag_review["source_json_path"])
        future, errors = apply_operations(
            direct,
            {
                "operations": [
                    {
                        "operation_key": "direct-path-change",
                        "creations": [],
                        "additions": [],
                        "changes": [
                            {
                                "object_key": "tag:1",
                                "json_path": "$.tag[0].parameter[0].value",
                                "before": "purchase",
                                "after": "purchase_complete",
                            }
                        ],
                        "remaps": [],
                        "deletions": [],
                        "renames": [],
                    }
                ]
            },
        )
        self.assertEqual([], errors)
        self.assertEqual("purchase_complete", future["tag"][0]["parameter"][0]["value"])

    def test_zones_and_google_tag_configs_are_reviewed_by_all_three_runs(self) -> None:
        data = sample_export()
        data["containerVersion"]["zone"] = [
            {
                "zoneId": "70",
                "name": "Partner FR child container",
                "childContainer": [{"publicId": "GTM-CHILD", "nickname": "Partner"}],
                "boundary": {
                    "condition": [condition("EQUALS", "partner", "partner")],
                    "customEvaluationTriggerId": ["10"],
                },
                "typeRestriction": {"enable": True, "whitelistedTypeId": ["html"]},
            }
        ]
        data["containerVersion"]["gtagConfig"] = [
            {
                "gtagConfigId": "71",
                "parameter": [
                    {"type": "TEMPLATE", "key": "tag_id", "value": "G-TEST123"},
                    {
                        "type": "TEMPLATE",
                        "key": "server_container_url",
                        "value": "https://collect.example.test",
                    },
                ],
            }
        ]
        export = self.root / "zone-and-gtag.json"
        export.write_text(json.dumps(data), encoding="utf-8")

        source = build_model(export)
        self.assertEqual(1, source["counts"]["zones"])
        self.assertEqual(1, source["counts"]["gtagConfigs"])
        self.assertTrue(
            any(edge["relation"] == "zone_boundary_trigger" for edge in source["trigger_edges"])
        )

        operational = audit_export(export)
        self.assertEqual(1, operational["counts"]["zones"])
        self.assertEqual(1, operational["counts"]["gtagConfigs"])
        self.assertEqual(
            {"zone:70", "gtagConfig:71"},
            {
                row["object_key"]
                for row in operational["lifecycle_matrix"]
                if row["layer"] in {"zone", "gtagConfig"}
            },
        )

        configuration = scaffold_configuration(export)
        configuration_keys = {row["object_key"] for row in configuration["rows"]}
        self.assertTrue({"zone:70", "gtagConfig:71"}.issubset(configuration_keys))
        self.assertIn(
            "collect.example.test",
            configuration["audit_context"]["server_routing_hosts"],
        )
        self.assertIn("FR", configuration["audit_context"]["markets"])
        for key in ("zone:70", "gtagConfig:71"):
            row = next(item for item in configuration["rows"] if item["object_key"] == key)
            self.assertTrue(row["required_contract_topics"])
            self.assertTrue(row["required_branch_reviews"])
        gtag_row = next(
            item for item in configuration["rows"] if item["object_key"] == "gtagConfig:71"
        )
        self.assertTrue(
            any(
                fact["json_path"].endswith(".type")
                and fact["value_type"] == "missing"
                for fact in gtag_row["source_absence_facts"]
            )
        )
        self.assertNotIn(
            "Unclassified external integration (collect.example.test)",
            {item["vendor"] for item in gtag_row["vendor_contexts"]},
        )

        architecture = scaffold_architecture(export)
        root_keys = {
            key
            for family in architecture["families"]
            for key in family["member_object_keys"]
        }
        self.assertTrue({"zone:70", "gtagConfig:71"}.issubset(root_keys))
        destination_comparison = next(
            row
            for row in architecture["comparisons"]
            if "shared_configured_destination" in row["comparison_types"]
            and {"tag:1", "gtagConfig:71"}.issubset(row["candidate_object_keys"])
        )
        self.assertIn(
            "consumer_destination_and_event_overlap",
            destination_comparison["discovery_methods"],
        )

        two_zone_cv = copy.deepcopy(data["containerVersion"])
        two_zone_cv["zone"].append(
            {
                "zoneId": "72",
                "name": "Partner child container alternate boundary",
                "childContainer": [{"publicId": "GTM-CHILD", "nickname": "Partner"}],
                "boundary": {"customEvaluationTriggerId": ["11"]},
            }
        )
        zone_comparison = next(
            row
            for row in relationship_candidates(two_zone_cv)
            if "shared_zone_child_container" in row["comparison_types"]
        )
        self.assertEqual(
            {"zone:70", "zone:72"}, set(zone_comparison["candidate_object_keys"])
        )

        complete_configuration_review = complete_configuration(export)
        configuration_errors, _ = validate_configuration(
            export,
            self.write_review("zone-gtag-configuration.json", complete_configuration_review),
        )
        self.assertEqual([], configuration_errors)
        complete_architecture_review = complete_architecture(export)
        architecture_errors, _ = validate_architecture(
            export,
            self.write_review("zone-gtag-architecture.json", complete_architecture_review),
        )
        self.assertEqual([], architecture_errors)

    def test_run1_detects_same_tag_payload_across_different_execution_routes(self) -> None:
        data = sample_export()
        original = data["containerVersion"]["tag"][0]
        alternate = copy.deepcopy(original)
        alternate.update(
            {
                "tagId": "99",
                "name": "GA4 purchase alternate control route",
                "firingTriggerId": ["11"],
                "blockingTriggerId": [],
                "setupTag": [],
                "teardownTag": [{"tagName": "Utility - Consent Defaults"}],
                "tagFiringOption": "ONCE_PER_EVENT",
                "priority": {"type": "INTEGER", "value": "7"},
                "liveOnly": True,
                "paused": True,
                "scheduleStartMs": "100",
                "scheduleEndMs": "200",
                "consentSettings": {"consentStatus": "needed"},
            }
        )
        data["containerVersion"]["tag"].append(alternate)
        export = self.root / "same-payload-different-route.json"
        export.write_text(json.dumps(data), encoding="utf-8")

        candidates = [
            row
            for row in audit_export(export)["findings"]
            if row["finding_type"] == "normalized_duplicate_tag_signature"
        ]
        self.assertTrue(
            any({"1", "99"}.issubset(set(row["object_ids"])) for row in candidates)
        )

    def test_run1_queues_same_contract_different_consent_controls(self) -> None:
        data = sample_export()
        alternate = copy.deepcopy(data["containerVersion"]["tag"][0])
        alternate.update(
            {
                "tagId": "99",
                "name": "GA4 purchase explicit consent route",
                "blockingTriggerId": [],
                "consentSettings": {"consentStatus": "needed"},
            }
        )
        data["containerVersion"]["tag"].append(alternate)
        export = self.root / "same-contract-consent-collision.json"
        export.write_text(json.dumps(data), encoding="utf-8")

        collision = next(
            row
            for row in audit_export(export)["findings"]
            if row["finding_type"]
            == "same_contract_different_consent_control_candidate"
        )
        self.assertEqual({"1", "99"}, set(collision["object_ids"]))
        self.assertEqual(
            ["purchase"],
            collision["shared_event_destination_contract"]["events"],
        )
        self.assertEqual(2, len(collision["consent_control_comparison"]))

    def test_consumed_object_deletion_requires_complete_remap_coverage(self) -> None:
        operation = {
            "creations": [],
            "additions": [],
            "changes": [],
            "remaps": [],
            "renames": [],
            "deletions": [
                {"object_key": "trigger:10", "reason": "Remove redundant trigger logic."}
            ],
            "minimum_aggressiveness": "Deep",
        }
        keys = {"trigger:10", "trigger:11", "tag:1"}
        consumers = {"trigger:10": {"tag:1"}, "trigger:11": set(), "tag:1": set()}
        errors = validate_structured_actions(operation, keys, "deletion test", consumers)
        self.assertTrue(any("requires remap coverage" in error for error in errors))

        operation["remaps"] = [
            {
                "from_object_key": "trigger:10",
                "to_object_key": "trigger:11",
                "consumer_object_keys": ["tag:1"],
            }
        ]
        self.assertEqual(
            [], validate_structured_actions(operation, keys, "deletion test", consumers)
        )

    def test_remaps_reject_cross_layer_deleted_targets_and_created_cycles(self) -> None:
        keys = {"trigger:12", "trigger:13", "tag:1"}
        consumers = {
            "trigger:12": {"trigger:13"},
            "trigger:13": {"trigger:12"},
            "tag:1": set(),
        }

        def operation(target: str, deletions: list[dict] | None = None) -> dict:
            return {
                "creations": [],
                "additions": [],
                "changes": [],
                "remaps": [
                    {
                        "from_object_key": "trigger:13",
                        "to_object_key": target,
                        "consumer_object_keys": ["trigger:12"],
                    }
                ],
                "renames": [],
                "deletions": deletions or [],
                "minimum_aggressiveness": "Deep",
            }

        cross_layer = validate_structured_actions(
            operation("tag:1"), keys, "cross-layer test", consumers
        )
        self.assertTrue(any("crosses GTM layers" in error for error in cross_layer))

        deleted_target = validate_structured_actions(
            operation(
                "trigger:12",
                [
                    {
                        "object_key": "trigger:12",
                        "reason": "Delete the obsolete target after consolidation.",
                    }
                ],
            ),
            keys,
            "deleted-target test",
            consumers,
        )
        self.assertTrue(any("remap target 'trigger:12' is also deleted" in error for error in deleted_target))

        cycle = validate_structured_actions(
            operation("trigger:12"), keys, "cycle test", consumers
        )
        self.assertTrue(any("creates a dependency cycle" in error for error in cycle))

    def test_operational_operation_set_rejects_duplicate_final_names(self) -> None:
        operation = duplicate_variable_operation()
        operation["renames"] = [
            {
                "object_key": "tag:1",
                "before": "GA4 - Purchase - All",
                "after": "Meta - Purchase - All",
            }
        ]
        errors = validate_operation_set(
            [operation],
            object_keys(self.export_path),
            object_consumer_map(self.export_path),
            object_name_map(self.export_path),
            "operational operation set",
        )
        self.assertTrue(any("duplicate final name" in error for error in errors))

        review = complete_operational(self.export_path)
        finding = next(
            row
            for row in review["findings"]
            if row["module_name"] == "duplicate_variable_paths"
            and set(row.get("object_ids", [])) == {"20", "21"}
        )
        finding.update(operation)
        finding["disposition"] = "cleanup_operation"
        finding["rationale"] = (
            "Variables 20 and 21 share ecommerce.items while tag 1 is proposed for an "
            "explicit final-name change in this operation."
        )
        review_errors, _ = validate_operational(
            self.export_path,
            self.write_review("duplicate-final-name-operational.json", review),
        )
        self.assertTrue(any("duplicate final name" in error for error in review_errors))

    def test_run2_rejects_conclusions_that_contradict_decisive_source_states(self) -> None:
        data = sample_export()
        data["containerVersion"]["tag"].append(
            {
                "tagId": "5",
                "name": "Marketing bundle",
                "type": "html",
                "parameter": [
                    {
                        "type": "TEMPLATE",
                        "key": "html",
                        "value": (
                            "<script>fbq('track','Purchase',{value:{{DLV - Value}}});"
                            "ttq.track('CompletePayment',{value:{{DLV - Value}}});"
                            "var s=document.createElement('script');"
                            "s.src='https://metrics.example.test/p.js';"
                            "document.head.appendChild(s);</script>"
                        ),
                    }
                ],
                "firingTriggerId": ["10"],
            }
        )
        data["containerVersion"]["zone"] = [
            {
                "zoneId": "70",
                "name": "Partner boundary",
                "childContainer": [
                    {"publicId": "GTM-CHILD"},
                    {"publicId": "GTM-CHILD"},
                    {},
                ],
                "boundary": {
                    "customEvaluationTriggerId": ["10", "999"],
                    "condition": "malformed",
                },
                "typeRestriction": {"enable": True, "whitelistedTypeId": []},
            }
        ]
        data["containerVersion"]["gtagConfig"] = [
            {
                "gtagConfigId": "80",
                "name": "Primary Google destination",
                "parameter": [
                    {"type": "TEMPLATE", "key": "measurementId", "value": "G-TEST123"},
                    {
                        "type": "TEMPLATE",
                        "key": "server_container_url",
                        "value": "https://collect.example.test",
                    },
                ],
            }
        ]
        data["containerVersion"]["customTemplate"] = [
            {
                "templateId": "90",
                "name": "Opaque template",
                "templateData": '{"__wm":"TEMPLATE","permissions":[]}',
            }
        ]
        export = self.root / "decisive-run2-states.json"
        export.write_text(json.dumps(data), encoding="utf-8")

        honest = complete_configuration(export)
        honest_errors, _ = validate_configuration(
            export,
            self.write_review("decisive-run2-honest.json", honest),
        )
        self.assertEqual([], honest_errors)
        gtag_row = next(
            item for item in honest["rows"] if item["object_key"] == "gtagConfig:80"
        )
        self.assertTrue(
            {"tag:1"}.issubset(
                {peer["object_key"] for peer in gtag_row["destination_peer_contexts"]}
            )
        )
        tag_row = next(
            item for item in honest["rows"] if item["object_key"] == "tag:1"
        )
        gtag_peer = next(
            peer
            for peer in tag_row["destination_peer_contexts"]
            if peer["object_key"] == "gtagConfig:80"
        )
        self.assertEqual(["collect.example.test"], gtag_peer["server_routing_hosts"])
        self.assertFalse(gtag_peer["type_present"])
        self.assertTrue(
            any(
                fact["json_path"].endswith(".type")
                and fact["value_type"] == "missing"
                for fact in tag_row["destination_peer_facts"]
            )
        )
        self.assertTrue(
            any(
                item["obligation_key"].startswith(
                    "peer_destination_contract_unproven:"
                )
                for item in tag_row["required_configuration_obligations"]
            )
        )
        value_row = next(
            item for item in honest["rows"] if item["object_key"] == "variable:24"
        )
        self.assertTrue(
            any(
                context["consumer_key"] == "tag:5"
                and "purchase" in context["events"]
                for context in value_row["consumer_dependency_contexts"]
            )
        )
        opaque_row = next(
            item
            for item in honest["rows"]
            if item["object_key"] == "customTemplate:90"
        )
        self.assertEqual(
            "unknown_opaque",
            opaque_row["technical_code_facts"]["returned_value_type"],
        )
        self.assertFalse(
            any(
                "no material external behavior limit" in value.lower()
                for value in opaque_row["technical_code_facts"][
                    "container_evidence_limits"
                ]
            )
        )

        poisoned = copy.deepcopy(honest)
        for key in (
            "tag:1",
            "tag:5",
            "zone:70",
            "gtagConfig:80",
            "customTemplate:90",
        ):
            row = next(item for item in poisoned["rows"] if item["object_key"] == key)
            row.update(
                {
                    "correctness_verdict": "Correct",
                    "correctness_basis": object_specific_text(
                        row,
                        "is incorrectly claimed to have no material configuration defect",
                        "correctness_basis",
                    ),
                    "defects": [],
                    "disposition": "keep",
                    "owner_question": "",
                }
            )
            for branch in row["configuration_branch_reviews"]:
                branch["correctness"] = "Correct"
            for check in row["logic_cross_checks"]:
                check["verdict"] = "Aligned"
            for check in row["contract_checks"]:
                check["verdict"] = "Compliant"
                if not check.get("source"):
                    check["source"] = "https://docs.example.test/official"
            for finding in row["technical_finding_reviews"]:
                if "no reviewable executable behavior" in finding[
                    "source_statement"
                ].lower():
                    finding["verdict"] = "False positive"

        errors, _ = validate_configuration(
            export,
            self.write_review("decisive-run2-poisoned.json", poisoned),
        )
        self.assertTrue(any("invalid_zone_boundary_field:condition" in error for error in errors))
        self.assertTrue(any("missing_required_field:type" in error for error in errors))
        self.assertTrue(any("opaque_custom_template_behavior" in error for error in errors))
        self.assertTrue(any("peer_destination_contract_unproven" in error for error in errors))
        self.assertTrue(any("unsupported value" in error for error in errors))
        self.assertTrue(any("reserved or non-production" in error for error in errors))

    def test_parser_fallback_requires_exact_segment_attestation(self) -> None:
        review = complete_configuration(self.export_path)
        row = next(item for item in review["rows"] if item["object_key"] == "variable:22")
        parser_review = next(
            item
            for item in row["technical_finding_reviews"]
            if item["finding_key"] == "parser:coverage"
        )
        parser_review["fallback_line_hashes"] = parser_review["fallback_line_hashes"][:-1]
        parser_review["manual_review_method"] = (
            "A generic line-by-line parser review was performed for this code."
        )
        errors, _ = validate_configuration(
            self.export_path,
            self.write_review("weak-parser-fallback.json", review),
        )
        self.assertTrue(any("every exported code segment hash" in error for error in errors))
        self.assertTrue(any("source-specific code behavior" in error for error in errors))

        segment_poisoned = complete_configuration(self.export_path)
        segment_row = next(
            item
            for item in segment_poisoned["rows"]
            if item["object_key"] == "variable:22"
        )
        segment_parser_review = next(
            item
            for item in segment_row["technical_finding_reviews"]
            if item["finding_key"] == "parser:coverage"
        )
        segment_parser_review["fallback_segment_reviews"][1]["behavior"] = (
            "This remaining code segment received a generic manual inspection only."
        )
        errors, _ = validate_configuration(
            self.export_path,
            self.write_review("generic-parser-segment.json", segment_poisoned),
        )
        self.assertTrue(any("parser fallback segment" in error for error in errors))

    def test_run2_rejects_code_semantic_reversal_and_risk_dismissal(self) -> None:
        data = sample_export()
        data["containerVersion"]["tag"][1]["parameter"][0]["value"] = (
            "<script>fbq('track', 'Purchase');</script>"
        )
        export = self.root / "semantic-code-reversal.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        review = complete_configuration(export)
        code_row = next(
            row for row in review["rows"] if row["object_key"] == "tag:2"
        )
        code_row["code_behavior_blocks"][0]["health_assessment"] = (
            "The exact fbq identifier is inspected, but Meta does not track or send the "
            "Purchase event from this segment."
        )
        parser_review = next(
            item
            for item in code_row["technical_finding_reviews"]
            if item["finding_key"] == "parser:coverage"
        )
        parser_review["fallback_segment_reviews"][0]["behavior"] = (
            "Mandatory line-by-line review cites fbq and Purchase, but states that Meta "
            "does not track or send the Purchase event from this exact source segment."
        )
        risk_key = next(
            item["finding_key"]
            for item in code_row["required_technical_findings"]
            if item["category"] in {"health", "security"}
        )
        risk_review = next(
            item
            for item in code_row["technical_finding_reviews"]
            if item["finding_key"] == risk_key
        )
        risk_review["verdict"] = "False positive"

        errors, _ = validate_configuration(
            export,
            self.write_review("semantic-code-reversal-review.json", review),
        )
        self.assertTrue(
            any("reverses source-proven segment behavior meta_event_send" in error for error in errors)
        )
        self.assertTrue(
            any("cannot be dismissed as a false positive" in error for error in errors)
        )
        dead_path_review = complete_configuration(export)
        dead_path_row = next(
            row
            for row in dead_path_review["rows"]
            if row["object_key"] == "tag:2"
        )
        dead_path_row["code_behavior_blocks"][0]["health_assessment"] = (
            "The source contains fbq track Purchase, but this is a dead code path with "
            "zero delivery."
        )
        dead_path_parser = next(
            item
            for item in dead_path_row["technical_finding_reviews"]
            if item["finding_key"] == "parser:coverage"
        )
        dead_path_parser["fallback_segment_reviews"][0]["behavior"] = (
            "Mandatory line-by-line inspection identifies fbq track Purchase, but calls "
            "the segment unreachable with no delivery."
        )
        dead_path_errors, _ = validate_configuration(
            export,
            self.write_review("dead-path-code-reversal.json", dead_path_review),
        )
        self.assertTrue(
            any("denies the executable effect" in error for error in dead_path_errors)
        )
        alias_expectations = {
            "Cleanup opportunity": "requires one concrete proposed_action",
            "Documented exception": "requires an evidence-bound exception_basis",
            "Owner decision needed": "requires a source-specific owner_question",
        }
        for verdict, expected_error in alias_expectations.items():
            with self.subTest(verdict=verdict):
                alias_review = complete_configuration(export)
                alias_row = next(
                    row
                    for row in alias_review["rows"]
                    if row["object_key"] == "tag:2"
                )
                alias_risk = next(
                    item
                    for item in alias_row["technical_finding_reviews"]
                    if item["finding_key"] == risk_key
                )
                alias_risk.update(
                    {
                        "verdict": verdict,
                        "rationale": (
                            "The source signal receives this generic disposition for the "
                            "current container."
                        ),
                        "proposed_action": "",
                        "exception_basis": "",
                        "owner_question": "",
                    }
                )
                alias_row["owner_question"] = ""
                alias_errors, _ = validate_configuration(
                    export,
                    self.write_review(
                        f"semantic-risk-alias-{verdict.lower().replace(' ', '-')}.json",
                        alias_review,
                    ),
                )
                self.assertTrue(
                    any(expected_error in error for error in alias_errors), alias_errors
                )

        confirmed_review = complete_configuration(export)
        confirmed_row = next(
            row
            for row in confirmed_review["rows"]
            if row["object_key"] == "tag:2"
        )
        confirmed_risk = next(
            item
            for item in confirmed_row["technical_finding_reviews"]
            if item["finding_key"] == risk_key
        )
        confirmed_risk["verdict"] = "Confirmed issue"
        confirmed_row.update(
            {
                "correctness_verdict": "Issue",
                "correctness_basis": object_specific_text(
                    confirmed_row,
                    "contains the confirmed custom-code risk and requires owner resolution",
                    "correctness_basis",
                ),
                "disposition": "owner_decision_needed",
                "owner_question": (
                    "Which owner will resolve the confirmed inline-script risk before this "
                    "tag remains active?"
                ),
                "defects": [
                    {
                        "defect_id": "TECH-001",
                        "statement": confirmed_risk["source_statement"],
                        "configured_effect": (
                            "The exact exported custom-code signal creates the confirmed "
                            "maintenance or security exposure."
                        ),
                        "expected_behavior": (
                            "The confirmed code risk must be corrected or accepted through a "
                            "source-bound owner decision."
                        ),
                        "evidence_anchors": [],
                        "code_line_hashes": [
                            confirmed_row["required_code_line_hashes"][0]
                        ],
                        "technical_finding_keys": [],
                    }
                ],
            }
        )
        confirmed_errors, _ = validate_configuration(
            export,
            self.write_review("unlinked-confirmed-technical-risk.json", confirmed_review),
        )
        self.assertTrue(
            any(
                f"confirmed technical issue {risk_key} must link" in error
                for error in confirmed_errors
            )
        )

    def test_run2_binds_malformed_controls_and_contradictions_to_every_verdict_layer(
        self,
    ) -> None:
        data = sample_export()
        data["containerVersion"]["tag"][0]["teardownTag"] = [{}, {}]
        data["containerVersion"]["trigger"].extend(
            [
                {
                    "triggerId": "915",
                    "name": "Impossible lead or signup conjunction",
                    "type": "CUSTOM_EVENT",
                    "filter": [
                        condition("EQUALS", "{{_event}}", "lead"),
                        condition("EQUALS", "{{_event}}", "signup"),
                    ],
                },
                {
                    "triggerId": "916",
                    "name": "Malformed trigger group members",
                    "type": "TRIGGER_GROUP",
                    "parameter": [
                        {
                            "type": "LIST",
                            "key": "triggerIds",
                            "list": ["10", {"type": "TEMPLATE", "value": "10"}, {}],
                        }
                    ],
                },
            ]
        )
        export = self.root / "malformed-controls-and-contradiction.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        review = complete_configuration(export)
        honest_errors, _ = validate_configuration(
            export, self.write_review("malformed-controls-honest.json", review)
        )
        self.assertEqual([], honest_errors)

        tag = next(
            row
            for row in reversed(review["rows"])
            if row["object_key"] == "tag:1"
        )
        malformed_sequences = [
            item
            for item in tag["required_configuration_obligations"]
            if item["obligation_key"].startswith("dependency:teardownTag:")
        ]
        self.assertEqual(2, len(malformed_sequences))
        group = next(
            row for row in review["rows"] if row["object_key"] == "trigger:916"
        )
        self.assertEqual(
            2,
            sum(
                item["obligation_key"].startswith("invalid_trigger_group_member:")
                for item in group["required_configuration_obligations"]
            ),
        )
        malformed_traces = [
            trace
            for trace in group["execution_dependency_traces"]
            if trace["resolution_state"] == "malformed"
        ]
        self.assertEqual(2, len(malformed_traces))
        self.assertEqual(
            {
                f"{group['source_json_path']}.parameter[0].list[0]",
                f"{group['source_json_path']}.parameter[0].list[2]",
            },
            {
                trace["source_reference_paths"][0]
                for trace in malformed_traces
            },
        )
        valid_trace = next(
            trace
            for trace in group["execution_dependency_traces"]
            if trace["reference"] == "10" and trace["resolution_state"] == "unique"
        )
        self.assertEqual(
            [f"{group['source_json_path']}.parameter[0].list[1].value"],
            valid_trace["source_reference_paths"],
        )
        contradictory = next(
            row for row in review["rows"] if row["object_key"] == "trigger:915"
        )
        self.assertTrue(
            any(
                item["obligation_key"].startswith("contradictory_equals:")
                for item in contradictory["required_configuration_obligations"]
            )
        )

        contradictory.update(
            {
                "correctness_verdict": "Correct",
                "disposition": "keep",
                "defects": [],
                "owner_question": "",
            }
        )
        for branch in contradictory["configuration_branch_reviews"]:
            branch["correctness"] = "Correct"
        for check in contradictory["logic_cross_checks"]:
            check["verdict"] = "Aligned"
            check["conclusion"] = (
                "The exported purpose and execution values were generally reviewed and "
                "declared aligned for this object."
            )
        errors, _ = validate_configuration(
            export, self.write_review("contradiction-fail-open.json", review)
        )
        self.assertTrue(any("contradictory_equals:" in error for error in errors))
        self.assertTrue(any("does not name deterministic obligation" in error for error in errors))

    def test_run2_rejects_duplicate_rows_branches_checks_findings_and_traces(self) -> None:
        review = complete_configuration(self.export_path)
        review["rows"].append(copy.deepcopy(review["rows"][0]))
        tag = next(
            row
            for row in reversed(review["rows"])
            if row["object_key"] == "tag:1"
        )
        tag["configuration_branch_reviews"].append(
            copy.deepcopy(tag["configuration_branch_reviews"][0])
        )
        tag["logic_cross_checks"].append(copy.deepcopy(tag["logic_cross_checks"][0]))
        tag["contract_checks"].append(copy.deepcopy(tag["contract_checks"][0]))
        if tag["reference_traces"]:
            tag["reference_traces"].append(copy.deepcopy(tag["reference_traces"][0]))
        code_row = next(
            row
            for row in review["rows"]
            if row["technical_finding_reviews"]
        )
        code_row["technical_finding_reviews"].append(
            copy.deepcopy(code_row["technical_finding_reviews"][0])
        )
        errors, _ = validate_configuration(
            self.export_path,
            self.write_review("duplicate-review-identities.json", review),
        )
        self.assertTrue(any("unique nonblank object keys" in error for error in errors))
        self.assertTrue(any("branch review paths must be unique" in error for error in errors))
        self.assertTrue(any("D3 logic check keys must be unique" in error for error in errors))
        self.assertTrue(any("unique nonblank topic keys" in error for error in errors))
        self.assertTrue(any("technical finding keys must be unique" in error for error in errors))
        if tag["reference_traces"]:
            self.assertTrue(any("trace references must be unique" in error for error in errors))

    def test_dynamic_script_load_is_network_behavior_with_full_endpoint_evidence(self) -> None:
        data = sample_export()
        data["containerVersion"]["tag"][1]["parameter"][0]["value"] = (
            "<script>var s=document.createElement('script');"
            "s.src='https://metrics.example.test/p.js';"
            "document.head.appendChild(s);</script>"
        )
        export = self.root / "dynamic-script-network.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        technical = next(
            row
            for row in extract_export(export)["rows"]
            if row["layer"] == "tag" and row["object_id"] == "2"
        )
        configuration = next(
            row
            for row in scaffold_configuration(export)["rows"]
            if row["object_key"] == "tag:2"
        )
        self.assertTrue(technical["network_calls"])
        self.assertIn("https://metrics.example.test/p.js", technical["external_scripts_loaded"])
        self.assertGreater(len(configuration["code_line_facts"]), 1)
        self.assertTrue(
            any(
                "metrics.example.test" in token
                for token in configuration["specificity_tokens"]
            )
        )

    def test_run3_requires_owner_or_action_for_unsafe_retention_candidates(self) -> None:
        data = sample_export()
        alternate = copy.deepcopy(data["containerVersion"]["tag"][0])
        alternate.update(
            {
                "tagId": "99",
                "name": "GA4 purchase alternate route",
                "firingTriggerId": ["11"],
                "blockingTriggerId": [],
                "setupTag": [],
                "consentSettings": {"consentStatus": "needed"},
            }
        )
        data["containerVersion"]["tag"].append(alternate)
        export = self.root / "unsafe-retention.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        review = complete_architecture(export)
        comparison = next(
            row
            for row in review["comparisons"]
            if "same_tag_payload_different_route" in row["comparison_types"]
            and {"tag:1", "tag:99"}.issubset(row["candidate_object_keys"])
        )
        comparison.update(
            {
                "relationship_verdict": "Intentional variant",
                "disposition": "keep",
                "owner_question": "",
            }
        )
        family = next(
            row
            for row in review["families"]
            if {"tag:1", "tag:99"}.issubset(row["member_object_keys"])
        )
        family.update(
            {
                "relationship_verdict": "Complementary",
                "disposition": "keep",
                "owner_question": "",
            }
        )
        errors, _ = validate_architecture(
            export,
            self.write_review("unsafe-retention-review.json", review),
        )
        self.assertTrue(any("not a source-proven intentional variant" in error for error in errors))
        self.assertTrue(any("family retention is unsupported" in error for error in errors))

    def test_run3_binds_cleanup_actions_and_discoveries_to_unsafe_candidates(self) -> None:
        data = sample_export()
        alternate = copy.deepcopy(data["containerVersion"]["tag"][0])
        alternate.update(
            {
                "tagId": "99",
                "name": "GA4 purchase alternate route",
                "firingTriggerId": ["11"],
                "blockingTriggerId": [],
                "consentSettings": {"consentStatus": "needed"},
            }
        )
        data["containerVersion"]["tag"].append(alternate)
        export = self.root / "architecture-operation-binding.json"
        export.write_text(json.dumps(data), encoding="utf-8")

        operation_review = complete_architecture(export)
        comparison = next(
            row
            for row in operation_review["comparisons"]
            if "same_tag_payload_different_route" in row["comparison_types"]
            and {"tag:1", "tag:99"}.issubset(row["candidate_object_keys"])
        )
        unrelated_operation = duplicate_variable_operation()
        unrelated_operation.update(
            {
                "operation_key": "unrelated-name-only-change",
                "canonical_object_key": "",
                "changes": [],
                "remaps": [],
                "deletions": [],
                "renames": [
                    {
                        "object_key": "tag:2",
                        "before": "Meta - Purchase - All",
                        "after": "Meta - Purchase - Primary",
                    }
                ],
            }
        )
        comparison.update(
            {
                "relationship_verdict": "Functional overlap",
                "disposition": "cleanup_operation",
                "owner_question": "",
                "operations": [unrelated_operation],
            }
        )
        errors, _ = validate_architecture(
            export,
            self.write_review("unrelated-architecture-operation.json", operation_review),
        )
        self.assertTrue(
            any("do not change any candidate member's behavior" in error for error in errors)
        )

        no_op_review = complete_architecture(export)
        no_op_comparison = next(
            row
            for row in no_op_review["comparisons"]
            if "same_tag_payload_different_route" in row["comparison_types"]
            and {"tag:1", "tag:99"}.issubset(row["candidate_object_keys"])
        )
        no_op_operation = copy.deepcopy(unrelated_operation)
        no_op_operation.update(
            {
                "operation_key": "candidate-no-op",
                "changes": [
                    {
                        "object_key": "tag:1",
                        "json_path": "$.containerVersion.tag[0].paused",
                        "before": False,
                        "after": False,
                    }
                ],
                "renames": [],
            }
        )
        no_op_comparison.update(
            {
                "relationship_verdict": "Functional overlap",
                "disposition": "cleanup_operation",
                "owner_question": "",
                "operations": [no_op_operation],
            }
        )
        errors, _ = validate_architecture(
            export,
            self.write_review("no-op-architecture-operation.json", no_op_review),
        )
        self.assertTrue(any("before and after values are identical" in error for error in errors))
        self.assertTrue(
            any("do not change any candidate member's behavior" in error for error in errors)
        )

        mismatched_path_review = complete_architecture(export)
        mismatched_comparison = next(
            row
            for row in mismatched_path_review["comparisons"]
            if "same_tag_payload_different_route" in row["comparison_types"]
            and {"tag:1", "tag:99"}.issubset(row["candidate_object_keys"])
        )
        mismatched_operation = copy.deepcopy(unrelated_operation)
        mismatched_operation.update(
            {
                "operation_key": "candidate-key-path-mismatch",
                "changes": [
                    {
                        "object_key": "tag:1",
                        "json_path": "$.containerVersion.tag[1].paused",
                        "before": False,
                        "after": True,
                    }
                ],
                "renames": [],
            }
        )
        mismatched_comparison.update(
            {
                "relationship_verdict": "Functional overlap",
                "disposition": "cleanup_operation",
                "owner_question": "",
                "operations": [mismatched_operation],
            }
        )
        errors, _ = validate_architecture(
            export,
            self.write_review(
                "mismatched-path-architecture-operation.json", mismatched_path_review
            ),
        )
        self.assertTrue(any("paired with an unrelated source path" in error for error in errors))
        self.assertTrue(
            any("do not change any candidate member's behavior" in error for error in errors)
        )

        discovery_review = complete_architecture(export)
        deterministic = next(
            row
            for row in discovery_review["comparisons"]
            if "same_tag_payload_different_route" in row["comparison_types"]
            and {"tag:1", "tag:99"}.issubset(row["candidate_object_keys"])
        )
        discovered = copy.deepcopy(deterministic)
        discovered.update(
            {
                "comparison_id": "DISC-UNSAFE-001",
                "comparison_origin": "analyst_discovered",
                "discovery_methods": ["normalized_condition_and_route_variants"],
                "relationship_verdict": "Intentional variant",
                "disposition": "keep",
                "owner_question": "",
                "operations": [],
            }
        )
        discovery_review["comparisons"].append(discovered)
        attestation = discovery_review["open_discovery_attestation"]
        attestation["discovered_comparison_ids"] = ["DISC-UNSAFE-001"]
        attestation["zero_discovery_rationale"] = ""
        method_review = next(
            item
            for item in attestation["method_reviews"]
            if item["method"] == "normalized_condition_and_route_variants"
        )
        method_review["additional_discovery_ids"] = ["DISC-UNSAFE-001"]
        errors, _ = validate_architecture(
            export,
            self.write_review("unsafe-discovered-retention.json", discovery_review),
        )
        self.assertTrue(
            any("not a source-proven intentional variant" in error for error in errors)
        )

        method_review["additional_discovery_ids"] = []
        errors, _ = validate_architecture(
            export,
            self.write_review("unattributed-discovery.json", discovery_review),
        )
        self.assertTrue(any("do not match the discoveries attributed" in error for error in errors))

    def test_run3_unsafe_discovery_inherits_policy_for_candidate_subsets(self) -> None:
        data = sample_export()
        data["containerVersion"]["gtagConfig"] = [
            {
                "gtagConfigId": "80",
                "name": "Primary Google destination",
                "type": "googtag",
                "parameter": [
                    {"type": "TEMPLATE", "key": "measurementId", "value": "G-TEST123"},
                    {
                        "type": "TEMPLATE",
                        "key": "server_container_url",
                        "value": "https://collect.example.test",
                    },
                ],
            }
        ]
        export = self.root / "unsafe-discovery-subset.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        review = complete_architecture(export)
        deterministic = next(
            row
            for row in review["comparisons"]
            if "browser_server_consent_deduplication_review" in row["comparison_types"]
        )
        candidate_keys = [
            deterministic["candidate_object_keys"][0],
            deterministic["candidate_object_keys"][-1],
        ]
        discovered = copy.deepcopy(deterministic)
        discovered.update(
            {
                "comparison_id": "DISC-UNSAFE-SUBSET-001",
                "comparison_origin": "analyst_discovered",
                "discovery_methods": ["semantic_name_and_business_term_variants"],
                "candidate_object_keys": candidate_keys,
                "member_assessments": [
                    item
                    for item in discovered["member_assessments"]
                    if item["object_key"] in candidate_keys
                ],
                "owner_question": unsafe_owner_question(
                    {"browser_server_consent_deduplication_review"}, candidate_keys
                ),
            }
        )
        discovered.pop("comparison_types", None)
        review["comparisons"].append(discovered)
        attestation = review["open_discovery_attestation"]
        attestation["discovered_comparison_ids"] = ["DISC-UNSAFE-SUBSET-001"]
        attestation["zero_discovery_rationale"] = ""
        next(
            item
            for item in attestation["method_reviews"]
            if item["method"] == "semantic_name_and_business_term_variants"
        )["additional_discovery_ids"] = ["DISC-UNSAFE-SUBSET-001"]

        errors, _ = validate_architecture(
            export,
            self.write_review("unsafe-discovery-subset-review.json", review),
        )
        self.assertTrue(
            any(
                "browser_server_consent_deduplication_review must be attributed" in error
                and "consent_sequence_and_server_route_conflicts" in error
                for error in errors
            )
        )

    def test_run3_rejects_generic_unsafe_questions_and_positive_runtime_claims(self) -> None:
        data = sample_export()
        data["containerVersion"]["gtagConfig"] = [
            {
                "gtagConfigId": "80",
                "name": "Primary Google destination",
                "type": "googtag",
                "parameter": [
                    {"type": "TEMPLATE", "key": "measurementId", "value": "G-TEST123"},
                    {
                        "type": "TEMPLATE",
                        "key": "server_container_url",
                        "value": "https://collect.example.test",
                    },
                ],
            }
        ]
        export = self.root / "architecture-negative-runtime-facts.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        review = complete_architecture(export)
        comparison = next(
            row
            for row in review["comparisons"]
            if "browser_server_consent_deduplication_review" in row["comparison_types"]
        )
        comparison.update(
            {
                "analyst_rationale": architecture_text(
                    comparison,
                    "analyst_rationale",
                    "Runtime event ID deduplication is unproven. Browser and server event IDs "
                    "are guaranteed identical and synchronized",
                ),
                "architecture_effect": architecture_text(
                    comparison,
                    "architecture_effect",
                    "Browser and server consent is verified equivalent and consistent",
                ),
                "owner_question": (
                    f"{' '.join(comparison['candidate_object_keys'][:2])} browser server "
                    "consent route deduplication canonical. Should this relationship be reviewed?"
                ),
            }
        )
        errors, _ = validate_architecture(
            export,
            self.write_review("positive-runtime-overclaim.json", review),
        )
        self.assertTrue(any("overclaims a complete, guaranteed" in error for error in errors))
        self.assertTrue(any("owner question must name" in error for error in errors))
        self.assertTrue(any("inside the interrogative clause" in error for error in errors))

        separated_review = complete_architecture(export)
        separated_comparison = next(
            row
            for row in separated_review["comparisons"]
            if "browser_server_consent_deduplication_review" in row["comparison_types"]
        )
        separated_comparison["analyst_rationale"] = architecture_text(
            separated_comparison,
            "analyst_rationale",
            "Consent and event-ID deduplication remain unproven from this export",
        )
        separated_comparison["architecture_effect"] = architecture_text(
            separated_comparison,
            "architecture_effect",
            "Consent " + ("distant context " * 30) + "is aligned and verified end to end",
        )
        separated_errors, _ = validate_architecture(
            export,
            self.write_review("separated-runtime-overclaim.json", separated_review),
        )
        self.assertTrue(
            any("overclaims a complete, guaranteed" in error for error in separated_errors)
        )

    def test_run3_generates_browser_server_consent_and_deduplication_family(self) -> None:
        data = sample_export()
        media_code = data["containerVersion"]["tag"][1]["parameter"][0]["value"]
        data["containerVersion"]["tag"][1]["parameter"][0]["value"] = media_code.replace(
            "</script>",
            "var s=document.createElement('script');"
            "s.src='https://metrics.example.test/p.js';"
            "document.head.appendChild(s);</script>",
        )
        data["containerVersion"]["gtagConfig"] = [
            {
                "gtagConfigId": "80",
                "name": "Primary Google destination",
                "type": "googtag",
                "parameter": [
                    {"type": "TEMPLATE", "key": "measurementId", "value": "G-TEST123"},
                    {
                        "type": "TEMPLATE",
                        "key": "server_container_url",
                        "value": "https://collect.example.test",
                    },
                ],
            }
        ]
        export = self.root / "browser-server-family.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        candidates = relationship_candidates(data["containerVersion"])
        family = next(
            row
            for row in candidates
            if "browser_server_consent_deduplication_review" in row["comparison_types"]
        )
        self.assertTrue(
            {"gtagConfig:80", "tag:1", "tag:2"}.issubset(family["candidate_object_keys"])
        )
        self.assertEqual(
            {
                "consent_sequence_and_server_route_conflicts",
                "consumer_destination_and_event_overlap",
                "terminal_source_formula_and_output_overlap",
            },
            set(family["discovery_methods"]),
        )
        self.assertIn(
            "metrics.example.test",
            next(
                row
                for row in candidates
                if {"tag:1", "tag:2"}.issubset(row["candidate_object_keys"])
            )["candidate_specificity_tokens"].get("tag:2", []),
        )
        for row in candidates:
            for terms in row.get("candidate_distinguishing_terms", {}).values():
                self.assertFalse(
                    any(
                        term in {"{}", "[]", "missing", "malformed", "script", "true"}
                        or any(
                            marker in term
                            for marker in ("<script", "document.", "window.", ".src", "ttq.")
                        )
                        for term in terms
                    )
                )

    def test_run3_cannot_hide_visible_relationship_inside_evidence_limit(self) -> None:
        review = complete_architecture(self.export_path)
        comparison = next(
            row
            for row in review["comparisons"]
            if "shared_execution_trigger" in row["comparison_types"]
        )
        comparison.update(
            {
                "relationship_verdict": "Container evidence limit",
                "disposition": "container_evidence_limit",
                "owner_question": (
                    "Which unseen runtime evidence proves that these visible routes remain separate?"
                ),
                "analyst_rationale": architecture_text(
                    comparison,
                    "analyst_rationale",
                    "The visible route is recorded while downstream runtime behavior is not visible",
                ),
                "architecture_effect": architecture_text(
                    comparison,
                    "architecture_effect",
                    "Visible configuration remains, but external evidence is unseen",
                ),
            }
        )
        errors, _ = validate_architecture(
            self.export_path,
            self.write_review("deterministic-evidence-limit.json", review),
        )
        self.assertTrue(
            any("deterministic source-visible relationship" in error for error in errors)
        )

    def test_run3_rejects_keep_for_zone_overlap_and_trigger_group_cycle(self) -> None:
        data = sample_export()
        data["containerVersion"]["trigger"].extend(
            [
                {
                    "triggerId": "30",
                    "name": "Cycle A",
                    "type": "TRIGGER_GROUP",
                    "parameter": [
                        {
                            "type": "LIST",
                            "key": "triggerIds",
                            "list": [{"type": "TEMPLATE", "value": "31"}],
                        }
                    ],
                },
                {
                    "triggerId": "31",
                    "name": "Cycle B",
                    "type": "TRIGGER_GROUP",
                    "parameter": [
                        {
                            "type": "LIST",
                            "key": "triggerIds",
                            "list": [{"type": "TEMPLATE", "value": "30"}],
                        }
                    ],
                },
            ]
        )
        data["containerVersion"]["zone"] = [
            {
                "zoneId": "70",
                "name": "Child A",
                "childContainer": [{"publicId": "GTM-CHILD"}],
                "boundary": {"customEvaluationTriggerId": ["10"]},
                "typeRestriction": {"enable": True, "whitelistedTypeId": ["html"]},
            },
            {
                "zoneId": "71",
                "name": "Child B",
                "childContainer": [{"publicId": "GTM-CHILD"}],
                "boundary": {"customEvaluationTriggerId": ["11"]},
                "typeRestriction": {"enable": True, "whitelistedTypeId": ["gaawe"]},
            },
        ]
        export = self.root / "unsafe-zone-cycle-retention.json"
        export.write_text(json.dumps(data), encoding="utf-8")
        review = complete_architecture(export)
        for comparison in review["comparisons"]:
            if set(comparison["comparison_types"]) & {
                "shared_zone_child_container",
                "cyclic_trigger_group_dependency",
            }:
                comparison.update(
                    {
                        "relationship_verdict": "Intentional variant",
                        "disposition": "keep",
                        "owner_question": "",
                    }
                )
        errors, _ = validate_architecture(
            export,
            self.write_review("unsafe-zone-cycle-retention-review.json", review),
        )
        self.assertTrue(any("multiple Zones governing one child" in error for error in errors))
        self.assertTrue(any("cyclic trigger-group dependency" in error for error in errors))

    def test_mandatory_operational_module_oracle_is_complete_and_fail_closed(self) -> None:
        expected = {
            "source_integrity",
            "inventory",
            "destination_inventory",
            "recognized_system_references",
            "missing_references",
            "duplicate_tag_names",
            "duplicate_trigger_names",
            "duplicate_variable_names",
            "duplicate_folder_names",
            "duplicate_zone_names",
            "duplicate_custom_template_names",
            "duplicate_client_names",
            "duplicate_transformation_names",
            "duplicate_tag_configurations",
            "normalized_duplicate_tag_signatures",
            "duplicate_trigger_logic",
            "duplicate_variable_logic",
            "duplicate_zone_configurations",
            "duplicate_google_tag_configurations",
            "duplicate_client_configurations",
            "duplicate_transformation_configurations",
            "duplicate_custom_template_configurations",
            "duplicate_variable_paths",
            "outdated_ua_styled_setup_objects",
            "unused_variables",
            "unused_triggers",
            "tags_without_firing_triggers",
            "unused_custom_templates",
            "unused_folders",
            "paused_tags",
            "used_only_by_paused_tags",
            "tag_sequence_structure",
            "tag_execution_controls",
            "single_member_trigger_groups",
            "trigger_group_structure",
            "zone_structure",
            "duplicate_custom_code",
            "variables_mirroring_builtins",
            "custom_variable_formula_logic",
            "consent_variable_logic",
            "media_tag_consent_route",
            "trigger_condition_lint",
            "ineffective_blocking_triggers",
            "unfiled_objects",
            "singleton_folders",
            "overloaded_folders",
            "name_hygiene",
            "naming_architecture_standardization",
        }
        self.assertEqual(expected, set(MANDATORY_OPERATIONAL_MODULES))
        scan = audit_export(self.export_path)
        self.assertEqual(expected, {row["module_name"] for row in scan["modules"]})
        mutated = copy.deepcopy(scan)
        mutated["modules"] = [
            row for row in mutated["modules"] if row["module_name"] != "tag_sequence_structure"
        ]
        self.assertTrue(
            any("tag_sequence_structure" in error for error in mandatory_module_errors(mutated))
        )

    def test_change_log_diff_is_field_level(self) -> None:
        before = container_version(sample_export())
        after = copy.deepcopy(before)
        after["tag"][0]["name"] = "GA4 - Purchase - Global"
        rows = diff_operations(before, after, "Direct", "Deep")
        self.assertEqual(1, len(rows))
        self.assertEqual("$.name", rows[0]["field_path"])
        change_log_path = self.root / "change-log.xlsx"
        build_change_log(
            {"kind": "gtm_field_level_change_log", "execution_mode": "planned", "changes": rows},
            change_log_path,
        )
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")
        workbook = load_workbook(change_log_path)
        self.assertLessEqual(len(workbook.sheetnames), 3)
        for sheet in workbook:
            self.assertLessEqual(sheet.max_column, 6)

    def test_change_log_supports_zone_and_google_tag_configuration_layers(self) -> None:
        before = container_version(sample_export())
        before["zone"] = [
            {
                "zoneId": "70",
                "name": "Partner Zone",
                "childContainer": [{"publicId": "GTM-CHILD"}],
                "boundary": {"customEvaluationTriggerId": ["10"]},
            }
        ]
        before["gtagConfig"] = [
            {
                "gtagConfigId": "71",
                "type": "GOOGLE_TAG",
                "parameter": [
                    {"type": "TEMPLATE", "key": "tag_id", "value": "G-TEST123"}
                ],
            }
        ]
        after = copy.deepcopy(before)
        after["zone"][0]["boundary"]["customEvaluationTriggerId"] = ["11"]
        after["gtagConfig"][0]["parameter"][0]["value"] = "G-NEW123"
        rows = diff_operations(before, after, "Direct", "Deep")
        self.assertEqual(
            {"Zone", "Google tag configuration"}, {row["layer"] for row in rows}
        )

        duplicate = copy.deepcopy(before)
        duplicate["zone"].append(copy.deepcopy(duplicate["zone"][0]))
        with self.assertRaisesRegex(ValueError, "change-log source fails integrity"):
            diff_operations(duplicate, after, "Direct", "Deep")

        invalid_artifact = self.root / "invalid-import-artifact.json"
        invalid_artifact.write_text(
            json.dumps({"tag": [{"name": "Missing tag identity", "type": "html"}]}),
            encoding="utf-8",
        )
        artifact_report = validate_artifact(invalid_artifact, None, "audit")
        self.assertEqual("fail", artifact_report["status"])
        self.assertIn(
            "artifact_source_integrity",
            {row["check"] for row in artifact_report["errors"]},
        )

    def test_executed_change_log_links_only_exact_approved_fields(self) -> None:
        operational, configuration, architecture = self.completed_reviews()
        align_duplicate_operation(operational, architecture)
        approved, errors = compile_operations(
            operational,
            configuration,
            architecture,
            "Direct GTM/MCP/API",
            "Deep",
            source_object_catalog(self.export_path),
        )
        self.assertEqual([], errors)
        future, apply_errors = apply_operations(sample_export(), approved)
        self.assertEqual([], apply_errors)
        after = container_version(future)
        after["tag"][0]["name"] = "Unexpected unapproved rename"
        rows = diff_operations(
            container_version(sample_export()),
            after,
            "Direct GTM/MCP/API",
            "Deep",
            approved,
            "executed",
        )
        deletion = next(row for row in rows if row["object_id"] == "21")
        unexpected = next(
            row
            for row in rows
            if row["object_id"] == "1" and row["field_path"] == "$.name"
        )
        self.assertTrue(deletion["operation_id"])
        self.assertEqual("Applied", deletion["status"])
        self.assertEqual("", unexpected["operation_id"])
        self.assertEqual("Blocked: missing approved operation link", unexpected["status"])

    def test_workbook_escapes_formula_text_and_privacy_scans_hidden_tabs(self) -> None:
        try:
            from openpyxl import Workbook, load_workbook
        except ImportError:
            self.skipTest("openpyxl is not installed")
        self.assertEqual("'=1+1", spreadsheet_safe_text("=1+1"))

        operational, configuration, architecture = self.completed_reviews()
        align_duplicate_operation(operational, architecture)
        payload, errors = compile_operations(
            operational,
            configuration,
            architecture,
            "Direct GTM/MCP/API",
            "Deep",
            source_object_catalog(self.export_path),
        )
        self.assertEqual([], errors)
        human, human_errors = build_rows(payload)
        self.assertEqual([], human_errors)
        human[0]["Problem / evidence"] = '=HYPERLINK("https://example.test","open")'
        workbook_path = self.root / "formula-safe.xlsx"
        build_workbook(
            {"source_file": self.export_path.name},
            build_model(self.export_path),
            operational,
            configuration,
            architecture,
            payload,
            {"rows": human},
            workbook_path,
        )
        rendered = load_workbook(workbook_path, data_only=False)
        formula_cell = rendered["02 Cleanup Plan"]["E2"]
        self.assertEqual("s", formula_cell.data_type)
        self.assertTrue(str(formula_cell.value).startswith("'="))
        rendered.close()

        privacy_path = self.root / "hidden-privacy.xlsx"
        raw = Workbook()
        raw.active.title = "Visible"
        hidden = raw.create_sheet("Hidden proof")
        hidden.sheet_state = "hidden"
        hidden["A1"] = "@".join(("analyst", "example.test"))
        raw.save(privacy_path)
        self.assertTrue(
            any("Hidden proof!A1" in finding for finding in scan_xlsx(privacy_path))
        )

    def test_privacy_helpers_redact_identity_and_sensitive_urls(self) -> None:
        address = "@".join(("jane.doe", "example.test"))
        text = f"Contact {address} or +33 6 12 34 56 78 token=secret-value"
        redacted = redact_text(text)
        self.assertNotIn(address, redacted)
        self.assertNotIn("secret-value", redacted)
        url_address = "@".join(("jane", "example.test"))
        url = sanitize_url(f"https://example.test/path?email={url_address}&utm_source=test")
        self.assertNotIn(url_address, url)
        self.assertIn("utm_source=%3Cvalue%3E", url)
        self.assertTrue(privacy_findings(text))

    def test_vendor_registry_is_current_and_structurally_valid(self) -> None:
        registry_path = ROOT / "references/03-rules/vendor-registry.toml"
        errors, warnings = validate_registry(registry_path, online=False, max_age_days=365)
        self.assertEqual([], errors)
        self.assertEqual([], warnings)
        self.assertGreaterEqual(len(load_registry(registry_path)["vendors"]), 15)
        self.assertEqual(
            "Universal Analytics (legacy)",
            vendor_record('{"type": "ua", "trackingId": "UA-123"}')["name"],
        )
        self.assertEqual(
            "Google Ads",
            vendor_record('{"type": "googtag", "tagId": "AW-123"}')["name"],
        )
        self.assertEqual(
            "GA4 / Google tag",
            vendor_record('{"type": "gaawe", "eventName": "purchase"}')["name"],
        )
        tiktok = next(
            vendor
            for vendor in load_registry(registry_path)["vendors"]
            if vendor["name"] == "TikTok"
        )
        self.assertEqual(["CompletePayment"], tiktok["unsupported_standard_events"])
        self.assertEqual(["CompletePayment=>Purchase"], tiktok["event_replacements"])

    def test_vendor_registry_rejects_malformed_event_lifecycle_metadata(self) -> None:
        registry_path = self.root / "malformed-vendor-registry.toml"
        registry_path.write_text(
            """schema_version = 1
reviewed_on = "2026-07-20"

[[vendors]]
name = "Example"
category = "media"
patterns = ["example"]
official_docs = ["http://docs.example.test"]
unsupported_standard_events = ["OldEvent", "OldEvent"]
event_replacements = ["DifferentEvent=>NewEvent", "broken"]
""",
            encoding="utf-8",
        )
        errors, _ = validate_registry(registry_path, online=False)
        self.assertTrue(any("absolute HTTPS" in error for error in errors))
        self.assertTrue(any("contains duplicates" in error for error in errors))
        self.assertTrue(any("not listed" in error for error in errors))
        self.assertTrue(any("must use old=>new" in error for error in errors))

    def test_vendor_url_check_falls_back_to_get_when_head_is_rejected(self) -> None:
        head_error = urllib.error.HTTPError(
            "https://example.invalid", 404, "HEAD rejected", {}, None
        )
        get_response = MagicMock()
        get_response.status = 200
        get_response.__enter__.return_value = get_response
        with patch(
            "gtm_vendor_registry.urllib.request.urlopen",
            side_effect=[head_error, get_response],
        ) as urlopen:
            self.assertIsNone(official_url_error("https://example.invalid"))
        self.assertEqual("HEAD", urlopen.call_args_list[0].args[0].method)
        self.assertEqual("GET", urlopen.call_args_list[1].args[0].method)

    def test_release_layout_allows_ignored_editable_install_metadata(self) -> None:
        errors = check_repository_layout(ROOT)
        self.assertEqual([], errors)

    def test_release_check_does_not_hide_git_tracking_failure(self) -> None:
        (self.root / ".git").mkdir()
        with patch(
            "check_release.subprocess.check_output",
            side_effect=subprocess.CalledProcessError(1, ["git", "ls-files"]),
        ), self.assertRaisesRegex(RuntimeError, "tracked release resources"):
            git_ls_files(self.root)

    def test_release_tag_uses_semver_and_matches_project_version(self) -> None:
        self.assertEqual([], check_release_tag("v1.0.0"))
        self.assertEqual([], check_release_tag("v1.1.0-rc.1"))
        self.assertEqual([], check_release_tag("v1.1.0+build.7"))
        self.assertTrue(check_release_tag("v2026.07.20.1"))
        self.assertTrue(check_release_tag("v01.0.0"))
        self.assertTrue(check_release_tag("1.0.0"))
        self.assertEqual([], check_project_version(ROOT, "v1.0.0"))
        self.assertTrue(check_project_version(ROOT, "v1.0.1"))

    def test_runtime_bundle_is_self_installable_and_excludes_repo_only_files(self) -> None:
        bundle = self.root / "runtime-bundle"
        build_skill_bundle(ROOT, bundle)
        for filename in ("SKILL.md", "LICENSE", "pyproject.toml"):
            self.assertTrue((bundle / filename).is_file())
        self.assertFalse((bundle / "README.md").exists())
        metadata = tomllib.loads((bundle / "pyproject.toml").read_text(encoding="utf-8"))
        self.assertIsInstance(metadata["project"]["readme"], dict)
        self.assertNotIn("README.md", json.dumps(metadata["project"]["readme"]))
        for relative in (
            "scripts/gtm_operational_review.py",
            "scripts/gtm_configuration_review.py",
            "scripts/gtm_architecture_review.py",
            "scripts/gtm_review_common.py",
            "scripts/gtm_review_shards.py",
            "scripts/gtm_three_run_gate.py",
        ):
            self.assertTrue((bundle / relative).is_file())
        self.assertFalse((bundle / "tests").exists())
        self.assertFalse((bundle / ".github").exists())
        self.assertFalse((bundle / "scripts/check_release.py").exists())
        self.assertFalse((bundle / "scripts/gtm_self_test.py").exists())


if __name__ == "__main__":
    unittest.main()
