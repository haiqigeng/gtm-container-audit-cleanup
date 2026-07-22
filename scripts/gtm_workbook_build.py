#!/usr/bin/env python3
"""Build the compact stakeholder GTM cleanup-plan workbook."""

from __future__ import annotations

import argparse
import json
import math
from pathlib import Path
from typing import Any

from gtm_privacy import redact_text, spreadsheet_safe_text

CANONICAL_SHEETS = [
    "01 Summary",
    "02 Cleanup Plan",
    "03 Operational Review",
    "04 Configuration Review",
    "05 Architecture Review",
    "06 Custom Code Review",
    "07 Reconciled Operations",
    "08 Source & Gates",
]

HEADER_FILL = "16324F"
HEADER_FONT = "FFFFFF"
ACCENT_FILL = "DCEEF2"
GRID_COLOR = "C8D2DC"
MAX_CELL_TEXT = 24000


def as_list(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def load_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def clean_text(value: Any) -> str:
    if isinstance(value, (dict, list)):
        text = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    else:
        text = str(value or "")
    text = redact_text(text)
    return spreadsheet_safe_text(text)


def join_text(values: list[Any]) -> str:
    return "; ".join(clean_text(value) for value in values if str(value or "").strip())


def cell_chunks(value: Any) -> list[str]:
    """Return lossless, formula-safe chunks that fit comfortably in an XLSX cell."""
    text = clean_text(value)
    if len(text) <= MAX_CELL_TEXT:
        return [text]
    return [
        spreadsheet_safe_text(text[start : start + MAX_CELL_TEXT])
        for start in range(0, len(text), MAX_CELL_TEXT)
    ]


def expanded_table_rows(
    rows: list[dict[str, Any]], headers: list[str], split_long_cells: bool
) -> list[list[str]]:
    expanded: list[list[str]] = []
    for row_number, row in enumerate(rows, start=2):
        chunks_by_column = [cell_chunks(row.get(header, "")) for header in headers]
        part_count = max(len(chunks) for chunks in chunks_by_column)
        if part_count > 1 and not split_long_cells:
            raise ValueError(
                f"visible workbook row {row_number} exceeds {MAX_CELL_TEXT} characters; "
                "summarize the user-facing row instead of truncating it"
            )
        for part in range(part_count):
            expanded.append(
                [
                    chunks[part] if part < len(chunks) else ""
                    for chunks in chunks_by_column
                ]
            )
    return expanded


def decision_text(row: dict[str, Any]) -> str:
    operation = row.get("operation") or {}
    operations = as_list(row.get("operations"))
    action = operation.get("exact_proposed_action") if operation else ""
    if not action and operations:
        action = join_text([item.get("exact_proposed_action") for item in operations])
    return join_text(
        [
            row.get("disposition"),
            row.get("owner_question"),
            action,
            row.get("confidence"),
        ]
    )


def operational_rows(review: dict[str, Any]) -> list[dict[str, str]]:
    rows = []
    for finding in as_list(review.get("findings")):
        rows.append(
            {
                "Finding": clean_text(finding.get("finding_id")),
                "Type / module": clean_text(
                    f"{finding.get('finding_type')} / {finding.get('module_name')}"
                ),
                "Affected objects": join_text(
                    [
                        f"{object_id} - {name}"
                        for object_id, name in zip(
                            as_list(finding.get("object_ids")),
                            as_list(finding.get("object_names")),
                            strict=False,
                        )
                    ]
                ),
                "Evidence": clean_text(finding.get("deterministic_evidence")),
                "Decision": join_text([finding.get("disposition"), finding.get("rationale")]),
                "Action / status": join_text(
                    [
                        finding.get("exact_proposed_action"),
                        finding.get("priority"),
                        finding.get("execution_readiness"),
                        finding.get("owner_question"),
                    ]
                ),
            }
        )
    return rows


def configuration_rows(review: dict[str, Any]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for item in as_list(review.get("rows")):
        object_label = clean_text(
            f"{item.get('object_key')} - {item.get('object_name')} [{item.get('object_type')}]"
        )
        rows.append(
            {
                "Object": object_label,
                "Review aspect": "Functional contract",
                "Configured behavior / assessment": join_text(
                    [
                        item.get("purpose"),
                        item.get("execution_logic"),
                        item.get("inputs_and_terminal_sources"),
                        item.get("configured_output_or_side_effect"),
                        item.get("consumer_contract"),
                        item.get("consent_and_sequence"),
                    ]
                ),
                "Verdict": join_text(
                    [item.get("correctness_verdict"), item.get("correctness_basis")]
                ),
                "Evidence": join_text(as_list(item.get("evidence_anchors"))),
                "Decision": decision_text(item),
            }
        )
        for check in as_list(item.get("logic_cross_checks")):
            rows.append(
                {
                    "Object": object_label,
                    "Review aspect": clean_text(f"D3 logic - {check.get('check_key')}"),
                    "Configured behavior / assessment": clean_text(check.get("conclusion")),
                    "Verdict": clean_text(check.get("verdict")),
                    "Evidence": join_text(as_list(check.get("evidence_anchors"))),
                    "Decision": "",
                }
            )
        for branch in as_list(item.get("configuration_branch_reviews")):
            rows.append(
                {
                    "Object": object_label,
                    "Review aspect": clean_text(
                        f"Branch - {branch.get('logic_role')} - {branch.get('json_path')}"
                    ),
                    "Configured behavior / assessment": join_text(
                        [branch.get("interpretation"), branch.get("configured_effect")]
                    ),
                    "Verdict": clean_text(branch.get("correctness")),
                    "Evidence": clean_text(
                        {"path": branch.get("json_path"), "value_hash": branch.get("value_hash")}
                    ),
                    "Decision": "",
                }
            )
        for trace in as_list(item.get("reference_traces")):
            rows.append(
                {
                    "Object": object_label,
                    "Review aspect": clean_text(f"Reference trace - {trace.get('reference')}"),
                    "Configured behavior / assessment": clean_text(trace.get("terminal_source")),
                    "Verdict": clean_text(trace.get("terminal_states")),
                    "Evidence": join_text(as_list(trace.get("evidence_anchors"))),
                    "Decision": "",
                }
            )
            for node in as_list(trace.get("node_reviews")):
                rows.append(
                    {
                        "Object": object_label,
                        "Review aspect": clean_text(
                            f"Trace node - {node.get('object_key')} - {node.get('object_name')}"
                        ),
                        "Configured behavior / assessment": join_text(
                            [
                                node.get("configured_function"),
                                node.get("configured_output"),
                                node.get("output_type_and_shape"),
                                node.get("availability_and_fallback"),
                                node.get("consumer_compatibility"),
                            ]
                        ),
                        "Verdict": clean_text(node.get("semantic_role")),
                        "Evidence": clean_text(node.get("configured_parameters")),
                        "Decision": "",
                    }
                )
        for contract in as_list(item.get("contract_checks")):
            rows.append(
                {
                    "Object": object_label,
                    "Review aspect": clean_text(
                        f"Vendor contract - {contract.get('contract_topic')}"
                    ),
                    "Configured behavior / assessment": join_text(
                        [contract.get("configured_value"), contract.get("expected_rule")]
                    ),
                    "Verdict": clean_text(contract.get("verdict")),
                    "Evidence": join_text(
                        [contract.get("source"), *as_list(contract.get("evidence_anchors"))]
                    ),
                    "Decision": "",
                }
            )
        for defect in as_list(item.get("defects")):
            rows.append(
                {
                    "Object": object_label,
                    "Review aspect": clean_text(f"Defect - {defect.get('defect_id')}"),
                    "Configured behavior / assessment": join_text(
                        [
                            defect.get("statement"),
                            defect.get("configured_effect"),
                            defect.get("expected_behavior"),
                        ]
                    ),
                    "Verdict": "Issue",
                    "Evidence": join_text(
                        [
                            *as_list(defect.get("evidence_anchors")),
                            *as_list(defect.get("code_line_hashes")),
                        ]
                    ),
                    "Decision": decision_text(item),
                }
            )
    return rows


def architecture_rows(review: dict[str, Any]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for item in as_list(review.get("families")):
        rows.append(
            {
                "Family / comparison": clean_text(
                    f"{item.get('family_id')} - {item.get('family_label')}"
                ),
                "Members / chain": join_text(as_list(item.get("chain_object_keys"))),
                "Business and execution logic": join_text(
                    [
                        item.get("business_action"),
                        item.get("family_purpose"),
                        item.get("execution_path_summary"),
                        item.get("payload_coherence"),
                        item.get("consent_and_sequence_coherence"),
                    ]
                ),
                "Relationship verdict": join_text(
                    [item.get("relationship_verdict"), item.get("analyst_rationale")]
                ),
                "Necessity / target state": join_text(
                    [item.get("necessity_and_ownership"), item.get("target_architecture")]
                ),
                "Decision": decision_text(item),
            }
        )
        for assessment in as_list(item.get("chain_assessments")):
            rows.append(
                {
                    "Family / comparison": clean_text(item.get("family_id")),
                    "Members / chain": clean_text(assessment.get("object_key")),
                    "Business and execution logic": join_text(
                        [
                            assessment.get("configured_role"),
                            assessment.get("necessity"),
                            assessment.get("distinguishing_configuration"),
                        ]
                    ),
                    "Relationship verdict": clean_text(assessment.get("status")),
                    "Necessity / target state": join_text(
                        as_list(assessment.get("evidence_anchors"))
                    ),
                    "Decision": "",
                }
            )
    for item in as_list(review.get("comparisons")):
        rows.append(
            {
                "Family / comparison": clean_text(item.get("comparison_id")),
                "Members / chain": clean_text(item.get("candidate_object_keys", [])),
                "Business and execution logic": join_text(
                    [item.get("candidate_basis"), item.get("architecture_effect")]
                ),
                "Relationship verdict": join_text(
                    [item.get("relationship_verdict"), item.get("analyst_rationale")]
                ),
                "Necessity / target state": join_text(
                    [
                        assessment.get("necessity")
                        for assessment in as_list(item.get("member_assessments"))
                    ]
                ),
                "Decision": decision_text(item),
            }
        )
    return rows


def code_rows(configuration: dict[str, Any]) -> list[dict[str, str]]:
    rows = []
    for item in as_list(configuration.get("rows")):
        if not item.get("required_code_line_hashes"):
            continue
        facts = item.get("technical_code_facts") or {}
        object_label = clean_text(f"{item.get('object_key')} - {item.get('object_name')}")
        facts_by_hash = {
            str(fact.get("line_hash") or ""): fact
            for fact in as_list(item.get("code_line_facts"))
        }
        for block in as_list(item.get("code_behavior_blocks")):
            line_facts = [
                facts_by_hash[line_hash]
                for line_hash in as_list(block.get("line_hashes"))
                if line_hash in facts_by_hash
            ]
            rows.append(
                {
                    "Object": object_label,
                    "Lines": clean_text(f"{block.get('start_line')}-{block.get('end_line')}"),
                    "Behavior": clean_text(block.get("purpose")),
                    "Inputs / outputs / side effects": join_text(
                        [block.get("inputs"), block.get("outputs"), block.get("side_effects")]
                    ),
                    "Code health / evidence": join_text(
                        [
                            block.get("health_assessment"),
                            *[fact.get("line_preview") for fact in line_facts],
                            facts.get("javascript_parser"),
                        ]
                    ),
                    "Decision": decision_text(item),
                }
            )
        for finding in as_list(item.get("technical_finding_reviews")):
            rows.append(
                {
                    "Object": object_label,
                    "Lines": clean_text(finding.get("finding_key")),
                    "Behavior": clean_text(finding.get("source_statement")),
                    "Inputs / outputs / side effects": clean_text(finding.get("rationale")),
                    "Code health / evidence": clean_text(finding.get("verdict")),
                    "Decision": decision_text(item),
                }
            )
    return rows


def operation_rows(payload: dict[str, Any]) -> list[dict[str, str]]:
    rows = []
    for operation in [
        *as_list(payload.get("operations")),
        *as_list(payload.get("deferred_operations")),
    ]:
        rows.append(
            {
                "Operation": clean_text(
                    f"{operation.get('operation_id')} / {operation.get('resolution_status')}"
                ),
                "Area / problem": clean_text(
                    f"{operation.get('area')} / {operation.get('problem_type')}"
                ),
                "Affected objects": clean_text(operation.get("affected_objects")),
                "Reason / target state": join_text(
                    [
                        operation.get("problem"),
                        operation.get("why_it_matters"),
                        operation.get("expected_clean_state"),
                    ]
                ),
                "Exact mutation": clean_text(
                    {
                        "action": operation.get("exact_proposed_action"),
                        "creations": operation.get("creations", []),
                        "additions": operation.get("additions", []),
                        "changes": operation.get("changes", []),
                        "remaps": operation.get("remaps", []),
                        "deletions": operation.get("deletions", []),
                        "renames": operation.get("renames", []),
                    }
                ),
                "Priority / QA / rollback": join_text(
                    [
                        operation.get("priority"),
                        operation.get("execution_readiness"),
                        operation.get("qa_steps"),
                        operation.get("rollback"),
                    ]
                ),
            }
        )
    return rows


def source_rows(
    manifest: dict[str, Any],
    source: dict[str, Any],
    operations: dict[str, Any],
) -> list[dict[str, str]]:
    return [
        {"Check": "Source file", "Result": clean_text(manifest.get("source_file"))},
        {"Check": "Source SHA-256", "Result": clean_text(manifest.get("source_sha256"))},
        {
            "Check": "Shared deterministic facts",
            "Result": clean_text(manifest.get("shared_facts_sha256")),
        },
        {
            "Check": "Audit context",
            "Result": clean_text(manifest.get("context_sha256")),
        },
        {
            "Check": "Source model coverage",
            "Result": clean_text(source.get("coverage_gate")),
        },
        {
            "Check": "Operational sanitation run",
            "Result": clean_text(
                (operations.get("run_statuses") or {}).get("operational_sanitation")
            ),
        },
        {
            "Check": "Configuration correctness run",
            "Result": clean_text(
                (operations.get("run_statuses") or {}).get("configuration_correctness")
            ),
        },
        {
            "Check": "Business architecture run",
            "Result": clean_text(
                (operations.get("run_statuses") or {}).get("business_architecture")
            ),
        },
        {"Check": "Execution route", "Result": clean_text(operations.get("route"))},
        {"Check": "Cleanup level", "Result": clean_text(operations.get("aggressiveness"))},
        {
            "Check": "Decision ledger records",
            "Result": clean_text(len(as_list(operations.get("decision_ledger")))),
        },
    ]


def add_table(
    sheet: Any,
    rows: list[dict[str, Any]],
    headers: list[str] | None = None,
    *,
    split_long_cells: bool = False,
) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    headers = headers or (list(rows[0]) if rows else ["Status"])
    if not rows:
        rows = [{headers[0]: "No rows"}]
    sheet.append(headers)
    for values in expanded_table_rows(rows, headers, split_long_cells):
        sheet.append(values)
    thin = Side(style="thin", color=GRID_COLOR)
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(color=HEADER_FONT, bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(bottom=thin)
    sheet.row_dimensions[1].height = 30
    widths = [16, 28, 42, 54, 54, 54] if len(headers) >= 6 else [28, 92]
    for row_number in range(2, sheet.max_row + 1):
        estimated_lines = 1
        for column_number, cell in enumerate(sheet[row_number], start=1):
            value = str(cell.value or "")
            width = widths[min(column_number - 1, len(widths) - 1)]
            estimated_lines = max(
                estimated_lines,
                value.count("\n") + max(1, math.ceil(len(value) / max(12, width * 1.25))),
            )
        sheet.row_dimensions[row_number].height = min(120, max(36, estimated_lines * 15))
        if row_number % 2 == 0:
            for cell in sheet[row_number]:
                cell.fill = PatternFill("solid", fgColor="F5F8FA")
        for cell in sheet[row_number]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=Side(style="hair", color=GRID_COLOR))
    for index in range(1, len(headers) + 1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = widths[
            min(index - 1, len(widths) - 1)
        ]
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False


def build_workbook(
    manifest: dict[str, Any],
    source: dict[str, Any],
    operational: dict[str, Any],
    configuration: dict[str, Any],
    architecture: dict[str, Any],
    operations: dict[str, Any],
    human_rows: dict[str, Any],
    output: Path,
) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to build XLSX output") from exc

    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = {name: workbook.create_sheet(name) for name in CANONICAL_SHEETS}
    operation_count = len(as_list(operations.get("operations")))
    deferred_count = len(as_list(operations.get("deferred_operations")))
    ledger = as_list(operations.get("decision_ledger"))
    owner_decisions = sum(
        1
        for row in ledger
        if row.get("disposition") == "owner_decision_needed"
    )
    evidence_limits = sum(
        1
        for row in ledger
        if row.get("disposition") == "container_evidence_limit"
    )
    retained_decisions = sum(1 for row in ledger if row.get("disposition") == "keep")
    documented_exceptions = sum(
        1 for row in ledger if row.get("disposition") == "documented_exception"
    )
    retained_families = [
        str(row.get("title") or row.get("decision_id") or "")
        for row in ledger
        if row.get("source_run") == "business_architecture"
        and str(row.get("decision_id") or "").startswith("FAM-")
        and row.get("disposition") == "keep"
    ]
    retained_family_summary = join_text(retained_families[:6])
    if len(retained_families) > 6:
        retained_family_summary += f"; +{len(retained_families) - 6} more retained families"
    priority_order = {"Critical": 0, "High": 1, "Medium": 2, "Low": 3}
    highest_impact = sorted(
        as_list(operations.get("operations")),
        key=lambda row: (
            priority_order.get(str(row.get("priority") or ""), 4),
            str(row.get("operation_id") or ""),
        ),
    )[:5]
    highest_impact_summary = join_text(
        [
            f"[{row.get('priority')}] {row.get('operation_id')}: "
            f"{row.get('problem')} -> "
            f"{row.get('exact_proposed_action')}"
            for row in highest_impact
        ]
    )
    if operations.get("aggressiveness") == "Undecided":
        overall_status = "Cleanup level decision required"
        if owner_decisions:
            overall_status += "; owner decisions required before cleanup approval"
    elif owner_decisions:
        overall_status = "Owner decisions required before cleanup approval"
    elif evidence_limits:
        overall_status = "Plan ready with documented container evidence limits"
    elif operation_count:
        overall_status = "Ready for human approval"
    else:
        overall_status = "Audit complete; no cleanup operation proposed"
    next_step = (
        "Select the cleanup level, then resolve any owner decisions before approval."
        if operations.get("aggressiveness") == "Undecided"
        else "Resolve the listed owner questions before approving cleanup."
        if owner_decisions
        else "Approve, reject, or amend the proposed operations before any GTM mutation."
    )
    summary = [
        {"Decision": "Overall status", "Value": overall_status},
        {"Decision": "Source", "Value": manifest.get("source_file")},
        {"Decision": "Objects reviewed", "Value": len(as_list(configuration.get("rows")))},
        {
            "Decision": "Operational findings reviewed",
            "Value": len(as_list(operational.get("findings"))),
        },
        {
            "Decision": "Business families reviewed",
            "Value": len(as_list(architecture.get("families"))),
        },
        {
            "Decision": "Retained / no-change decisions",
            "Value": retained_decisions,
        },
        {"Decision": "Documented owner exceptions", "Value": documented_exceptions},
        {
            "Decision": "Retained business-family architecture",
            "Value": retained_family_summary or "No retained family decision recorded",
        },
        {
            "Decision": "Highest-impact proposed actions",
            "Value": highest_impact_summary or "No cleanup operation proposed",
        },
        {"Decision": "Proposed operations", "Value": operation_count},
        {"Decision": "Deferred operations", "Value": deferred_count},
        {"Decision": "Owner decisions", "Value": owner_decisions},
        {"Decision": "Container evidence limits", "Value": evidence_limits},
        {
            "Decision": "Projected object counts",
            "Value": join_text(
                [
                    f"{layer}: {counts.get('before', 0)} -> {counts.get('after', 0)} "
                    f"({counts.get('delta', 0):+d})"
                    for layer, counts in sorted(
                        (operations.get("projected_object_counts") or {}).items()
                    )
                ]
            )
            or "No count-changing operation proposed",
        },
        {"Decision": "Execution route", "Value": operations.get("route")},
        {"Decision": "Cleanup level", "Value": operations.get("aggressiveness")},
        {
            "Decision": "Next step",
            "Value": next_step,
        },
    ]
    add_table(sheets["01 Summary"], summary, ["Decision", "Value"])
    add_table(sheets["02 Cleanup Plan"], as_list(human_rows.get("rows")))
    add_table(
        sheets["03 Operational Review"],
        operational_rows(operational),
        split_long_cells=True,
    )
    add_table(
        sheets["04 Configuration Review"],
        configuration_rows(configuration),
        split_long_cells=True,
    )
    add_table(
        sheets["05 Architecture Review"],
        architecture_rows(architecture),
        split_long_cells=True,
    )
    add_table(
        sheets["06 Custom Code Review"],
        code_rows(configuration),
        split_long_cells=True,
    )
    add_table(
        sheets["07 Reconciled Operations"],
        operation_rows(operations),
        split_long_cells=True,
    )
    add_table(
        sheets["08 Source & Gates"],
        source_rows(manifest, source, operations),
        split_long_cells=True,
    )
    for name in CANONICAL_SHEETS[2:]:
        sheets[name].sheet_state = "hidden"
    workbook.active = 0
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("package_dir", type=Path)
    parser.add_argument("operations", type=Path)
    parser.add_argument("human_rows", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    build_workbook(
        load_json(args.package_dir / "audit_package_manifest.json"),
        load_json(args.package_dir / "source_model.json"),
        load_json(args.package_dir / "operational_review.json"),
        load_json(args.package_dir / "configuration_review.json"),
        load_json(args.package_dir / "architecture_review.json"),
        load_json(args.operations),
        load_json(args.human_rows),
        args.output,
    )
    print(json.dumps({"output": str(args.output), "tabs": CANONICAL_SHEETS}))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
