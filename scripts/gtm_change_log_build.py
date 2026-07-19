#!/usr/bin/env python3
"""Build a separate human-facing XLSX change log from a field-level GTM diff."""

from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any

from gtm_diff_operations import csv_row
from gtm_workbook_build import add_table


def as_list(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def build_change_log(payload: dict[str, Any], output: Path) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to build XLSX output") from exc

    changes = as_list(payload.get("changes"))
    actions = Counter(str(row.get("action") or "") for row in changes)
    statuses = Counter(str(row.get("status") or "") for row in changes)
    summary = [
        {"Decision": "Log type", "Value": payload.get("execution_mode", "planned")},
        {"Decision": "Field-level changes", "Value": len(changes)},
        {"Decision": "Actions", "Value": json.dumps(actions, ensure_ascii=False, sort_keys=True)},
        {"Decision": "Statuses", "Value": json.dumps(statuses, ensure_ascii=False, sort_keys=True)},
        {
            "Decision": "Validation",
            "Value": "Verify every applied row through GTM readback, export diff, and dependency validation.",
        },
        {
            "Decision": "Next step",
            "Value": "Resolve unlinked or unverified rows before publish readiness.",
        },
    ]

    workbook = Workbook()
    workbook.remove(workbook.active)
    summary_sheet = workbook.create_sheet("01 Change Log Summary")
    detail_sheet = workbook.create_sheet("02 Change Log Details")
    proof_sheet = workbook.create_sheet("03 Field Diff Proof")
    add_table(summary_sheet, summary, ["Decision", "Value"])
    summary_sheet.column_dimensions["A"].width = 30
    summary_sheet.column_dimensions["B"].width = 90
    detail_rows = [csv_row(row) for row in changes]
    add_table(
        detail_sheet,
        detail_rows,
        [
            "Change ID",
            "Area / object",
            "Change made",
            "Before",
            "After",
            "Reason / QA / status",
        ],
    )
    proof_rows = [
        {
            "Change / operation": (
                f"{row.get('change_id')} / {row.get('operation_id') or 'unlinked'}"
            ),
            "Object": (
                f"{row.get('layer')} {row.get('object_id')} / "
                f"{row.get('before_name')} -> {row.get('after_name')}"
            ),
            "Field / action": (
                f"{row.get('change_category')} / {row.get('action')} / "
                f"{row.get('field_path')} / {row.get('route')} / "
                f"{row.get('aggressiveness')}"
            ),
            "Before / after": (
                f"Before: {row.get('before_value')}\nAfter: {row.get('after_value')}"
            ),
            "Reason / impact": (
                f"{row.get('reason')} Impact: {row.get('functional_impact')}"
                + (f" Blocker: {row.get('blocker')}" if row.get("blocker") else "")
            ),
            "QA / rollback / status": (
                f"QA: {row.get('qa_method')} ({row.get('qa_status')}). "
                f"Rollback: {row.get('rollback')} Status: {row.get('status')}"
            ),
        }
        for row in changes
    ]
    add_table(proof_sheet, proof_rows)
    proof_sheet.sheet_state = "hidden"
    workbook.active = 0
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("field_diff", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()
    payload = json.loads(args.field_diff.read_text(encoding="utf-8"))
    build_change_log(payload, args.output)
    print(json.dumps({"output": str(args.output), "changes": payload.get("changeCount", 0)}))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
