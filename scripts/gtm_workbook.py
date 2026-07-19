#!/usr/bin/env python3
"""Dependency-light workbook helpers for GTM audit validation scripts."""

from __future__ import annotations

import re
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = text.strip("_")
    return text


def column_index(cell_ref: str) -> int:
    letters = "".join(ch for ch in cell_ref if ch.isalpha()).upper()
    index = 0
    for char in letters:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return index - 1


def xml_text(node: ElementTree.Element | None) -> str:
    if node is None:
        return ""
    return "".join(node.itertext())


def workbook_target_path(target: str) -> str:
    return "xl/" + target.lstrip("/") if not target.startswith("xl/") else target


def rows_from_values(parsed_rows: list[list[Any]]) -> list[dict[str, Any]]:
    if not parsed_rows:
        return []
    headers = [normalize_header(value) for value in parsed_rows[0]]
    rows = []
    for values in parsed_rows[1:]:
        if not values or all(value in (None, "") for value in values):
            continue
        rows.append({headers[i]: values[i] for i in range(min(len(headers), len(values)))})
    return rows


def load_xlsx_workbook_openpyxl(path: Path) -> dict[str, list[dict[str, Any]]]:
    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise RuntimeError("openpyxl is unavailable") from exc

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    result: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        values = [list(row) for row in sheet.iter_rows(values_only=True)]
        result[sheet_name] = rows_from_values(values)
    workbook.close()
    return result


def load_xlsx_workbook_stdlib(path: Path) -> dict[str, list[dict[str, Any]]]:
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pkgrel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }

    with zipfile.ZipFile(path) as archive:
        workbook = ElementTree.fromstring(archive.read("xl/workbook.xml"))
        rels = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))

        rel_targets = {
            rel.attrib["Id"]: rel.attrib["Target"]
            for rel in rels.findall("pkgrel:Relationship", ns)
        }

        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            shared_root = ElementTree.fromstring(archive.read("xl/sharedStrings.xml"))
            for item in shared_root.findall("main:si", ns):
                shared_strings.append(xml_text(item))

        result: dict[str, list[dict[str, Any]]] = {}
        for sheet in workbook.findall("main:sheets/main:sheet", ns):
            name = sheet.attrib.get("name", "")
            rel_id = sheet.attrib.get(f"{{{ns['rel']}}}id")
            target = rel_targets.get(rel_id or "")
            if not name or not target:
                continue
            sheet_target = workbook_target_path(target)
            if sheet_target not in archive.namelist():
                continue

            parsed_rows: list[list[str]] = []
            sheet_root = ElementTree.fromstring(archive.read(sheet_target))
            for row in sheet_root.findall("main:sheetData/main:row", ns):
                values: list[str] = []
                for cell in row.findall("main:c", ns):
                    ref = cell.attrib.get("r", "")
                    index = column_index(ref) if ref else len(values)
                    while len(values) <= index:
                        values.append("")
                    cell_type = cell.attrib.get("t", "")
                    if cell_type == "s":
                        raw = xml_text(cell.find("main:v", ns))
                        values[index] = shared_strings[int(raw)] if raw else ""
                    elif cell_type == "inlineStr":
                        values[index] = xml_text(cell.find("main:is", ns))
                    else:
                        values[index] = xml_text(cell.find("main:v", ns))
                if any(value != "" for value in values):
                    parsed_rows.append(values)
            result[name] = rows_from_values(parsed_rows)
    return result


def load_xlsx_workbook(path: Path) -> dict[str, list[dict[str, Any]]]:
    try:
        import openpyxl  # type: ignore  # noqa: F401
    except ImportError:
        return load_xlsx_workbook_stdlib(path)
    try:
        return load_xlsx_workbook_openpyxl(path)
    except Exception as exc:  # noqa: BLE001 - preserve artifact parsing failure.
        raise RuntimeError(
            "Unable to read XLSX workbook with openpyxl; refusing the stdlib reader "
            "because parser failure may indicate a malformed artifact."
        ) from exc
