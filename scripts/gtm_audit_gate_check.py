#!/usr/bin/env python3
"""Validate the compact GTM cleanup-plan workbook structure and content."""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
from typing import Any

from gtm_workbook import load_xlsx_workbook
from gtm_workbook_build import CANONICAL_SHEETS

PLACEHOLDER_PATTERNS = (
    re.compile(r"\b(?:tbd|to" r"do|lorem ipsum)\b", re.I),
    re.compile(r"\b(?:configuration|code) (?:reviewed|inspected)\b", re.I),
    re.compile(r"\b(?:review|check) (?:later|in gtm)\b", re.I),
)


def as_list(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def duplicate_columns(sheet_name: str, rows: list[dict[str, Any]]) -> list[str]:
    if not rows:
        return []
    headers = list(rows[0])
    columns = {
        header: tuple(str(row.get(header, "") or "").strip() for row in rows) for header in headers
    }
    errors = []
    for index, left in enumerate(headers):
        if not any(columns[left]):
            continue
        for right in headers[index + 1 :]:
            if columns[left] == columns[right]:
                errors.append(f"{sheet_name}: duplicate column content in {left!r} and {right!r}")
    return errors


def placeholder_errors(sheet_name: str, rows: list[dict[str, Any]]) -> list[str]:
    errors = []
    for row_number, row in enumerate(rows, start=2):
        for header, value in row.items():
            text = str(value or "")
            if any(pattern.search(text) for pattern in PLACEHOLDER_PATTERNS):
                errors.append(
                    f"{sheet_name}!{header} row {row_number}: generic or deferred wording"
                )
    return errors


def workbook_structure_errors(
    workbook_rows: dict[str, list[dict[str, Any]]]
) -> list[str]:
    errors: list[str] = []
    if list(workbook_rows) != CANONICAL_SHEETS:
        errors.append("workbook tabs do not match the canonical eight-tab order")
    if len(workbook_rows) > 8:
        errors.append("workbook has more than eight tabs")
    for name, rows in workbook_rows.items():
        headers = list(rows[0]) if rows else []
        if len(headers) > 6:
            errors.append(f"{name}: more than six columns")
        errors.extend(duplicate_columns(name, rows))
        errors.extend(placeholder_errors(name, rows))
    if "Change Log" in workbook_rows:
        errors.append("cleanup plan must not contain a change-log tab")
    return errors


def rendered_cell_errors(sheet: Any) -> list[str]:
    errors: list[str] = []
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "[truncated]" in cell.value.lower():
                errors.append(f"{sheet.title}!{cell.coordinate}: proof text was truncated")
            if cell.data_type == "f":
                errors.append(f"{sheet.title}!{cell.coordinate}: formulas are not allowed")
            elif (
                isinstance(cell.value, str)
                and cell.value.lstrip().startswith(("=", "+", "-", "@", "\t", "\r", "\n"))
                and not cell.value.startswith("'")
            ):
                errors.append(
                    f"{sheet.title}!{cell.coordinate}: formula-like text is not escaped"
                )
    return errors


def rendered_sheet_errors(sheet: Any) -> list[str]:
    errors: list[str] = []
    if sheet.max_column > 6:
        errors.append(f"{sheet.title}: rendered workbook exceeds six columns")
    if any((dimension.width or 0) > 92 for dimension in sheet.column_dimensions.values()):
        errors.append(f"{sheet.title}: column width exceeds 92")
    if any((dimension.height or 0) > 120 for dimension in sheet.row_dimensions.values()):
        errors.append(f"{sheet.title}: row height exceeds 120")
    errors.extend(rendered_cell_errors(sheet))
    return errors


def rendered_workbook_errors(workbook_path: Path) -> list[str]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        return [f"openpyxl is required for workbook state validation: {exc}"]
    workbook = load_workbook(workbook_path, read_only=False, data_only=False)
    errors: list[str] = []
    visible = [sheet.title for sheet in workbook if sheet.sheet_state == "visible"]
    if visible != CANONICAL_SHEETS[:2]:
        errors.append("only Summary and Cleanup Plan may be visible")
    for sheet in workbook:
        errors.extend(rendered_sheet_errors(sheet))
    workbook.close()
    return errors


def operations_alignment_errors(
    workbook_rows: dict[str, list[dict[str, Any]]], operations: dict[str, Any]
) -> list[str]:
    errors: list[str] = []
    owner_count = sum(
        1
        for row in as_list(operations.get("decision_ledger"))
        if row.get("disposition") == "owner_decision_needed"
    )
    evidence_limit_count = sum(
        1
        for row in as_list(operations.get("decision_ledger"))
        if row.get("disposition") == "container_evidence_limit"
    )
    cleanup_rows = workbook_rows.get("02 Cleanup Plan", [])
    expected_operation_ids = {
        str(row.get("operation_id") or "")
        for row in as_list(operations.get("operations"))
        if row.get("operation_id")
    }
    rendered_operation_ids = [
        value
        for row in cleanup_rows
        for value in re.findall(r"\bOP-\d{4}\b", str(row.get("id") or ""))
    ]
    if set(rendered_operation_ids) != expected_operation_ids:
        missing = sorted(expected_operation_ids - set(rendered_operation_ids))
        unknown = sorted(set(rendered_operation_ids) - expected_operation_ids)
        if missing:
            errors.append(
                "Cleanup Plan omits operation IDs: " + ", ".join(missing)
            )
        if unknown:
            errors.append(
                "Cleanup Plan contains unknown operation IDs: " + ", ".join(unknown)
            )
    duplicates = sorted(
        operation_id
        for operation_id in set(rendered_operation_ids)
        if rendered_operation_ids.count(operation_id) > 1
    )
    if duplicates:
        errors.append(
            "Cleanup Plan repeats operation IDs across visible rows: "
            + ", ".join(duplicates)
        )

    expected_owner_ids = {
        str(row.get("decision_id") or "")
        for row in as_list(operations.get("decision_ledger"))
        if row.get("disposition") == "owner_decision_needed"
        and row.get("decision_id")
    }
    rendered_ids = [str(row.get("id") or "") for row in cleanup_rows]
    missing_owner_ids = sorted(
        owner_id for owner_id in expected_owner_ids if rendered_ids.count(owner_id) != 1
    )
    if missing_owner_ids:
        errors.append(
            "Cleanup Plan must show each owner decision exactly once: "
            + ", ".join(missing_owner_ids)
        )
    if evidence_limit_count and rendered_ids.count("SCOPE-001") != 1:
        errors.append("Cleanup Plan must contain one consolidated evidence-boundary row")
    if not evidence_limit_count and "SCOPE-001" in rendered_ids:
        errors.append("Cleanup Plan contains an evidence-boundary row without source decisions")
    allowed_standalone_ids = expected_owner_ids | ({"SCOPE-001"} if evidence_limit_count else set())
    for identifier in rendered_ids:
        if re.search(r"\bOP-\d{4}\b", identifier) or identifier in allowed_standalone_ids:
            continue
        errors.append(f"Cleanup Plan contains an unlinked visible row ID: {identifier!r}")
    summary_values = {
        str(row.get("decision") or ""): str(row.get("value") or "")
        for row in workbook_rows.get("01 Summary", [])
    }
    overall_status = summary_values.get("Overall status", "").lower()
    if owner_count and "owner decisions required" not in overall_status:
        errors.append("Summary status does not expose unresolved owner decisions")
    action_completeness = operations.get("action_completeness") or {}
    if action_completeness.get("status") != "pass":
        errors.append("cleanup plan action completeness is not pass")
        if "incomplete cleanup plan" not in overall_status:
            errors.append("Summary status does not expose incomplete cleanup actions")
    if set((operations.get("run_statuses") or {}).values()) != {"complete"}:
        errors.append("operations do not record three complete review runs")
    return errors


def validate_workbook(
    workbook_path: Path, operations_path: Path | None = None
) -> tuple[list[str], list[str]]:
    errors: list[str] = []
    warnings: list[str] = []
    if not workbook_path.is_file():
        return [f"workbook does not exist: {workbook_path}"], warnings
    workbook_rows = load_xlsx_workbook(workbook_path)
    errors.extend(workbook_structure_errors(workbook_rows))
    errors.extend(rendered_workbook_errors(workbook_path))

    if operations_path:
        operations = json.loads(operations_path.read_text(encoding="utf-8"))
        errors.extend(operations_alignment_errors(workbook_rows, operations))
    return errors, warnings


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--operations", type=Path)
    parser.add_argument("--pretty", action="store_true")
    args = parser.parse_args()
    errors, warnings = validate_workbook(args.workbook, args.operations)
    report = {
        "kind": "gtm_cleanup_workbook_gate",
        "status": "pass" if not errors else "fail",
        "workbook": str(args.workbook),
        "errors": errors,
        "warnings": warnings,
    }
    print(json.dumps(report, indent=2 if args.pretty else None))
    for warning in warnings:
        print(f"WARNING: {warning}", file=sys.stderr)
    for error in errors:
        print(f"ERROR: {error}", file=sys.stderr)
    return 1 if errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
